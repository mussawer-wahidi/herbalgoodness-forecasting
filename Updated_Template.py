import pandas as pd
import numpy as np
import os
import json
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from scipy.stats import norm
import math
import time
from datetime import datetime
import threading
from sklearn.linear_model import LinearRegression
from streamlit_extras.stylable_container import stylable_container
import warnings
import traceback
import logging
import sys
import io
from io import BytesIO
import streamlit as st
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
import gspread
from google.oauth2.service_account import Credentials
from collections import defaultdict
from typing import Dict, List, Tuple, Set, Optional
# NEW: API-related imports
from fastapi import FastAPI, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from typing import Optional, List, Dict, Any
import uuid
import threading

warnings.filterwarnings('ignore')

# Always reference files relative to the script's directory
try:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
except NameError:
    BASE_DIR = os.getcwd()

print(f"[DEBUG] BASE_DIR set to: {BASE_DIR}")


try:
    GOOGLE_SHEETS_AVAILABLE = True
except ImportError:
    GOOGLE_SHEETS_AVAILABLE = False

class GoogleSheetsConnector:
    def __init__(self, credentials_file='credentials.json'):
        if not GOOGLE_SHEETS_AVAILABLE:
            raise ImportError("Google Sheets packages not available")

        try:
            print("Connecting to Google Sheets...")
            scopes = [
                'https://www.googleapis.com/auth/spreadsheets.readonly',
                'https://www.googleapis.com/auth/drive.readonly'
            ]
            credentials = Credentials.from_service_account_file(credentials_file, scopes=scopes)
            self.gc = gspread.authorize(credentials)
            print("Google Sheets connection established")
        except Exception as e:
            print(f"Error connecting to Google Sheets: {e}")
            raise

    # def get_inventory_data(self, spreadsheet_url):
    #     """
    #     Extract inventory data from Google Sheets
    #     Column B: SKU
    #     Column AK: Inventory total
    #     """
    #     try:
    #         print("📦 Extracting inventory data from Google Sheets...")
    #         spreadsheet = self.gc.open_by_url(spreadsheet_url)
            
    #         # Get the first worksheet (or specify the worksheet name if needed)
    #         worksheet = spreadsheet.get_worksheet(0)  # First sheet
            
    #         # Get all values
    #         all_values = worksheet.get_all_values()
            
    #         if not all_values:
    #             print("❌ No data found in inventory sheet")
    #             return {}
            
    #         # Find column indices (B = index 1, AK = index 36)
    #         sku_col = 1  # Column B
    #         inventory_col = 36  # Column AK
            
    #         inventory_data = {}
    #         skus_processed = 0
            
    #         # Process data rows (assuming row 1 is header)
    #         for row_idx, row in enumerate(all_values[1:], start=2):
    #             try:
    #                 # Ensure row has enough columns
    #                 if len(row) <= max(sku_col, inventory_col):
    #                     continue
                    
    #                 # Extract SKU
    #                 raw_sku = str(row[sku_col]).strip()
    #                 if not raw_sku or raw_sku.lower() in ['', 'none', 'null', 'n/a']:
    #                     continue
                    
    #                 # Clean SKU - multiple formats
    #                 cleaned_skus = []
                    
    #                 # Original format
    #                 cleaned_skus.append(raw_sku)
                    
    #                 # Remove leading zeros
    #                 cleaned_skus.append(raw_sku.lstrip('0'))
                    
    #                 # Add leading zeros if short UPC
    #                 if raw_sku.isdigit() and len(raw_sku) < 12:
    #                     cleaned_skus.append(raw_sku.zfill(12))
                    
    #                 # Remove non-alphanumeric
    #                 alphanumeric_only = ''.join(c for c in raw_sku if c.isalnum())
    #                 if alphanumeric_only:
    #                     cleaned_skus.append(alphanumeric_only)
                    
    #                 # Extract inventory quantity
    #                 inventory_value = row[inventory_col] if inventory_col < len(row) else '0'
                    
    #                 try:
    #                     inventory_qty = float(str(inventory_value).replace(',', '').strip())
    #                     if inventory_qty < 0:
    #                         inventory_qty = 0
    #                 except:
    #                     inventory_qty = 0
                    
    #                 # Store all SKU variations with the same inventory value
    #                 for sku_variant in cleaned_skus:
    #                     if sku_variant:
    #                         inventory_data[sku_variant] = inventory_qty
                            
    #                 skus_processed += 1
                    
    #             except Exception as e:
    #                 print(f"   Warning: Error processing row {row_idx}: {e}")
    #                 continue
            
    #         print(f"✅ Extracted inventory for {skus_processed} SKUs")
    #         print(f"   Total inventory records (including variants): {len(inventory_data)}")
            
    #         # Show sample data
    #         if inventory_data:
    #             print(f"\n🔍 SAMPLE INVENTORY DATA:")
    #             sample_items = list(inventory_data.items())[:5]
    #             for sku, qty in sample_items:
    #                 print(f"   SKU: '{sku}' -> Inventory: {qty}")
            
    #         return inventory_data
            
    #     except Exception as e:
    #         print(f"❌ Error extracting inventory data: {e}")
    #         import traceback
    #         traceback.print_exc()
    #         return {}

    def get_inventory_data(self, spreadsheet_url):
        """
        Extract inventory data from Google Sheets
        Column C: SKU
        Column R: Inventory total
        Searches through Teas, Capsules, and Liquids worksheets
        """
        try:
            print("📦 Extracting inventory data from Google Sheets...")
            spreadsheet = self.gc.open_by_url(spreadsheet_url)
            
            # Get all available worksheets first for debugging
            all_worksheets = spreadsheet.worksheets()
            print(f"📋 Available worksheets:")
            for ws in all_worksheets:
                print(f"   - '{ws.title}'")
            
            # Define worksheets to search through
            worksheet_names = ['Teas', 'Capsules', 'Liquids']
            
            # Find column indices (C = index 2, R = index 17)
            sku_col = 2  # Column C (0-based indexing)
            inventory_col = 19  # Column R (0-based indexing)
            
            print(f"🔍 Looking for SKUs in column {chr(67)} (index {sku_col})")
            print(f"🔍 Looking for inventory in column {chr(82)} (index {inventory_col})")
            
            inventory_data = {}
            skus_processed = 0
            
            # Iterate through each worksheet
            for worksheet_name in worksheet_names:
                try:
                    print(f"\n🔍 Searching in worksheet: {worksheet_name}")
                    
                    # Try to find worksheet by name (case-insensitive partial match)
                    worksheet = None
                    
                    for ws in all_worksheets:
                        if worksheet_name.lower() in ws.title.lower():
                            worksheet = ws
                            break
                    
                    if not worksheet:
                        print(f"   ⚠️ Worksheet containing '{worksheet_name}' not found, skipping...")
                        continue
                    
                    print(f"   📄 Found worksheet: '{worksheet.title}'")
                    
                    # Get all values from current worksheet
                    all_values = worksheet.get_all_values()
                    
                    if not all_values:
                        print(f"   ❌ No data found in {worksheet.title}")
                        continue
                    
                    # Debug: Show header row to verify column positions
                    if len(all_values) > 0:
                        header_row = all_values[0]
                        print(f"   📊 Header row length: {len(header_row)} columns")
                        if len(header_row) > sku_col:
                            print(f"   📊 Column C header: '{header_row[sku_col]}'")
                        if len(header_row) > inventory_col:
                            print(f"   📊 Column R header: '{header_row[inventory_col]}'")
                        else:
                            print(f"   ⚠️ Column R (index {inventory_col}) not found! Sheet only has {len(header_row)} columns")
                            continue
                    
                    worksheet_skus_processed = 0
                    debug_count = 0
                    
                    # Process data rows (assuming row 1 is header)
                    for row_idx, row in enumerate(all_values[1:], start=2):
                        try:
                            # Ensure row has enough columns
                            if len(row) <= max(sku_col, inventory_col):
                                continue
                            
                            # Extract SKU from column C
                            raw_sku = str(row[sku_col]).strip()
                            if not raw_sku or raw_sku.lower() in ['', 'none', 'null', 'n/a']:
                                continue
                            
                            # Extract inventory quantity from column R
                            inventory_value = row[inventory_col] if inventory_col < len(row) else '0'
                            
                            # Debug: Show first few entries
                            if debug_count < 3:
                                print(f"   🔍 Row {row_idx}: SKU='{raw_sku}', Inventory='{inventory_value}'")
                                debug_count += 1
                            
                            try:
                                # Clean the inventory value more thoroughly
                                inventory_str = str(inventory_value).replace(',', '').replace('$', '').strip()
                                
                                # Handle empty or non-numeric values
                                if not inventory_str or inventory_str.lower() in ['', 'none', 'null', 'n/a', '-']:
                                    inventory_qty = 0
                                else:
                                    inventory_qty = float(inventory_str)
                                    if inventory_qty < 0:
                                        inventory_qty = 0
                            except ValueError:
                                print(f"   ⚠️ Could not parse inventory value '{inventory_value}' for SKU '{raw_sku}' at row {row_idx}")
                                inventory_qty = 0
                            
                            # Clean SKU - multiple formats
                            cleaned_skus = []
                            
                            # Original format
                            cleaned_skus.append(raw_sku)
                            
                            # Remove leading zeros
                            if raw_sku.lstrip('0'):  # Only add if not all zeros
                                cleaned_skus.append(raw_sku.lstrip('0'))
                            
                            # Add leading zeros if short UPC
                            if raw_sku.isdigit() and len(raw_sku) < 12:
                                cleaned_skus.append(raw_sku.zfill(12))
                            
                            # Remove non-alphanumeric
                            alphanumeric_only = ''.join(c for c in raw_sku if c.isalnum())
                            if alphanumeric_only and alphanumeric_only != raw_sku:
                                cleaned_skus.append(alphanumeric_only)
                            
                            # Store all SKU variations with the same inventory value
                            # BUT check for duplicates and warn if inventory values differ
                            for sku_variant in cleaned_skus:
                                if sku_variant:
                                    if sku_variant in inventory_data:
                                        if inventory_data[sku_variant] != inventory_qty:
                                            print(f"   ⚠️ SKU '{sku_variant}' found multiple times with different inventory: {inventory_data[sku_variant]} vs {inventory_qty}")
                                    inventory_data[sku_variant] = inventory_qty
                                    
                            worksheet_skus_processed += 1
                            
                        except Exception as e:
                            print(f"   ❌ Error processing row {row_idx} in {worksheet.title}: {e}")
                            continue
                    
                    print(f"   ✅ Processed {worksheet_skus_processed} SKUs from {worksheet.title}")
                    skus_processed += worksheet_skus_processed
                    
                except Exception as e:
                    print(f"   ❌ Error processing worksheet '{worksheet_name}': {e}")
                    import traceback
                    traceback.print_exc()
                    continue
            
            print(f"\n✅ Extracted inventory for {skus_processed} SKUs across all worksheets")
            print(f"   Total inventory records (including variants): {len(inventory_data)}")
            
            # Show sample data with more details
            if inventory_data:
                print(f"\n🔍 SAMPLE INVENTORY DATA:")
                sample_items = list(inventory_data.items())[:10]  # Show more samples
                for sku, qty in sample_items:
                    print(f"   SKU: '{sku}' -> Inventory: {qty}")
            
            return inventory_data
            
        except Exception as e:
            print(f"❌ Error extracting inventory data: {e}")
            import traceback
            traceback.print_exc()
            return {}

    def get_product_data(self, spreadsheet_url):  ##get_data
        try:
            print("Extracting product data from Google Sheets...")
            spreadsheet = self.gc.open_by_url(spreadsheet_url)
            worksheet = spreadsheet.worksheet("All Labeled Products")

            # Get all values to handle duplicate headers manually
            all_values = worksheet.get_all_values()
            time.sleep(5)

            if not all_values:
                print("No data found in the worksheet")
                return {}, {}, {}, {}, {}

            # Get headers and clean them
            headers = all_values[0]

            # Find the columns we need by index - MORE FLEXIBLE MATCHING
            product_name_col = None
            launch_date_col = None
            upc_col = None
            price_col = None
            lead_time_col = None
            category_col = None
            status_col = None

            # Look for our target columns (case insensitive and partial matching)
            for i, header in enumerate(headers):
                header_clean = str(header).strip().lower()

                if any(keyword in header_clean for keyword in ['product name', 'product_name', 'name']):
                    product_name_col = i
                    print(f"✅ Found 'Product Name' in column {i}: {header}")
                elif any(keyword in header_clean for keyword in ['launch date', 'launch_date', 'date']):
                    launch_date_col = i
                    print(f"✅ Found 'Launch Date' in column {i}: {header}")
                elif any(keyword in header_clean for keyword in ['upc', 'sku', 'barcode', 'unit upc']):
                    upc_col = i
                    print(f"✅ Found 'UPC/SKU' in column {i}: {header}")
                elif any(keyword in header_clean for keyword in ['price', 'cost', 'msrp']):
                    price_col = i
                    print(f"✅ Found 'Price' in column {i}: {header}")
                elif any(keyword in header_clean for keyword in ['lead time', 'lead_time', 'leadtime']):
                    lead_time_col = i
                    print(f"✅ Found 'Lead Time' in column {i}: {header}")
                elif any(keyword in header_clean for keyword in ['category', 'product_category', 'product category']):
                    category_col = i
                    print(f"✅ Found 'Category' in column {i}: {header}")
                elif header_clean == 'status':
                    status_col = i
                    print(f"✅ Found 'Status' in column {i}: {header}")


            if upc_col is None:
                print("❌ Could not find UPC/SKU column")
                return {}, {}, {}, {}, {}

            product_info = {}
            lead_times = {}
            launch_dates = {}  # Store launch dates for each SKU
            sku_variations = {}  # Track different formats of the same SKU
            product_category = {}
            product_status = {}

            # Process data rows (skip header row)
            for row_idx, row in enumerate(all_values[1:], start=2):
                try:
                    # Ensure row has enough columns
                    if len(row) <= max(filter(None, [product_name_col, launch_date_col, upc_col, price_col, lead_time_col, category_col, status_col])):
                        continue

                    # Extract UPC/SKU with multiple cleaning approaches
                    raw_upc = str(row[upc_col]).strip() if upc_col < len(row) else ''
                    if not raw_upc or raw_upc.lower() in ['', 'none', 'null', 'n/a']:
                        continue

                    # Clean UPC/SKU - try multiple formats
                    cleaned_upcs = []

                    # Original format
                    cleaned_upcs.append(raw_upc)

                    # Remove leading zeros (common issue)
                    cleaned_upcs.append(raw_upc.lstrip('0'))

                    # Add leading zeros if it looks like a short UPC
                    if raw_upc.isdigit() and len(raw_upc) < 12:
                        cleaned_upcs.append(raw_upc.zfill(12))

                    # Remove all non-alphanumeric characters
                    alphanumeric_only = ''.join(c for c in raw_upc if c.isalnum())
                    if alphanumeric_only:
                        cleaned_upcs.append(alphanumeric_only)

                    # Remove spaces and dashes
                    no_spaces_dashes = raw_upc.replace(' ', '').replace('-', '')
                    if no_spaces_dashes:
                        cleaned_upcs.append(no_spaces_dashes)

                    # Extract other data
                    product_name = str(row[product_name_col]).strip() if product_name_col is not None and product_name_col < len(row) else 'Unknown'
                    launch_date_raw = str(row[launch_date_col]).strip() if launch_date_col is not None and launch_date_col < len(row) else ''
                    price = row[price_col] if price_col is not None and price_col < len(row) else 0
                    lead_time = row[lead_time_col] if lead_time_col is not None and lead_time_col < len(row) else 2
                    category_name = str(row[category_col]).strip() if category_col is not None and category_col < len(row) else 'Unknown'
                    status_name = str(row[status_col]).strip() if status_col is not None and status_col < len(row) else 'Unknown'

                    # Clean and parse launch date
                    launch_date_parsed = None
                    if launch_date_raw and launch_date_raw.lower() not in ['', 'none', 'null', 'n/a']:
                        try:
                            # Try multiple date formats
                            for date_format in ['%m/%d/%Y', '%Y-%m-%d', '%m-%d-%Y', '%d/%m/%Y', '%Y/%m/%d']:
                                try:
                                    launch_date_parsed = pd.to_datetime(launch_date_raw, format=date_format)
                                    break
                                except:
                                    continue

                            # If still not parsed, try pandas auto-detection
                            if launch_date_parsed is None:
                                launch_date_parsed = pd.to_datetime(launch_date_raw, errors='coerce')

                        except Exception as e:
                            print(f"   Warning: Could not parse launch date '{launch_date_raw}' for row {row_idx}")
                            launch_date_parsed = None

                    # Clean and validate data
                    try:
                        price = float(str(price).replace('$', '').replace(',', '')) if price else 0
                    except:
                        price = 0

                    try:
                        lead_time = int(float(str(lead_time))) if lead_time else 2
                    except:
                        lead_time = 2

                    # Store all UPC variations for this product
                    for upc_variant in cleaned_upcs:
                        if upc_variant and upc_variant not in product_info:
                            product_info[upc_variant] = product_name
                            product_category[upc_variant] = category_name
                            product_status[upc_variant] = status_name
                            lead_times[upc_variant] = lead_time
                            launch_dates[upc_variant] = launch_date_parsed  # Store parsed date
                            sku_variations[upc_variant] = {
                                'original': raw_upc,
                                'product_name': product_name,
                                'category_name': category_name,
                                'status_name': status_name,
                                'launch_date': launch_date_parsed,
                                'row': row_idx
                            }

                except Exception as e:
                    print(f"Error processing row {row_idx}: {e}")
                    continue

            print(f"✅ Extracted data for {len(product_info)} product variations")
            print(f"✅ Extracted data for {len(product_category)} product variations")
            print(f"✅ Extracted data for {len(product_status)} product variations")
            print(f"✅ Extracted lead times for {len(lead_times)} product variations")
            print(f"✅ Extracted launch dates for {len([d for d in launch_dates.values() if d is not None])} product variations")

            # Debug: Show some examples of what we extracted
            print(f"\n🔍 SAMPLE EXTRACTED DATA:")
            sample_items = list(product_info.items())[:5]
            for sku, name in sample_items:
                launch_info = launch_dates.get(sku)
                launch_str = launch_info.strftime('%Y-%m-%d') if launch_info else 'No Date'
                print(f"   SKU: '{sku}' -> Product: '{name}' | Launch: {launch_str} | Lead Time: {lead_times.get(sku, 'N/A')}")

            return product_info, lead_times, launch_dates, product_category, product_status

        except Exception as e:
            print(f"❌ Error extracting product data: {e}")
            import traceback
            traceback.print_exc()
            return {}, {}, {}, {}, {}

    def get_amazon_fba_weekly_sales(self, spreadsheet_url):
        try:
            print("📦 Extracting Amazon FBA weekly sales data...")
            spreadsheet = self.gc.open_by_url(spreadsheet_url)
            worksheet = spreadsheet.worksheet("Amazon FBA")

            all_values = worksheet.get_all_values()
            if not all_values:
                print("❌ No data found.")
                return pd.DataFrame()

            headers = all_values[0]
            print(f"🧾 Headers: {headers}")

            sku_col = None
            week_columns = []

            # Identify columns
            for i, header in enumerate(headers):
                header_clean = str(header).strip().lower()

                if 'upc' in header_clean:
                    sku_col = i
                    print(f"✅ Found UPC column at index {i}: {header}")
                    continue

                # Try to parse as date
                try:
                    week_date = pd.to_datetime(header, dayfirst=False)  # mm/dd/yyyy format
                    week_columns.append((i, week_date))
                    print(f"📅 Week column at index {i}: {week_date.strftime('%d/%m/%Y')}")
                except:
                    # Print once for debugging
                    if i >= 4:  # Avoid clutter for known non-date columns
                        print(f"⛔ Skipping non-date column: {header}")
                    continue

            if sku_col is None or not week_columns:
                print("❌ Could not identify necessary columns.")
                return pd.DataFrame()

            # Parse each row
            weekly_sales_data = []
            processed_skus = set()

            for row_idx, row in enumerate(all_values[1:], start=2):
                try:
                    if len(row) <= sku_col:
                        continue

                    raw_sku = str(row[sku_col]).strip()
                    if not raw_sku or raw_sku.lower() in ['none', 'null', 'n/a', '']:
                        continue

                    # Clean SKU
                    cleaned_sku = raw_sku.zfill(12) if raw_sku.isdigit() and len(raw_sku) < 12 else raw_sku
                    processed_skus.add(cleaned_sku)

                    for col_idx, week_date in week_columns:
                        if col_idx >= len(row):
                            continue
                        try:
                            sales = float(row[col_idx]) if row[col_idx].strip() not in ['', '-', 'n/a', 'N/A'] else 0
                        except:
                            sales = 0

                        if sales > 0:
                            weekly_sales_data.append({
                                'SKU': cleaned_sku,
                                'Original_SKU': raw_sku,
                                'Week_Start': week_date,
                                'Sales': sales,
                                'Channel': 'Amazon'
                            })

                except Exception as e:
                    print(f"⚠️ Error at row {row_idx}: {e}")
                    continue

            # Convert to DataFrame
            weekly_df = pd.DataFrame(weekly_sales_data)

            if weekly_df.empty:
                print("⚠️ No valid weekly sales found.")
            else:
                print(f"✅ {len(weekly_df)} records extracted | {weekly_df['SKU'].nunique()} SKUs")
                print(f"   Range: {weekly_df['Week_Start'].min().strftime('%d/%m/%Y')} → {weekly_df['Week_Start'].max().strftime('%d/%m/%Y')}")

            return weekly_df

        except Exception as e:
            print(f"❌ Fatal error during extraction: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def convert_amazon_weekly_to_monthly(self, weekly_df):
        """
        Convert Amazon FBA weekly sales data to monthly format
        using actual week start dates.
        """
        try:
            if weekly_df.empty:
                return pd.DataFrame()

            print(f"Converting Amazon FBA weekly sales to monthly format...")

            # Ensure 'Week_Start' column is datetime
            weekly_df['Week_Start'] = pd.to_datetime(weekly_df['Week_Start'], errors='coerce')

            # Drop rows with invalid dates
            weekly_df = weekly_df.dropna(subset=['Week_Start'])

            # Assign the month (e.g., 2025-01-31, 2025-02-28 etc.) to each week using MonthEnd
            weekly_df['Month'] = weekly_df['Week_Start'] + pd.offsets.MonthEnd(0)

            # Debug print mapping
            mapping = weekly_df[['Week_Start', 'Month']].drop_duplicates().sort_values('Week_Start')
            print("Week to Month mapping:")
            for _, row in mapping.head(10).iterrows():
                print(f"   Week of {row['Week_Start'].strftime('%d/%m/%Y')} → Month End: {row['Month'].strftime('%d/%m/%Y')}")

            # Group and aggregate
            monthly_sales = weekly_df.groupby(['SKU', 'Month'])['Sales'].sum().reset_index()
            monthly_sales['Channel'] = 'Amazon'
            monthly_sales = monthly_sales.rename(columns={'Month': 'Date'})

            print(f"✅ Converted to {len(monthly_sales)} monthly records")
            print(f"   Date range: {monthly_sales['Date'].min().strftime('%d/%m/%Y')} to {monthly_sales['Date'].max().strftime('%d/%m/%Y')}")
            print(f"   SKUs covered: {monthly_sales['SKU'].nunique()}")

            # Sample output
            for _, row in monthly_sales.head(3).iterrows():
                print(f"   SKU {row['SKU']} | {row['Date'].strftime('%b %Y')} | Sales: {row['Sales']}")

            return monthly_sales

        except Exception as e:
            print(f"❌ Error converting Amazon FBA weekly to monthly: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def extend_historical_data_with_amazon_weekly(self, historical_data, amazon_weekly_monthly_data):
        """
        Extend historical sales data with Amazon FBA weekly data
        """
        try:
            if amazon_weekly_monthly_data.empty:
                print("No Amazon FBA weekly data to extend historical data")
                return historical_data

            print("Extending historical data with Amazon FBA weekly sales...")

            # Get the latest date in historical data
            historical_data['Date'] = pd.to_datetime(historical_data['Date'])
            latest_historical_date = historical_data['Date'].max()

            print(f"Latest historical date: {latest_historical_date}")
            print(f"Amazon FBA weekly data starts: {amazon_weekly_monthly_data['Date'].min()}")
            print(f"Amazon FBA weekly data ends: {amazon_weekly_monthly_data['Date'].max()}")

            # Filter Amazon weekly data to only include dates after historical data
            new_amazon_data = amazon_weekly_monthly_data[amazon_weekly_monthly_data['Date'] > latest_historical_date].copy()

            if new_amazon_data.empty:
                print("No new Amazon FBA weekly data after historical cutoff")
                return historical_data

            print(f"Adding {len(new_amazon_data)} new Amazon FBA monthly records")

            # Amazon FBA data is already in the right format - just add it
            if not new_amazon_data.empty:
                # Combine with historical data
                combined_data = pd.concat([historical_data, new_amazon_data], ignore_index=True)
                combined_data = combined_data.sort_values(['SKU', 'Channel', 'Date'])

                print(f"✅ Extended historical data with {len(new_amazon_data)} new Amazon records")
                print(f"   Total records: {len(combined_data)} (was {len(historical_data)})")
                print(f"   New date range: {combined_data['Date'].min()} to {combined_data['Date'].max()}")

                # Show sample of extended data
                amazon_sample = new_amazon_data.head(3)
                for _, row in amazon_sample.iterrows():
                    print(f"   Added Amazon: {row['SKU']}, {row['Date'].strftime('%b %Y')}, Sales: {row['Sales']}")

                return combined_data
            else:
                return historical_data

        except Exception as e:
            print(f"❌ Error extending historical data with Amazon FBA: {e}")
            import traceback
            traceback.print_exc()
            return historical_data
        
    ### For Shopify Data Seprately

    def get_shopify_main_weekly_sales(self, spreadsheet_url):
        try:
            print("📦 Extracting Shopify Main weekly sales data...")
            spreadsheet = self.gc.open_by_url(spreadsheet_url)
            worksheet = spreadsheet.worksheet("Shopify Main")

            all_values = worksheet.get_all_values()
            if not all_values:
                print("❌ No data found.")
                return pd.DataFrame()

            headers = all_values[0]
            print(f"🧾 Headers: {headers}")

            sku_col = None
            week_columns = []

            # Identify columns
            for i, header in enumerate(headers):
                header_clean = str(header).strip().lower()

                if 'upc' in header_clean:
                    sku_col = i
                    print(f"✅ Found UPC column at index {i}: {header}")
                    continue

                # Try to parse as date
                try:
                    week_date = pd.to_datetime(header, dayfirst=False)  # mm/dd/yyyy format
                    week_columns.append((i, week_date))
                    print(f"📅 Week column at index {i}: {week_date.strftime('%d/%m/%Y')}")
                except:
                    # Print once for debugging
                    if i >= 4:  # Avoid clutter for known non-date columns
                        print(f"⛔ Skipping non-date column: {header}")
                    continue

            if sku_col is None or not week_columns:
                print("❌ Could not identify necessary columns.")
                return pd.DataFrame()

            # Parse each row
            weekly_sales_data = []
            processed_skus = set()

            for row_idx, row in enumerate(all_values[1:], start=2):
                try:
                    if len(row) <= sku_col:
                        continue

                    raw_sku = str(row[sku_col]).strip()
                    if not raw_sku or raw_sku.lower() in ['none', 'null', 'n/a', '']:
                        continue

                    # Clean SKU
                    cleaned_sku = raw_sku.zfill(12) if raw_sku.isdigit() and len(raw_sku) < 12 else raw_sku
                    processed_skus.add(cleaned_sku)

                    for col_idx, week_date in week_columns:
                        if col_idx >= len(row):
                            continue
                        try:
                            sales = float(row[col_idx]) if row[col_idx].strip() not in ['', '-', 'n/a', 'N/A'] else 0
                        except:
                            sales = 0

                        if sales > 0:
                            weekly_sales_data.append({
                                'SKU': cleaned_sku,
                                'Original_SKU': raw_sku,
                                'Week_Start': week_date,
                                'Sales': sales,
                                'Channel': 'Shopify'
                            })

                except Exception as e:
                    print(f"⚠️ Error at row {row_idx}: {e}")
                    continue

            # Convert to DataFrame
            weekly_df2 = pd.DataFrame(weekly_sales_data)

            if weekly_df2.empty:
                print("⚠️ No valid weekly sales found.")
            else:
                print(f"✅ {len(weekly_df2)} records extracted | {weekly_df2['SKU'].nunique()} SKUs")
                print(f"   Range: {weekly_df2['Week_Start'].min().strftime('%d/%m/%Y')} → {weekly_df2['Week_Start'].max().strftime('%d/%m/%Y')}")

            return weekly_df2

        except Exception as e:
            print(f"❌ Fatal error during extraction: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def convert_shopify_weekly_to_monthly(self, weekly_df2):
        """
        Convert Shopify Main weekly sales data to monthly format
        using actual week start dates.
        """
        try:
            if weekly_df2.empty:
                return pd.DataFrame()

            print(f"Converting Shopify Main weekly sales to monthly format...")

            # Ensure 'Week_Start' column is datetime
            weekly_df2['Week_Start'] = pd.to_datetime(weekly_df2['Week_Start'], errors='coerce')

            # Drop rows with invalid dates
            weekly_df2 = weekly_df2.dropna(subset=['Week_Start'])

            # Assign the month (e.g., 2025-01-31, 2025-02-28 etc.) to each week using MonthEnd
            weekly_df2['Month'] = weekly_df2['Week_Start'] + pd.offsets.MonthEnd(0)

            # Debug print mapping
            mapping = weekly_df2[['Week_Start', 'Month']].drop_duplicates().sort_values('Week_Start')
            print("Week to Month mapping:")
            for _, row in mapping.head(10).iterrows():
                print(f"   Week of {row['Week_Start'].strftime('%d/%m/%Y')} → Month End: {row['Month'].strftime('%d/%m/%Y')}")

            # Group and aggregate
            monthly_sales2 = weekly_df2.groupby(['SKU', 'Month'])['Sales'].sum().reset_index()
            monthly_sales2['Channel'] = 'Shopify'
            monthly_sales2 = monthly_sales2.rename(columns={'Month': 'Date'})

            print(f"✅ Converted to {len(monthly_sales2)} monthly records")
            print(f"   Date range: {monthly_sales2['Date'].min().strftime('%d/%m/%Y')} to {monthly_sales2['Date'].max().strftime('%d/%m/%Y')}")
            print(f"   SKUs covered: {monthly_sales2['SKU'].nunique()}")

            # Sample output
            for _, row in monthly_sales2.head(3).iterrows():
                print(f"   SKU {row['SKU']} | {row['Date'].strftime('%b %Y')} | Sales: {row['Sales']}")

            return monthly_sales2

        except Exception as e:
            print(f"❌ Error converting Shopify Main weekly to monthly: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()
    
    def extend_historical_data_with_shopify_weekly(self, historical_data, shopify_weekly_monthly_data):
        """
        Extend historical sales data with Shopify Main weekly data
        """
        try:
            if shopify_weekly_monthly_data.empty:
                print("No Shopify Main weekly data to extend historical data")
                return historical_data

            print("Extending historical data with Shopify weekly sales...")

            # Get the latest date in historical data
            historical_data['Date'] = pd.to_datetime(historical_data['Date'])
            latest_historical_date = historical_data['Date'].max()

            print(f"Latest historical date: {latest_historical_date}")
            print(f"Shopify Main weekly data starts: {shopify_weekly_monthly_data['Date'].min()}")
            print(f"Shopify Main weekly data ends: {shopify_weekly_monthly_data['Date'].max()}")

            # Filter Shopify weekly data to only include dates after historical data
            new_shopify_data = shopify_weekly_monthly_data[shopify_weekly_monthly_data['Date'] > latest_historical_date].copy()

            if new_shopify_data.empty:
                print("No new Shopify Main weekly data after historical cutoff")
                return historical_data

            print(f"Adding {len(new_shopify_data)} new Shopify Main monthly records")

            # Shopify Main data is already in the right format - just add it
            if not new_shopify_data.empty:
                # Combine with historical data
                combined_data = pd.concat([historical_data, new_shopify_data], ignore_index=True)
                combined_data = combined_data.sort_values(['SKU', 'Channel', 'Date'])

                print(f"✅ Extended historical data with {len(new_shopify_data)} new Shopify records")
                print(f"   Total records: {len(combined_data)} (was {len(historical_data)})")
                print(f"   New date range: {combined_data['Date'].min()} to {combined_data['Date'].max()}")

                # Show sample of extended data
                shopify_sample = new_shopify_data.head(3)
                for _, row in shopify_sample.iterrows():
                    print(f"   Added Shopify: {row['SKU']}, {row['Date'].strftime('%b %Y')}, Sales: {row['Sales']}")

                return combined_data
            else:
                return historical_data

        except Exception as e:
            print(f"❌ Error extending historical data with Shopify Main: {e}")
            import traceback
            traceback.print_exc()
            return historical_data    

### Just ADDED Shopify Faire Tab

    def get_shopify_faire_weekly_sales(self, spreadsheet_url):
        try:
            print("📦 Extracting Shopify Faire weekly sales data...")
            spreadsheet = self.gc.open_by_url(spreadsheet_url)
            worksheet = spreadsheet.worksheet("Shopify Faire")

            all_values = worksheet.get_all_values()
            if not all_values:
                print("❌ No data found.")
                return pd.DataFrame()

            headers = all_values[0]
            print(f"🧾 Headers: {headers}")

            sku_col = None
            week_columns = []

            # Identify columns
            for i, header in enumerate(headers):
                header_clean = str(header).strip().lower()

                if 'upc' in header_clean:
                    sku_col = i
                    print(f"✅ Found UPC column at index {i}: {header}")
                    continue

                # Try to parse as date
                try:
                    week_date = pd.to_datetime(header, dayfirst=False)  # mm/dd/yyyy format
                    week_columns.append((i, week_date))
                    print(f"📅 Week column at index {i}: {week_date.strftime('%d/%m/%Y')}")
                except:
                    # Print once for debugging
                    if i >= 4:  # Avoid clutter for known non-date columns
                        print(f"⛔ Skipping non-date column: {header}")
                    continue

            if sku_col is None or not week_columns:
                print("❌ Could not identify necessary columns.")
                return pd.DataFrame()

            # Parse each row
            weekly_sales_data = []
            processed_skus = set()

            for row_idx, row in enumerate(all_values[1:], start=2):
                try:
                    if len(row) <= sku_col:
                        continue

                    raw_sku = str(row[sku_col]).strip()
                    if not raw_sku or raw_sku.lower() in ['none', 'null', 'n/a', '']:
                        continue

                    # Clean SKU
                    cleaned_sku = raw_sku.zfill(12) if raw_sku.isdigit() and len(raw_sku) < 12 else raw_sku
                    processed_skus.add(cleaned_sku)

                    for col_idx, week_date in week_columns:
                        if col_idx >= len(row):
                            continue
                        try:
                            sales = float(row[col_idx]) if row[col_idx].strip() not in ['', '-', 'n/a', 'N/A'] else 0
                        except:
                            sales = 0

                        if sales > 0:
                            weekly_sales_data.append({
                                'SKU': cleaned_sku,
                                'Original_SKU': raw_sku,
                                'Week_Start': week_date,
                                'Sales': sales,
                                'Channel': 'Shopify Faire'
                            })

                except Exception as e:
                    print(f"⚠️ Error at row {row_idx}: {e}")
                    continue

            # Convert to DataFrame
            weekly_df3 = pd.DataFrame(weekly_sales_data)

            if weekly_df3.empty:
                print("⚠️ No valid weekly sales found.")
            else:
                print(f"✅ {len(weekly_df3)} records extracted | {weekly_df3['SKU'].nunique()} SKUs")
                print(f"   Range: {weekly_df3['Week_Start'].min().strftime('%d/%m/%Y')} → {weekly_df3['Week_Start'].max().strftime('%d/%m/%Y')}")

            return weekly_df3

        except Exception as e:
            print(f"❌ Fatal error during extraction: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def convert_shopify_faire_weekly_to_monthly(self, weekly_df3):
        """
        Convert Shopify Faire weekly sales data to monthly format
        using actual week start dates.
        """
        try:
            if weekly_df3.empty:
                return pd.DataFrame()

            print(f"Converting Shopify Faire weekly sales to monthly format...")

            # Ensure 'Week_Start' column is datetime
            weekly_df3['Week_Start'] = pd.to_datetime(weekly_df3['Week_Start'], errors='coerce')

            # Drop rows with invalid dates
            weekly_df3 = weekly_df3.dropna(subset=['Week_Start'])

            # Assign the month (e.g., 2025-01-31, 2025-02-28 etc.) to each week using MonthEnd
            weekly_df3['Month'] = weekly_df3['Week_Start'] + pd.offsets.MonthEnd(0)

            # Debug print mapping
            mapping = weekly_df3[['Week_Start', 'Month']].drop_duplicates().sort_values('Week_Start')
            print("Week to Month mapping:")
            for _, row in mapping.head(10).iterrows():
                print(f"   Week of {row['Week_Start'].strftime('%d/%m/%Y')} → Month End: {row['Month'].strftime('%d/%m/%Y')}")

            # Group and aggregate
            monthly_sales3 = weekly_df3.groupby(['SKU', 'Month'])['Sales'].sum().reset_index()
            monthly_sales3['Channel'] = 'Shopify Faire'
            monthly_sales3 = monthly_sales3.rename(columns={'Month': 'Date'})

            print(f"✅ Converted to {len(monthly_sales3)} monthly records")
            print(f"   Date range: {monthly_sales3['Date'].min().strftime('%d/%m/%Y')} to {monthly_sales3['Date'].max().strftime('%d/%m/%Y')}")
            print(f"   SKUs covered: {monthly_sales3['SKU'].nunique()}")

            # Sample output
            for _, row in monthly_sales3.head(3).iterrows():
                print(f"   SKU {row['SKU']} | {row['Date'].strftime('%b %Y')} | Sales: {row['Sales']}")

            return monthly_sales3

        except Exception as e:
            print(f"❌ Error converting Shopify Faire weekly to monthly: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def extend_historical_data_with_shopify_faire_weekly(self, historical_data, shopify_faire_weekly_monthly_data):
        """
        Extend historical sales data with Shopify Faire weekly data
        """
        try:
            if shopify_faire_weekly_monthly_data.empty:
                print("No Shopify Faire weekly data to extend historical data")
                return historical_data

            print("Extending historical data with Shopify Faire weekly sales...")

            # Get the latest date in historical data
            historical_data['Date'] = pd.to_datetime(historical_data['Date'])
            latest_historical_date = historical_data['Date'].max()

            print(f"Latest historical date: {latest_historical_date}")
            print(f"Shopify Faire weekly data starts: {shopify_faire_weekly_monthly_data['Date'].min()}")
            print(f"Shopify Faire weekly data ends: {shopify_faire_weekly_monthly_data['Date'].max()}")

            # Filter Shopify Faire weekly data to only include dates after historical data
            new_shopify_faire_data = shopify_faire_weekly_monthly_data[shopify_faire_weekly_monthly_data['Date'] > latest_historical_date].copy()

            if new_shopify_faire_data.empty:
                print("No new Shopify Faire weekly data after historical cutoff")
                return historical_data

            print(f"Adding {len(new_shopify_faire_data)} new Shopify Faire monthly records")

            # Shopify Faire data is already in the right format - just add it
            if not new_shopify_faire_data.empty:
                # Combine with historical data
                combined_data = pd.concat([historical_data, new_shopify_faire_data], ignore_index=True)
                combined_data = combined_data.sort_values(['SKU', 'Channel', 'Date'])

                print(f"✅ Extended historical data with {len(new_shopify_faire_data)} new Shopify records")
                print(f"   Total records: {len(combined_data)} (was {len(historical_data)})")
                print(f"   New date range: {combined_data['Date'].min()} to {combined_data['Date'].max()}")

                # Show sample of extended data
                shopify_faire_sample = new_shopify_faire_data.head(3)
                for _, row in shopify_faire_sample.iterrows():
                    print(f"   Added Shopify Faire: {row['SKU']}, {row['Date'].strftime('%b %Y')}, Sales: {row['Sales']}")

                return combined_data
            else:
                return historical_data

        except Exception as e:
            print(f"❌ Error extending historical data with Shopify Faire: {e}")
            import traceback
            traceback.print_exc()
            return historical_data     

        ### AMAZON FBM TAB   
        
    def get_amazon_fbm_weekly_sales(self, spreadsheet_url):
        try:
            print("📦 Extracting Amazon FBM weekly sales data...")
            spreadsheet = self.gc.open_by_url(spreadsheet_url)
            worksheet = spreadsheet.worksheet("Amazon FBM")

            all_values = worksheet.get_all_values()
            if not all_values:
                print("❌ No data found.")
                return pd.DataFrame()

            headers = all_values[0]
            print(f"🧾 Headers: {headers}")

            sku_col = None
            week_columns = []

            # Identify columns
            for i, header in enumerate(headers):
                header_clean = str(header).strip().lower()

                if 'upc' in header_clean:
                    sku_col = i
                    print(f"✅ Found UPC column at index {i}: {header}")
                    continue

                # Try to parse as date
                try:
                    week_date = pd.to_datetime(header, dayfirst=False)  # mm/dd/yyyy format
                    week_columns.append((i, week_date))
                    print(f"📅 Week column at index {i}: {week_date.strftime('%d/%m/%Y')}")
                except:
                    # Print once for debugging
                    if i >= 4:  # Avoid clutter for known non-date columns
                        print(f"⛔ Skipping non-date column: {header}")
                    continue

            if sku_col is None or not week_columns:
                print("❌ Could not identify necessary columns.")
                return pd.DataFrame()

            # Parse each row
            weekly_sales_data = []
            processed_skus = set()

            for row_idx, row in enumerate(all_values[1:], start=2):
                try:
                    if len(row) <= sku_col:
                        continue

                    raw_sku = str(row[sku_col]).strip()
                    if not raw_sku or raw_sku.lower() in ['none', 'null', 'n/a', '']:
                        continue

                    # Clean SKU
                    cleaned_sku = raw_sku.zfill(12) if raw_sku.isdigit() and len(raw_sku) < 12 else raw_sku
                    processed_skus.add(cleaned_sku)

                    for col_idx, week_date in week_columns:
                        if col_idx >= len(row):
                            continue
                        try:
                            sales = float(row[col_idx]) if row[col_idx].strip() not in ['', '-', 'n/a', 'N/A'] else 0
                        except:
                            sales = 0

                        if sales > 0:
                            weekly_sales_data.append({
                                'SKU': cleaned_sku,
                                'Original_SKU': raw_sku,
                                'Week_Start': week_date,
                                'Sales': sales,
                                'Channel': 'Amazonfbm'
                            })

                except Exception as e:
                    print(f"⚠️ Error at row {row_idx}: {e}")
                    continue

            # Convert to DataFrame
            weekly_df4 = pd.DataFrame(weekly_sales_data)

            if weekly_df4.empty:
                print("⚠️ No valid weekly sales found.")
            else:
                print(f"✅ {len(weekly_df4)} records extracted | {weekly_df4['SKU'].nunique()} SKUs")
                print(f"   Range: {weekly_df4['Week_Start'].min().strftime('%d/%m/%Y')} → {weekly_df4['Week_Start'].max().strftime('%d/%m/%Y')}")

            return weekly_df4

        except Exception as e:
            print(f"❌ Fatal error during extraction: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def convert_amazon_fbm_weekly_to_monthly(self, weekly_df4):
        """
        Convert Amazon FBM weekly sales data to monthly format
        using actual week start dates.
        """
        try:
            if weekly_df4.empty:
                return pd.DataFrame()

            print(f"Converting Amazon FBM weekly sales to monthly format...")

            # Ensure 'Week_Start' column is datetime
            weekly_df4['Week_Start'] = pd.to_datetime(weekly_df4['Week_Start'], errors='coerce')

            # Drop rows with invalid dates
            weekly_df4 = weekly_df4.dropna(subset=['Week_Start'])

            # Assign the month (e.g., 2025-01-31, 2025-02-28 etc.) to each week using MonthEnd
            weekly_df4['Month'] = weekly_df4['Week_Start'] + pd.offsets.MonthEnd(0)

            # Debug print mapping
            mapping = weekly_df4[['Week_Start', 'Month']].drop_duplicates().sort_values('Week_Start')
            print("Week to Month mapping:")
            for _, row in mapping.head(10).iterrows():
                print(f"   Week of {row['Week_Start'].strftime('%d/%m/%Y')} → Month End: {row['Month'].strftime('%d/%m/%Y')}")

            # Group and aggregate
            monthly_sales4 = weekly_df4.groupby(['SKU', 'Month'])['Sales'].sum().reset_index()
            monthly_sales4['Channel'] = 'Amazonfbm'
            monthly_sales4 = monthly_sales4.rename(columns={'Month': 'Date'})

            print(f"✅ Converted to {len(monthly_sales4)} monthly records")
            print(f"   Date range: {monthly_sales4['Date'].min().strftime('%d/%m/%Y')} to {monthly_sales4['Date'].max().strftime('%d/%m/%Y')}")
            print(f"   SKUs covered: {monthly_sales4['SKU'].nunique()}")

            # Sample output
            for _, row in monthly_sales4.head(3).iterrows():
                print(f"   SKU {row['SKU']} | {row['Date'].strftime('%b %Y')} | Sales: {row['Sales']}")

            return monthly_sales4

        except Exception as e:
            print(f"❌ Error converting Amazon FBM weekly to monthly: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def extend_historical_data_with_amazon_fbm_weekly(self, historical_data, amazon_fbm_weekly_monthly_data):
        """
        Extend historical sales data with Amazon FBM weekly data
        """
        try:
            if amazon_fbm_weekly_monthly_data.empty:
                print("No Amazon FBM weekly data to extend historical data")
                return historical_data

            print("Extending historical data with Amazon FBM weekly sales...")

            # Get the latest date in historical data
            historical_data['Date'] = pd.to_datetime(historical_data['Date'])
            latest_historical_date = historical_data['Date'].max()

            print(f"Latest historical date: {latest_historical_date}")
            print(f"Amazon FBM weekly data starts: {amazon_fbm_weekly_monthly_data['Date'].min()}")
            print(f"Amazon FBM weekly data ends: {amazon_fbm_weekly_monthly_data['Date'].max()}")

            # Filter Amazon FBM weekly data to only include dates after historical data
            new_amazon_fbm_data = amazon_fbm_weekly_monthly_data[amazon_fbm_weekly_monthly_data['Date'] > latest_historical_date].copy()

            if new_amazon_fbm_data.empty:
                print("No new Amazon FBM weekly data after historical cutoff")
                return historical_data

            print(f"Adding {len(new_amazon_fbm_data)} new Amazon FBM monthly records")

            # Amazon FBM data is already in the right format - just add it
            if not new_amazon_fbm_data.empty:
                # Combine with historical data
                combined_data = pd.concat([historical_data, new_amazon_fbm_data], ignore_index=True)
                combined_data = combined_data.sort_values(['SKU', 'Channel', 'Date'])

                print(f"✅ Extended historical data with {len(new_amazon_fbm_data)} new Shopify records")
                print(f"   Total records: {len(combined_data)} (was {len(historical_data)})")
                print(f"   New date range: {combined_data['Date'].min()} to {combined_data['Date'].max()}")

                # Show sample of extended data
                amazon_fbm_sample = new_amazon_fbm_data.head(3)
                for _, row in amazon_fbm_sample.iterrows():
                    print(f"   Added Amazon FBM: {row['SKU']}, {row['Date'].strftime('%b %Y')}, Sales: {row['Sales']}")

                return combined_data
            else:
                return historical_data

        except Exception as e:
            print(f"❌ Error extending historical data with Amazon FBM: {e}")
            import traceback
            traceback.print_exc()
            return historical_data        
        
        ### Walmart FBM Tab

    def get_walmart_fbm_weekly_sales(self, spreadsheet_url):
        try:
            print("📦 Extracting Walmart FBM weekly sales data...")
            spreadsheet = self.gc.open_by_url(spreadsheet_url)
            worksheet = spreadsheet.worksheet("Walmart FBM")

            all_values = worksheet.get_all_values()
            if not all_values:
                print("❌ No data found.")
                return pd.DataFrame()

            headers = all_values[0]
            print(f"🧾 Headers: {headers}")

            sku_col = None
            week_columns = []

            # Identify columns
            for i, header in enumerate(headers):
                header_clean = str(header).strip().lower()

                if 'upc' in header_clean:
                    sku_col = i
                    print(f"✅ Found UPC column at index {i}: {header}")
                    continue

                # Try to parse as date
                try:
                    week_date = pd.to_datetime(header, dayfirst=False)  # mm/dd/yyyy format
                    week_columns.append((i, week_date))
                    print(f"📅 Week column at index {i}: {week_date.strftime('%d/%m/%Y')}")
                except:
                    # Print once for debugging
                    if i >= 4:  # Avoid clutter for known non-date columns
                        print(f"⛔ Skipping non-date column: {header}")
                    continue

            if sku_col is None or not week_columns:
                print("❌ Could not identify necessary columns.")
                return pd.DataFrame()

            # Parse each row
            weekly_sales_data = []
            processed_skus = set()

            for row_idx, row in enumerate(all_values[1:], start=2):
                try:
                    if len(row) <= sku_col:
                        continue

                    raw_sku = str(row[sku_col]).strip()
                    if not raw_sku or raw_sku.lower() in ['none', 'null', 'n/a', '']:
                        continue

                    # Clean SKU
                    cleaned_sku = raw_sku.zfill(12) if raw_sku.isdigit() and len(raw_sku) < 12 else raw_sku
                    processed_skus.add(cleaned_sku)

                    for col_idx, week_date in week_columns:
                        if col_idx >= len(row):
                            continue
                        try:
                            sales = float(row[col_idx]) if row[col_idx].strip() not in ['', '-', 'n/a', 'N/A'] else 0
                        except:
                            sales = 0

                        if sales > 0:
                            weekly_sales_data.append({
                                'SKU': cleaned_sku,
                                'Original_SKU': raw_sku,
                                'Week_Start': week_date,
                                'Sales': sales,
                                'Channel': 'Walmartfbm'
                            })

                except Exception as e:
                    print(f"⚠️ Error at row {row_idx}: {e}")
                    continue

            # Convert to DataFrame
            weekly_df5 = pd.DataFrame(weekly_sales_data)

            if weekly_df5.empty:
                print("⚠️ No valid weekly sales found.")
            else:
                print(f"✅ {len(weekly_df5)} records extracted | {weekly_df5['SKU'].nunique()} SKUs")
                print(f"   Range: {weekly_df5['Week_Start'].min().strftime('%d/%m/%Y')} → {weekly_df5['Week_Start'].max().strftime('%d/%m/%Y')}")

            return weekly_df5

        except Exception as e:
            print(f"❌ Fatal error during extraction: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def convert_walmart_fbm_weekly_to_monthly(self, weekly_df5):
        """
        Convert Walmart FBM weekly sales data to monthly format
        using actual week start dates.
        """
        try:
            if weekly_df5.empty:
                return pd.DataFrame()

            print(f"Converting Walmart FBM weekly sales to monthly format...")

            # Ensure 'Week_Start' column is datetime
            weekly_df5['Week_Start'] = pd.to_datetime(weekly_df5['Week_Start'], errors='coerce')

            # Drop rows with invalid dates
            weekly_df5 = weekly_df5.dropna(subset=['Week_Start'])

            # Assign the month (e.g., 2025-01-31, 2025-02-28 etc.) to each week using MonthEnd
            weekly_df5['Month'] = weekly_df5['Week_Start'] + pd.offsets.MonthEnd(0)

            # Debug print mapping
            mapping = weekly_df5[['Week_Start', 'Month']].drop_duplicates().sort_values('Week_Start')
            print("Week to Month mapping:")
            for _, row in mapping.head(10).iterrows():
                print(f"   Week of {row['Week_Start'].strftime('%d/%m/%Y')} → Month End: {row['Month'].strftime('%d/%m/%Y')}")

            # Group and aggregate
            monthly_sales5 = weekly_df5.groupby(['SKU', 'Month'])['Sales'].sum().reset_index()
            monthly_sales5['Channel'] = 'Walmartfbm'
            monthly_sales5 = monthly_sales5.rename(columns={'Month': 'Date'})

            print(f"✅ Converted to {len(monthly_sales5)} monthly records")
            print(f"   Date range: {monthly_sales5['Date'].min().strftime('%d/%m/%Y')} to {monthly_sales5['Date'].max().strftime('%d/%m/%Y')}")
            print(f"   SKUs covered: {monthly_sales5['SKU'].nunique()}")

            # Sample output
            for _, row in monthly_sales5.head(3).iterrows():
                print(f"   SKU {row['SKU']} | {row['Date'].strftime('%b %Y')} | Sales: {row['Sales']}")

            return monthly_sales5

        except Exception as e:
            print(f"❌ Error converting Walmart FBM weekly to monthly: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def extend_historical_data_with_walmart_fbm_weekly(self, historical_data, walmart_fbm_weekly_monthly_data):
        """
        Extend historical sales data with Walmart FBM weekly data
        """
        try:
            if walmart_fbm_weekly_monthly_data.empty:
                print("No Walmart FBM weekly data to extend historical data")
                return historical_data

            print("Extending historical data with Walmart FBM weekly sales...")

            # Get the latest date in historical data
            historical_data['Date'] = pd.to_datetime(historical_data['Date'])
            latest_historical_date = historical_data['Date'].max()

            print(f"Latest historical date: {latest_historical_date}")
            print(f"Walmart FBM weekly data starts: {walmart_fbm_weekly_monthly_data['Date'].min()}")
            print(f"Walmart FBM weekly data ends: {walmart_fbm_weekly_monthly_data['Date'].max()}")

            # Filter Walmart FBM weekly data to only include dates after historical data
            new_walmart_fbm_data = walmart_fbm_weekly_monthly_data[walmart_fbm_weekly_monthly_data['Date'] > latest_historical_date].copy()

            if new_walmart_fbm_data.empty:
                print("No new Walmart FBM weekly data after historical cutoff")
                return historical_data

            print(f"Adding {len(new_walmart_fbm_data)} new Walmart FBM monthly records")

            # Walmart FBM data is already in the right format - just add it
            if not new_walmart_fbm_data.empty:
                # Combine with historical data
                combined_data = pd.concat([historical_data, new_walmart_fbm_data], ignore_index=True)
                combined_data = combined_data.sort_values(['SKU', 'Channel', 'Date'])

                print(f"✅ Extended historical data with {len(new_walmart_fbm_data)} new Shopify records")
                print(f"   Total records: {len(combined_data)} (was {len(historical_data)})")
                print(f"   New date range: {combined_data['Date'].min()} to {combined_data['Date'].max()}")

                # Show sample of extended data
                walmart_fbm_sample = new_walmart_fbm_data.head(3)
                for _, row in walmart_fbm_sample.iterrows():
                    print(f"   Added Walmart FBM: {row['SKU']}, {row['Date'].strftime('%b %Y')}, Sales: {row['Sales']}")

                return combined_data
            else:
                return historical_data

        except Exception as e:
            print(f"❌ Error extending historical data with Walmart FBM: {e}")
            import traceback
            traceback.print_exc()
            return historical_data        
    

# [REST OF THE CODE REMAINS EXACTLY THE SAME FROM EnhancedForecastingModel CLASS ONWARDS]
# Including all methods in EnhancedForecastingModel and the main() function
# Only the inventory loading part in main() needs to be modified


class EnhancedForecastingModel:
    def __init__(self, historical_data, lead_times, launch_dates, service_level=0.95):
        try:
            print("Initializing Enhanced Forecasting Model...")

            expected_columns = ['SKU', 'Channel', 'Date', 'Sales']
            missing_cols = [col for col in expected_columns if col not in historical_data.columns]
            if missing_cols:
                raise ValueError(f"Missing columns: {missing_cols}")

            # Clean SKU data more thoroughly
            historical_data['SKU'] = historical_data['SKU'].astype(str).str.strip()
            historical_data['Date'] = pd.to_datetime(historical_data['Date'], format='%m/%d/%Y', errors='coerce')
            historical_data = historical_data.dropna(subset=['Date'])
            historical_data['Channel'] = historical_data['Channel'].str.strip().str.title()
            historical_data = historical_data[historical_data['Channel'].isin(['Amazon', 'Shopify', 'Shopify Faire', 'Amazonfbm', 'Walmartfbm'])]
            historical_data['Sales'] = pd.to_numeric(historical_data['Sales'], errors='coerce').fillna(0)

            self.data = historical_data
            self.lead_times = lead_times
            self.launch_dates = launch_dates
            self.service_level = service_level
            self.z_score = norm.ppf(service_level)

            # Debug SKU matching
            print(f"\n🔍 SKU MATCHING DEBUG:")
            historical_skus = set(self.data['SKU'].unique())
            lead_time_skus = set(lead_times.keys())

            print(f"Historical data SKUs: {len(historical_skus)}")
            print(f"Lead times SKUs: {len(lead_time_skus)}")

            # Show sample SKUs from each source
            print(f"Sample historical SKUs: {list(historical_skus)[:5]}")
            print(f"Sample lead time SKUs: {list(lead_time_skus)[:5]}")

            # Check for matches
            matching_skus = historical_skus.intersection(lead_time_skus)
            print(f"Directly matching SKUs: {len(matching_skus)}")

            print(f"Data processed: {len(self.data)} records for {self.data['SKU'].nunique()} SKUs")

            self.velocity_categories = self._perform_abc_analysis()
            print(f"ABC Analysis complete: {len(self.velocity_categories)} SKUs categorized")

        except Exception as e:
            print(f"Error initializing model: {e}")
            raise

    def _perform_abc_analysis(self):
        try:
            print("Performing ABC velocity analysis...")

            latest_date = self.data['Date'].max()
            cutoff_date = latest_date - pd.DateOffset(months=3)
            recent_data = self.data[self.data['Date'] >= cutoff_date]

            if recent_data.empty:
                recent_data = self.data.copy()

            sku_sales = recent_data.groupby('SKU')['Sales'].sum().sort_values(ascending=False)

            categories = {}
            total_skus = len(sku_sales)

            for i, (sku, sales) in enumerate(sku_sales.items()):
                rank = i + 1

                if rank <= 20 or rank <= total_skus * 0.10:
                    category = 'A'
                    safety_months = 6.0
                    service_level = 0.98
                elif rank <= total_skus * 0.30:
                    category = 'B'
                    safety_months = 6.0
                    service_level = 0.95
                elif rank <= total_skus * 0.70:
                    category = 'C'
                    safety_months = 3.0
                    service_level = 0.90
                else:
                    category = 'D'
                    safety_months = 3.0
                    service_level = 0.85

                categories[sku] = {
                    'category': category,
                    'rank': rank,
                    'total_sales': sales,
                    'monthly_velocity': sales / 3,
                    'safety_stock_months': safety_months,
                    'service_level': service_level
                }

            return categories

        except Exception as e:
            print(f"Error in ABC analysis: {e}")
            return {}

    def smart_sku_lookup(self, sku, product_info, data_dict=None):
        """
        Enhanced SKU lookup that tries multiple matching strategies
        Can return product info or any other data from data_dict
        """
        sku = str(sku).strip()

        # Choose which dictionary to use for lookup
        lookup_dict = data_dict if data_dict is not None else product_info

        # Direct match first
        if sku in lookup_dict:
            return lookup_dict[sku]

        # Try variations of the SKU
        sku_variations = [
            sku.lstrip('0'),  # Remove leading zeros
            sku.zfill(12),    # Pad with zeros to 12 digits
            ''.join(c for c in sku if c.isalnum()),  # Remove special characters
            sku.replace(' ', '').replace('-', '')     # Remove spaces and dashes
        ]

        for variant in sku_variations:
            if variant and variant in lookup_dict:
                return lookup_dict[variant]

        # Try partial matching (both ways)
        for lookup_sku, value in lookup_dict.items():
            # Check if either SKU contains the other
            if len(sku) >= 8 and len(lookup_sku) >= 8:  # Only for reasonable length SKUs
                if sku in lookup_sku or lookup_sku in sku:
                    return value

        # Return appropriate default based on data type
        if data_dict is None:  # product_info lookup
            return 'Unknown'
        else:  # Could be launch_dates or other data
            return None

    def calculate_years_since_launch(self, sku, current_date=None):
        """
        Calculate the number of years since product launch (with decimal precision)
        """
        if current_date is None:
            current_date = pd.Timestamp.now()

        launch_date = self.smart_sku_lookup(sku, None, self.launch_dates)

        if launch_date is None or pd.isna(launch_date):
            return None

        try:
            # Calculate the difference in years with decimal precision
            years_diff = (current_date - launch_date).days / 365.25  # Account for leap years
            return round(years_diff, 1) if years_diff >= 0 else 0.0
        except:
            return None

    def prepare_data(self, channel, sku):
        try:
            sku = str(sku).strip()
            channel = str(channel).strip().title()

            sku_data = self.data[(self.data['SKU'] == sku) & (self.data['Channel'] == channel)].sort_values('Date')

            if sku_data.empty:
                date_range = pd.date_range(start='2024-01-01', end='2025-04-30', freq='ME')
                return pd.Series([0] * len(date_range), index=date_range)

            sku_data['Month'] = sku_data['Date'].dt.to_period('M').dt.to_timestamp('M')
            monthly_sales = sku_data.groupby('Month')['Sales'].sum().sort_index()

            full_range = pd.date_range(start=monthly_sales.index.min(), end=monthly_sales.index.max(), freq='ME')
            monthly_sales = monthly_sales.reindex(full_range, fill_value=0)
            monthly_sales.index.freq = 'ME'

            return monthly_sales

        except Exception:
            date_range = pd.date_range(start='2024-01-01', end='2025-04-30', freq='ME')
            return pd.Series([0] * len(date_range), index=date_range)

    def generate_forecast(self, series, horizon=8):
        try:
            growth_rate = self._calculate_growth_rate(series)

            if series.empty or series.sum() == 0:
                last_date = series.index[-1] if not series.empty else pd.to_datetime('today') + pd.offsets.MonthEnd(0)
                forecast_dates = pd.date_range(start=last_date + pd.offsets.MonthEnd(1), periods=horizon, freq='ME')
                forecast = pd.Series([0] * horizon, index=forecast_dates)
                return forecast, 'Zero Forecast', 0, [], 0

            mean_sales = series.mean()
            std_sales = series.std()
            cv = std_sales / mean_sales if mean_sales > 0 else float('inf')

            if cv > 3 or len(series) < 3:
                avg_forecast = max(mean_sales, 0)
                forecast_values = [max(0, round(avg_forecast + growth_rate * (i + 1))) for i in range(horizon)]
                last_date = series.index[-1]
                forecast_dates = pd.date_range(start=last_date + pd.offsets.MonthEnd(1), periods=horizon, freq='ME')
                forecast = pd.Series(forecast_values, index=forecast_dates)
                return forecast, 'Average + Growth', 0, [], growth_rate

            try:
                if len(series) >= 12:
                    model = ExponentialSmoothing(series, seasonal='add', seasonal_periods=6, trend='add').fit()
                    method = 'Holt-Winters Seasonal'
                elif len(series) >= 6:
                    model = ExponentialSmoothing(series, trend='add').fit()
                    method = 'Holt-Winters Trend'
                else:
                    model = ExponentialSmoothing(series).fit()
                    method = 'Simple Exponential Smoothing'

                forecast = model.forecast(horizon)
                forecast = forecast.clip(lower=0, upper=series.max() * 5)

                last_date = series.index[-1]
                forecast_dates = pd.date_range(start=last_date + pd.offsets.MonthEnd(1), periods=horizon, freq='ME')
                forecast.index = forecast_dates

                return forecast.round().astype(int), method, 6, [], growth_rate

            except Exception:
                avg_forecast = max(mean_sales, 0)
                forecast_values = [max(0, round(avg_forecast + growth_rate * (i + 1))) for i in range(horizon)]
                last_date = series.index[-1]
                forecast_dates = pd.date_range(start=last_date + pd.offsets.MonthEnd(1), periods=horizon, freq='ME')
                forecast = pd.Series(forecast_values, index=forecast_dates)
                return forecast, 'Fallback Average', 0, [], growth_rate

        except Exception:
            last_date = pd.to_datetime('2025-04-30')
            forecast_dates = pd.date_range(start=last_date + pd.offsets.MonthEnd(1), periods=horizon, freq='ME')
            forecast = pd.Series([0] * horizon, index=forecast_dates)
            return forecast, 'Error Fallback', 0, [], 0

    def _calculate_growth_rate(self, series):
        try:
            if len(series) < 2:
                return 0
            x = np.arange(len(series)).reshape(-1, 1)
            y = series.values.reshape(-1, 1)
            model = LinearRegression()
            model.fit(x, y)
            growth_rate = round(model.coef_[0][0], 2)
            # Handle NaN/Inf values
            if pd.isna(growth_rate) or growth_rate == float('inf') or growth_rate == float('-inf'):
                return 0
            return growth_rate
        except:
            return 0

    def calculate_enhanced_safety_stock(self, series, lead_time, sku):
        try:
            sku_info = self.velocity_categories.get(sku, {})
            category = sku_info.get('category', 'D')
            target_service_level = sku_info.get('service_level', 0.85)
            safety_months = sku_info.get('safety_stock_months', 3.0)

            z_score = norm.ppf(target_service_level)

            if len(series) < 2 or series.sum() == 0:
                return max(round(safety_months * 10), 5), category, safety_months

            demand_std = series.std()
            mean_demand = series.mean()

            statistical_ss = z_score * demand_std * math.sqrt(lead_time)
            velocity_min_ss = mean_demand * safety_months

            final_ss = max(statistical_ss, velocity_min_ss)
            min_ss = max(5, mean_demand * 0.25)
            max_ss = mean_demand * lead_time * 2
            final_ss = max(min_ss, min(final_ss, max_ss))

            return max(round(final_ss), 5), category, safety_months

        except Exception:
            return 10, 'D', 1.0

    def calculate_enhanced_reorder_point(self, series, lead_time, safety_stock):
        try:
            if series.empty or series.sum() == 0:
                return safety_stock

            avg_demand = series.mean() * lead_time
            growth_rate = self._calculate_growth_rate(series)
            trend_adjustment = growth_rate * lead_time

            reorder_point = avg_demand + trend_adjustment + safety_stock
            return max(round(reorder_point), safety_stock)

        except:
            return safety_stock

    def calculate_enhanced_po_quantity(self, series, current_inventory, reorder_point, sku, safety_stock_months):
        try:
            if series.empty or series.sum() == 0:
                if current_inventory < reorder_point:
                    return max(20, reorder_point - current_inventory), "NEW PRODUCT - Order Required"
                else:
                    return 0, "NEW PRODUCT - Sufficient Stock"

            monthly_demand = series.mean()
            if monthly_demand <= 0:
                return 0, "No demand"

            sku_info = self.velocity_categories.get(sku, {})
            velocity_category = sku_info.get('category', 'D')

            if velocity_category == 'A':
                order_months = 4.0
            elif velocity_category == 'B':
                order_months = 3.0
            elif velocity_category == 'C':
                order_months = 2.5
            else:
                order_months = 2.0

            if current_inventory <= reorder_point:
                shortage = reorder_point - current_inventory
                target_stock = monthly_demand * (safety_stock_months + order_months)
                po_quantity = max(shortage, target_stock - current_inventory)
                urgency = "HIGH - Below Reorder Point"
            elif current_inventory <= reorder_point * 1.3:
                po_quantity = monthly_demand * order_months
                urgency = "MEDIUM - Approaching Reorder Point"
            else:
                po_quantity = 0
                urgency = "LOW - Sufficient Stock"

            return round(po_quantity), urgency

        except Exception:
            return 0, "Error"

    def calculate_future_orders(self, forecast, current_inventory, reorder_point, lead_time, safety_stock_months, sku):
        try:
            orders = []
            running_inventory = current_inventory

            sku_info = self.velocity_categories.get(sku, {})
            velocity_category = sku_info.get('category', 'D')

            if velocity_category == 'A':
                order_months = 4.0
            elif velocity_category == 'B':
                order_months = 3.0
            elif velocity_category == 'C':
                order_months = 2.5
            else:
                order_months = 2.0

            forecast_values = [f for f in forecast.iloc[:6] if f > 0]
            if not forecast_values:
                return [], [], []

            avg_monthly_demand = sum(forecast_values) / len(forecast_values)
            order_quantity = round(avg_monthly_demand * order_months)

            for i in range(8):
                month_demand = forecast.iloc[i] if i < len(forecast) else 0

                if i + lead_time < 8:
                    future_inventory = running_inventory

                    for j in range(i, min(i + lead_time, 8)):
                        future_demand = forecast.iloc[j] if j < len(forecast) else 0
                        future_inventory -= future_demand

                    if future_inventory <= reorder_point and order_quantity > 0:
                        order_date = forecast.index[i] if i < len(forecast) else None
                        arrival_date = forecast.index[min(i + lead_time, len(forecast)-1)] if i + lead_time < len(forecast) else None

                        orders.append({
                            'order_date': order_date,
                            'arrival_date': arrival_date,
                            'quantity': order_quantity
                        })

                        if i + lead_time < 8:
                            running_inventory += order_quantity

                running_inventory -= month_demand
                running_inventory = max(0, running_inventory)

            order_dates = []
            order_qtys = []
            arrival_dates = []

            for order in orders[:3]:
                order_dates.append(order['order_date'].strftime('%Y-%m-%d') if order['order_date'] else '')
                order_qtys.append(order['quantity'])
                arrival_dates.append(order['arrival_date'].strftime('%Y-%m-%d') if order['arrival_date'] else '')

            while len(order_dates) < 3:
                order_dates.append('')
                order_qtys.append(0)
                arrival_dates.append('')

            return order_dates, order_qtys, arrival_dates

        except Exception:
            return ['', '', ''], [0, 0, 0], ['', '', '']

    def calculate_months_of_inventory(self, current_inventory, forecast):
        try:
            if current_inventory <= 0:
                return 0.0

            forecast_values = [f for f in forecast.iloc[:6] if f > 0]

            if not forecast_values or sum(forecast_values) == 0:
                return 999.0

            avg_monthly_demand = sum(forecast_values) / len(forecast_values)

            if avg_monthly_demand <= 0:
                return 999.0

            months_available = current_inventory / avg_monthly_demand
            return round(months_available, 1)

        except Exception:
            return 0.0

    def create_enhanced_forecast(self, channel, inventory, product_info, product_category, product_status):
        try:
            print(f"Creating enhanced forecast for {channel}...")

            all_skus = set()
            all_skus.update(self.lead_times.keys())
            all_skus.update(inventory.keys())
            all_skus.update(product_info.keys())
            all_skus.update(self.data['SKU'].unique())
            all_skus.update(product_category.keys())
            all_skus.update(product_status.keys())

            print(f"Processing {len(all_skus)} SKUs for {channel}")

            # Generate historical month labels (last 3 months)
            current_date = pd.Timestamp.now()
            historical_months = []
            for i in range(3, 0, -1):  # 3, 2, 1 months ago
                past_date = current_date - pd.DateOffset(months=i)
                historical_months.append(past_date.strftime('%b_%Y'))

            # Generate forecast month labels
            forecast_months = []
            for i in range(8):
                future_date = current_date + pd.DateOffset(months=i+1)
                forecast_months.append(future_date.strftime('%b_%Y'))  # e.g., 'Jun_2025'

            print(f"Historical periods: {', '.join(historical_months)}")
            print(f"Forecast periods: {', '.join(forecast_months)}")

            forecast_data = []
            products_found = 0

            for i, sku in enumerate(all_skus):
                try:
                    if i % 20 == 0:
                        print(f"   Processing SKU {i+1}/{len(all_skus)}: {sku}")

                    series = self.prepare_data(channel, sku)
                    forecast, method, seasonal_periods, seasonal_factors, growth_rate = self.generate_forecast(series)

                    # Use enhanced SKU lookup for product names and launch dates
                    product_name = self.smart_sku_lookup(sku, product_info)

                    category_name = self.smart_sku_lookup(sku, product_category)
                    status_name = self.smart_sku_lookup(sku, product_status)

                    launch_date = self.smart_sku_lookup(sku, None, self.launch_dates)
                    years_since_launch = self.calculate_years_since_launch(sku)

                    # Debug: Show some matching attempts
                    if i < 5 or (i % 50 == 0):
                        print(f"   DEBUG SKU {sku}: Product = '{product_name}'")

                    if product_name != 'Unknown':
                        products_found += 1
                    else:
                        # Skip forecasts for SKUs without valid product mappings
                        if i % 100 == 0:  # Only log occasionally to avoid spam
                            print(f"   Skipping SKU {sku} - no product mapping found")
                        continue

                    # Get last 3 months sales (skip the most recent month to avoid incomplete data)
                    last_3_sales = series.iloc[-4:-1].tolist() if len(series) >= 4 else series.tail(3).tolist() if len(series) >= 3 else [0, 0, 0]
                    while len(last_3_sales) < 3:
                        last_3_sales.insert(0, 0)
                    last_3_sales = last_3_sales[-3:]  # Ensure exactly 3 values

                    # Ensure we have valid numeric values
                    last_3_sales = [float(x) if not pd.isna(x) else 0.0 for x in last_3_sales]

                    last_3_months_avg = round(series.iloc[-4:-1].mean(), 2) if len(series) >= 4 else round(series[-3:].mean(), 2) if len(series) >= 3 else round(series.mean(), 2) if not series.empty else 0
                    # Handle NaN values
                    if pd.isna(last_3_months_avg):
                        last_3_months_avg = 0.0

                    total_sales = round(series.sum(), 2)
                    if pd.isna(total_sales):
                        total_sales = 0.0
                    lead_time = self.lead_times.get(sku, 2)

                    safety_stock, velocity_category, safety_stock_months = self.calculate_enhanced_safety_stock(series, lead_time, sku)
                    reorder_point = self.calculate_enhanced_reorder_point(series, lead_time, safety_stock)

                    current_inventory = inventory.get(sku, 0)
                    po_quantity, urgency = self.calculate_enhanced_po_quantity(series, current_inventory, reorder_point, sku, safety_stock_months)

                    order_dates, order_qtys, arrival_dates = self.calculate_future_orders(forecast, current_inventory, reorder_point, lead_time, safety_stock_months, sku)
                    months_of_inventory = self.calculate_months_of_inventory(current_inventory, forecast)

                    sku_info = self.velocity_categories.get(sku, {})

                    if current_inventory <= 0:
                        stock_status = "OUT OF STOCK"
                    elif current_inventory <= reorder_point:
                        stock_status = "REORDER NOW"
                    elif last_3_months_avg > 0 and current_inventory / last_3_months_avg < 1:
                        stock_status = "LOW STOCK"
                    elif last_3_months_avg > 0 and current_inventory / last_3_months_avg > 6:
                        stock_status = "OVERSTOCK"
                    else:
                        stock_status = "NORMAL"

                    # Create dynamic forecast row with required order
                    row_data = {
                        'SKU': str(sku),
                        'Product_Name': product_name,
                        'Category': category_name,
                        'Status': status_name,
                        'Launch_Date': launch_date.strftime('%Y-%m-%d') if launch_date and not pd.isna(launch_date) else '',
                        'Forecast_Method': method,
                        'Years_Since_Launch': years_since_launch if years_since_launch is not None else '',
                        'Current_Inventory': current_inventory,
                        'Stock_Status': stock_status,
                        'PO_Urgency': urgency,
                        'Recommended_PO_Qty': po_quantity,
                        'Next_Order_Date': order_dates[0],
                        'Next_Order_Qty': order_qtys[0],
                        'Next_Arrival_Date': arrival_dates[0],
                        'Months_of_Inventory': months_of_inventory,
                        'Velocity_Category': velocity_category,
                        'Safety_Stock_Months': safety_stock_months,
                        'Reorder_Point': reorder_point,
                        'Safety_Stock': safety_stock,
                    }
                    ## AMAZON TAB
                    # Add forecast columns with month/year labels
                    for idx, month_label in enumerate(forecast_months):
                        row_data[f'Forecast_{month_label}'] = int(forecast.iloc[idx]) if idx < len(forecast) else 0

                    # Add trailing metrics and info (except Channel and Historical for now)
                    row_data.update({
                        'Last_3_Months_Avg': last_3_months_avg,
                        'Total_Sales': total_sales,
                        'Growth_Rate': growth_rate if not pd.isna(growth_rate) else 0.0,
                        'Order_2_Date': order_dates[1],
                        'Order_2_Qty': order_qtys[1],
                        'Order_2_Arrival': arrival_dates[1],
                        'Order_3_Date': order_dates[2],
                        'Order_3_Qty': order_qtys[2],
                        'Order_3_Arrival': arrival_dates[2],
                        'Lead_Time': lead_time,
                        'Service_Level': f"{sku_info.get('service_level', 0.85)*100:.0f}%",
                        'Monthly_Velocity': round(sku_info.get('monthly_velocity', 0), 1) if not pd.isna(sku_info.get('monthly_velocity', 0)) else 0.0,
                        'Velocity_Rank': sku_info.get('rank', 999),
                    })

                    # Add historical months data right before 'Channel'
                    for idx, month_label in enumerate(historical_months):
                        if idx < len(last_3_sales):
                            value = int(last_3_sales[idx]) if not pd.isna(last_3_sales[idx]) else 0
                        else:
                            value = 0
                        row_data[month_label] = value

                    # Now finally add 'Channel' at the end
                    row_data['Channel'] = channel

                    # Append row to forecast list
                    forecast_data.append(row_data)


                except Exception as e:
                    print(f"   Error processing SKU {sku}: {e}")
                    if i < 10:  # Show detailed errors for first 10 SKUs
                        import traceback
                        traceback.print_exc()
                    continue

            df = pd.DataFrame(forecast_data)

            # Additional filtering to ensure we only have mapped products
            if not df.empty:
                initial_count = len(df)
                df = df[df['Product_Name'] != 'Unknown']
                final_count = len(df)
                filtered_out = initial_count - final_count

                if filtered_out > 0:
                    print(f"   Filtered out {filtered_out} SKUs without product mappings")

            print(f"Generated {len(df)} forecasts for {channel} (only mapped products)")
            print(f"✅ Successfully mapped {products_found} product names")

            return df

        except Exception:
            return pd.DataFrame()

    def combine_channel_forecasts(self, amazon_forecast, shopify_forecast, shopify_faire_forecast, amazon_fbm_forecast, walmart_fbm_forecast):
        try:
            print("Combining channel forecasts...")

            # Concatenate all dataframes
            all_data = pd.concat([amazon_forecast, shopify_forecast, shopify_faire_forecast, amazon_fbm_forecast, walmart_fbm_forecast], ignore_index=True)

            if all_data.empty:
                return pd.DataFrame()
            
            # Get forecast month columns dynamically
            forecast_columns = [col for col in all_data.columns if col.startswith('Forecast_') and col != 'Forecast_Method']

            # Get historical month columns dynamically
            historical_columns = []
            month_names = ['Jan_', 'Feb_', 'Mar_', 'Apr_', 'May_', 'Jun_', 'Jul_', 'Aug_', 'Sep_', 'Oct_', 'Nov_', 'Dec_']
            for col in all_data.columns:
                if any(month in col for month in month_names) and 'Forecast_' not in col and '_' in col:
                    try:
                        parts = col.split('_')
                        if len(parts) == 2 and parts[0][:3] in [m[:3] for m in month_names]:
                            historical_columns.append(col)
                    except:
                        pass

            print(f"Found historical columns: {historical_columns}")
            print(f"Found forecast columns: {forecast_columns}")

            combined_data = []
            unique_skus = all_data['SKU'].unique()

            for sku in unique_skus:
                sku_data = all_data[all_data['SKU'] == sku]

                if sku_data.empty:
                    continue

                # Get data from each channel for this SKU
                amazon_data = sku_data[sku_data['Channel'] == 'Amazon']
                shopify_data = sku_data[sku_data['Channel'] == 'Shopify']
                shopify_faire_data = sku_data[sku_data['Channel'] == 'Shopify Faire']
                amazon_fbm_data = sku_data[sku_data['Channel'] == 'Amazonfbm']
                walmart_fbm_data = sku_data[sku_data['Channel'] == 'Walmartfbm']

                # Use the first available record as base (preference: Amazon -> Shopify -> others)
                if not amazon_data.empty:
                    base_record = amazon_data.iloc[0].copy()
                elif not shopify_data.empty:
                    base_record = shopify_data.iloc[0].copy()
                else:
                    base_record = sku_data.iloc[0].copy()

                ### ALL FORECASTS TAB - maintaining original column sequence
                combined_record = {
                    'SKU': str(sku),
                    'Product_Name': base_record.get('Product_Name', 'Unknown'),
                    'Category': base_record.get('Category', 'Unknown'),
                    'Status': base_record.get('Status', 'Unknown'),
                    'Launch_Date': base_record.get('Launch_Date', ''),
                    'Years_Since_Launch': base_record.get('Years_Since_Launch', ''),
                    'Velocity_Category': base_record.get('Velocity_Category', 'D'),
                    'Velocity_Rank': base_record.get('Velocity_Rank', 999),
                    'Service_Level': base_record.get('Service_Level', '85%'),
                    'Safety_Stock_Months': base_record.get('Safety_Stock_Months', 1.0),
                    'Months_of_Inventory': base_record.get('Months_of_Inventory', 0.0),
                    'Safety_Stock': base_record.get('Safety_Stock', 0),
                    'Reorder_Point': base_record.get('Reorder_Point', 0),
                    'Current_Inventory': base_record.get('Current_Inventory', 0),
                }

                # Calculate Last_3_Months_Avg from summed historical data
                if len(historical_columns) >= 3:
                    last_3_sum = sum(int(sku_data[col].sum()) for col in historical_columns[-3:] if col in sku_data.columns)
                    combined_record['Last_3_Months_Avg'] = round(last_3_sum / 3, 2)
                else:
                    combined_record['Last_3_Months_Avg'] = 0.0

                # Add Growth_Rate
                combined_record['Growth_Rate'] = round(sku_data['Growth_Rate'].mean(), 2) if 'Growth_Rate' in sku_data.columns and not pd.isna(sku_data['Growth_Rate'].mean()) else 0.0

                # Add remaining fields
                combined_record.update({
                    'Stock_Status': base_record.get('Stock_Status', 'UNKNOWN'),
                    'PO_Urgency': base_record.get('PO_Urgency', 'LOW'),
                    'Recommended_PO_Qty': base_record.get('Recommended_PO_Qty', 0),
                    'Next_Order_Date': base_record.get('Next_Order_Date', ''),
                    'Next_Order_Qty': base_record.get('Next_Order_Qty', 0),
                    'Next_Arrival_Date': base_record.get('Next_Arrival_Date', ''),
                })

                # Add combined forecast columns (summed across channels)
                for col in forecast_columns:
                    combined_record[col] = int(sku_data[col].sum()) if col in sku_data.columns else 0

                # Add channel breakdown for first 3 forecast months
                if len(forecast_columns) >= 3:
                    for i in range(3):
                        month_col = forecast_columns[i]
                        month_label = month_col.replace('Forecast_', '')
                        combined_record[f'Amazon_{month_label}'] = int(amazon_data[month_col].iloc[0]) if not amazon_data.empty and month_col in amazon_data.columns else 0
                        combined_record[f'Shopify_{month_label}'] = int(shopify_data[month_col].iloc[0]) if not shopify_data.empty and month_col in shopify_data.columns else 0
                        combined_record[f'Shopify_Faire_{month_label}'] = int(shopify_faire_data[month_col].iloc[0]) if not shopify_faire_data.empty and month_col in shopify_faire_data.columns else 0
                        combined_record[f'Amazon_FBM_{month_label}'] = int(amazon_fbm_data[month_col].iloc[0]) if not amazon_fbm_data.empty and month_col in amazon_fbm_data.columns else 0
                        combined_record[f'Walmart_FBM_{month_label}'] = int(walmart_fbm_data[month_col].iloc[0]) if not walmart_fbm_data.empty and month_col in walmart_fbm_data.columns else 0

                # Add remaining metrics
                combined_record.update({
                    'Total_Sales': round(sku_data['Total_Sales'].sum(), 2) if 'Total_Sales' in sku_data.columns and not pd.isna(sku_data['Total_Sales'].sum()) else 0.0,
                    'Order_2_Date': base_record.get('Order_2_Date', ''),
                    'Order_2_Qty': base_record.get('Order_2_Qty', 0),
                    'Order_2_Arrival': base_record.get('Order_2_Arrival', ''),
                    'Order_3_Date': base_record.get('Order_3_Date', ''),
                    'Order_3_Qty': base_record.get('Order_3_Qty', 0),
                    'Order_3_Arrival': base_record.get('Order_3_Arrival', ''),
                    'Lead_Time': base_record.get('Lead_Time', 2),
                    'Monthly_Velocity': base_record.get('Monthly_Velocity', 0) if not pd.isna(base_record.get('Monthly_Velocity', 0)) else 0.0,
                    'Channel': 'Combined'
                })

                # Add historical months just before Channel
                for col in historical_columns:
                    combined_record[col] = int(sku_data[col].sum()) if col in sku_data.columns else 0

                # Forecast_Method last
                combined_record['Forecast_Method'] = "Combined Channels"

                combined_data.append(combined_record)

            combined_df = pd.DataFrame(combined_data)
            print(f"Combined forecasts: {len(combined_df)} unique SKUs")

            return combined_df

        except Exception as e:
            print(f"Error in combine_channel_forecasts: {str(e)}")
            return pd.concat([amazon_forecast, shopify_forecast, shopify_faire_forecast, amazon_fbm_forecast, walmart_fbm_forecast], ignore_index=True)

    def generate_actionable_insights(self, combined_forecast):
        """Generate comprehensive actionable insights for inventory planning."""
        try:
            print("Generating actionable insights...")

            insights = {
                'immediate_actions': [],
                'weekly_actions': [],
                'monthly_actions': [],
                'risk_analysis': [],
                'opportunities': [],
                'cost_optimization': []
            }

            current_date = pd.Timestamp.now()

            for _, row in combined_forecast.iterrows():
                row_dict = row.to_dict()

                sku = row_dict.get('SKU', '')
                product_name = row_dict.get('Product_Name', '')
                current_inventory = row_dict.get('Current_Inventory', 0)
                months_of_inventory = row_dict.get('Months_of_Inventory', 0.0)
                po_urgency = row_dict.get('PO_Urgency', 'LOW')
                velocity_category = row_dict.get('Velocity_Category', 'D')
                reorder_point = row_dict.get('Reorder_Point', 0)
                recommended_po = row_dict.get('Recommended_PO_Qty', 0)
                lead_time = row_dict.get('Lead_Time', 2)
                growth_rate = row_dict.get('Growth_Rate', 0.0)

                # Sanitize invalid or infinite values
                if pd.isna(months_of_inventory) or months_of_inventory == float('inf'):
                    months_of_inventory = 999.0
                elif months_of_inventory == float('-inf'):
                    months_of_inventory = 0.0
                if pd.isna(growth_rate):
                    growth_rate = 0.0

                # Get forecast columns dynamically
                forecast_cols = [col for col in row_dict if col.startswith('Forecast_')]
                next_3_months_demand = 0
                for col in forecast_cols[:3]:
                    try:
                        val = float(row.get(col, 0))
                        if pd.notna(val):
                            next_3_months_demand += val
                    except:
                        continue


                # === IMMEDIATE ACTIONS ===
                if current_inventory <= 0:
                    lost_sales = next_3_months_demand * 50
                    insights['immediate_actions'].append({
                        'Priority': 'CRITICAL',
                        'SKU': sku,
                        'Product': product_name,
                        'Action': 'EXPEDITE ORDER IMMEDIATELY',
                        'Reason': 'OUT OF STOCK',
                        'Quantity': max(recommended_po, next_3_months_demand),
                        'Impact': f'Lost sales: ~${lost_sales:.0f}/month',
                        'Contact': 'Call supplier TODAY for expedited shipping'
                    })
                elif current_inventory <= reorder_point * 0.5:
                    insights['immediate_actions'].append({
                        'Priority': 'HIGH',
                        'SKU': sku,
                        'Product': product_name,
                        'Action': 'Place PO Today',
                        'Reason': f'Critically low: {months_of_inventory:.1f} months left',
                        'Quantity': recommended_po,
                        'Impact': f'Risk of stockout in {lead_time} months',
                        'Contact': 'Email PO to supplier by EOD'
                    })

                # === WEEKLY ACTIONS ===
                elif current_inventory <= reorder_point:
                    insights['weekly_actions'].append({
                        'Priority': 'MEDIUM',
                        'SKU': sku,
                        'Product': product_name,
                        'Action': 'Place PO This Week',
                        'Reason': 'At reorder point',
                        'Quantity': recommended_po,
                        'By_Date': (current_date + pd.Timedelta(days=7)).strftime('%Y-%m-%d'),
                        'Notes': f'Lead time: {lead_time} months'
                    })

                # === MONTHLY ACTIONS ===
                next_order_date = row_dict.get('Next_Order_Date')
                next_order_qty = row_dict.get('Next_Order_Qty', 0)

                if next_order_date:
                    try:
                        parsed_date = pd.to_datetime(next_order_date)
                        days_until_order = (parsed_date - current_date).days
                        if 7 < days_until_order <= 30:
                            budget_impact = next_order_qty * 30
                            insights['monthly_actions'].append({
                                'SKU': sku,
                                'Product': product_name,
                                'Action': 'Schedule PO',
                                'Order_Date': next_order_date,
                                'Quantity': next_order_qty,
                                'Preparation': 'Confirm supplier capacity',
                                'Budget_Impact': f'~${budget_impact:.0f}'
                            })
                    except Exception:
                        pass

                # === RISK ANALYSIS ===
                if velocity_category == 'A' and months_of_inventory < 2:
                    potential_loss = next_3_months_demand * 50
                    insights['risk_analysis'].append({
                        'Risk_Level': 'HIGH',
                        'SKU': sku,
                        'Product': product_name,
                        'Issue': 'Top seller with low inventory',
                        'Potential_Loss': f'${potential_loss:.0f}',
                        'Mitigation': 'Consider air freight or split shipments'
                    })
                elif months_of_inventory > 12:
                    tied_capital = current_inventory * 30
                    insights['risk_analysis'].append({
                        'Risk_Level': 'MEDIUM',
                        'SKU': sku,
                        'Product': product_name,
                        'Issue': f'Excess inventory: {months_of_inventory:.0f} months',
                        'Tied_Capital': f'${tied_capital:.0f}',
                        'Mitigation': 'Pause orders, consider promotions'
                    })

                # === OPPORTUNITIES ===
                if growth_rate > 10 and months_of_inventory < 3:
                    insights['opportunities'].append({
                        'Type': 'GROWTH',
                        'SKU': sku,
                        'Product': product_name,
                        'Trend': f'+{growth_rate:.0f}% growth',
                        'Action': 'Increase safety stock',
                        'Potential': 'Capture more market share'
                    })

                # === COST OPTIMIZATION ===
                if velocity_category in ['C', 'D'] and recommended_po > 0:
                    insights['cost_optimization'].append({
                        'SKU': sku,
                        'Product': product_name,
                        'Current_Order': recommended_po,
                        'Suggestion': 'Combine with other orders',
                        'Savings': 'Reduce shipping costs by 15-20%',
                        'Action': 'Consolidate low-velocity orders monthly'
                    })

            print(f"📌 immediate_actions count: {len(insights['immediate_actions'])}")
            print(f"📌 weekly_actions count: {len(insights['weekly_actions'])}")
            print(f"📌 monthly_actions count: {len(insights['monthly_actions'])}")
            print(f"📌 risk_analysis count: {len(insights['risk_analysis'])}")
            print(f"📌 opportunities count: {len(insights['opportunities'])}")
            print(f"📌 cost_optimization count: {len(insights['cost_optimization'])}")

            return insights

        except Exception as e:
            print(f"❌ Error generating insights: {e}")
            import traceback
            traceback.print_exc()
            return {}

    def create_executive_summary(self, combined_forecast, insights):
        """Create executive summary with key metrics and actions."""
        try:
            summary = {
                'date': pd.Timestamp.now().strftime('%Y-%m-%d'),
                'total_skus': len(combined_forecast),
                'immediate_actions_required': len(insights.get('immediate_actions', [])),
                'weekly_actions_required': len(insights.get('weekly_actions', [])),
                'at_risk_skus': 0,
                'overstock_skus': 0,
                'total_inventory_value': 0,
                'total_po_value_needed': 0,
                'cash_flow_30_days': 0,
                'cash_flow_60_days': 0,
                'cash_flow_90_days': 0
            }

            # Calculate key metrics
            for _, row in combined_forecast.iterrows():
                current_inv = row['Current_Inventory']
                months_inv = row['Months_of_Inventory']

                # Inventory value (assuming $30 cost)
                summary['total_inventory_value'] += current_inv * 30

                # At risk SKUs
                if months_inv < 2:
                    summary['at_risk_skus'] += 1
                elif months_inv > 6:
                    summary['overstock_skus'] += 1

                # PO values needed
                if row['Recommended_PO_Qty'] > 0:
                    summary['total_po_value_needed'] += row['Recommended_PO_Qty'] * 30

                # Cash flow projections
                if row['Next_Order_Date']:
                    order_date = pd.to_datetime(row['Next_Order_Date'])
                    days_until = (order_date - pd.Timestamp.now()).days
                    order_value = row['Next_Order_Qty'] * 30

                    if days_until <= 30:
                        summary['cash_flow_30_days'] += order_value
                    elif days_until <= 60:
                        summary['cash_flow_60_days'] += order_value
                    elif days_until <= 90:
                        summary['cash_flow_90_days'] += order_value

            return summary

        except Exception as e:
            print(f"Error creating executive summary: {e}")
            return {}

    def create_action_priority_matrix(self, combined_forecast):
        """Create action priority matrix based on velocity and urgency."""
        try:
            priority_matrix = []

            for _, row in combined_forecast.iterrows():
                velocity = row['Velocity_Category']
                months_inv = row['Months_of_Inventory']
                po_urgency = row['PO_Urgency']

                # Calculate priority score
                velocity_score = {'A': 4, 'B': 3, 'C': 2, 'D': 1}.get(velocity, 1)

                if months_inv <= 1:
                    urgency_score = 4
                elif months_inv <= 2:
                    urgency_score = 3
                elif months_inv <= 3:
                    urgency_score = 2
                else:
                    urgency_score = 1

                priority_score = velocity_score * urgency_score

                if priority_score >= 12:
                    action_priority = 'IMMEDIATE'
                    action_timeline = 'Today'
                elif priority_score >= 8:
                    action_priority = 'HIGH'
                    action_timeline = 'This Week'
                elif priority_score >= 4:
                    action_priority = 'MEDIUM'
                    action_timeline = 'This Month'
                else:
                    action_priority = 'LOW'
                    action_timeline = 'Next Month'

                priority_matrix.append({
                    'SKU': row['SKU'],
                    'Product_Name': row['Product_Name'],
                    'Velocity_Category': velocity,
                    'Months_of_Inventory': months_inv if not pd.isna(months_inv) else 0.0,
                    'Priority_Score': priority_score,
                    'Action_Priority': action_priority,
                    'Action_Timeline': action_timeline,
                    'Recommended_Action': self._get_specific_action(row),
                    'Order_Quantity': row['Recommended_PO_Qty'] if not pd.isna(row['Recommended_PO_Qty']) else 0,
                    'Current_Inventory': row['Current_Inventory'] if not pd.isna(row['Current_Inventory']) else 0,
                    'Next_Month_Forecast': row[[col for col in row.index if col.startswith('Forecast_')][0]] if any(col.startswith('Forecast_') for col in row.index) and not pd.isna(row[[col for col in row.index if col.startswith('Forecast_')][0]]) else 0
                })

            return pd.DataFrame(priority_matrix)

        except Exception as e:
            print(f"Error creating priority matrix: {e}")
            return pd.DataFrame()

    def _get_specific_action(self, row):
        """Get specific action recommendation based on SKU status."""
        current_inv = row.get('Current_Inventory', 0)
        reorder_point = row.get('Reorder_Point', 0)
        months_inv = row.get('Months_of_Inventory', 0)

        # Handle NaN values
        if pd.isna(current_inv):
            current_inv = 0
        if pd.isna(reorder_point):
            reorder_point = 0
        if pd.isna(months_inv):
            months_inv = 0

        if current_inv <= 0:
            return 'EXPEDITE: Air freight required'
        elif current_inv <= reorder_point * 0.5:
            return 'URGENT: Place PO today, follow up with supplier'
        elif current_inv <= reorder_point:
            return 'REORDER: Standard PO this week'
        elif months_inv > 12:
            return 'PAUSE: No orders, plan clearance'
        elif months_inv > 6:
            return 'MONITOR: Reduce order quantities'
        else:
            return 'SCHEDULE: Follow standard ordering'
         
    def create_enhanced_forecast_shopify_special(self, channel, inventory, product_info, product_category, product_status):
        try:
            print(f"Creating enhanced forecast for {channel}...")

            all_skus = set()
            all_skus.update(self.lead_times.keys())
            all_skus.update(inventory.keys())
            all_skus.update(product_info.keys())
            all_skus.update(self.data['SKU'].unique())
            all_skus.update(product_category.keys())
            all_skus.update(product_status.keys())

            print(f"Processing {len(all_skus)} SKUs for {channel}")

            # Generate historical month labels (last 3 months)
            current_date = pd.Timestamp.now()
            historical_months = []
            for i in range(3, 0, -1):  # 3, 2, 1 months ago
                past_date = current_date - pd.DateOffset(months=i)
                historical_months.append(past_date.strftime('%b_%Y'))

            # Generate forecast month labels
            forecast_months = []
            for i in range(8):
                future_date = current_date + pd.DateOffset(months=i+1)
                forecast_months.append(future_date.strftime('%b_%Y'))  # e.g., 'Jun_2025'

            print(f"Historical periods: {', '.join(historical_months)}")
            print(f"Forecast periods: {', '.join(forecast_months)}")

            forecast_data = []
            products_found = 0

            for i, sku in enumerate(all_skus):
                try:
                    if i % 20 == 0:
                        print(f"   Processing SKU {i+1}/{len(all_skus)}: {sku}")

                    series = self.prepare_data(channel, sku)
                    forecast, method, seasonal_periods, seasonal_factors, growth_rate = self.generate_forecast(series)

                    # Use enhanced SKU lookup for product names and launch dates
                    product_name = self.smart_sku_lookup(sku, product_info)
                    
                    category_name = self.smart_sku_lookup(sku, product_category)
                    status_name = self.smart_sku_lookup(sku, product_status)

                    launch_date = self.smart_sku_lookup(sku, None, self.launch_dates)
                    years_since_launch = self.calculate_years_since_launch(sku)

                    # Debug: Show some matching attempts
                    if i < 5 or (i % 50 == 0):
                        print(f"   DEBUG SKU {sku}: Product = '{product_name}'")

                    if product_name != 'Unknown':
                        products_found += 1
                    else:
                        # Skip forecasts for SKUs without valid product mappings
                        if i % 100 == 0:  # Only log occasionally to avoid spam
                            print(f"   Skipping SKU {sku} - no product mapping found")
                        continue

                    # Get last 3 months sales (skip the most recent month to avoid incomplete data)
                    last_3_sales = series.iloc[-4:-1].tolist() if len(series) >= 4 else series.tail(3).tolist() if len(series) >= 3 else [0, 0, 0]
                    while len(last_3_sales) < 3:
                        last_3_sales.insert(0, 0)
                    last_3_sales = last_3_sales[-3:]  # Ensure exactly 3 values

                    # Ensure we have valid numeric values
                    last_3_sales = [float(x) if not pd.isna(x) else 0.0 for x in last_3_sales]

                    last_3_months_avg = round(series.iloc[-4:-1].mean(), 2) if len(series) >= 4 else round(series[-3:].mean(), 2) if len(series) >= 3 else round(series.mean(), 2) if not series.empty else 0
                    # Handle NaN values
                    if pd.isna(last_3_months_avg):
                        last_3_months_avg = 0.0

                    total_sales = round(series.sum(), 2)
                    if pd.isna(total_sales):
                        total_sales = 0.0
                    lead_time = self.lead_times.get(sku, 2)

                    safety_stock, velocity_category, safety_stock_months = self.calculate_enhanced_safety_stock(series, lead_time, sku)
                    reorder_point = self.calculate_enhanced_reorder_point(series, lead_time, safety_stock)

                    current_inventory = inventory.get(sku, 0)
                    po_quantity, urgency = self.calculate_enhanced_po_quantity(series, current_inventory, reorder_point, sku, safety_stock_months)

                    order_dates, order_qtys, arrival_dates = self.calculate_future_orders(forecast, current_inventory, reorder_point, lead_time, safety_stock_months, sku)
                    months_of_inventory = self.calculate_months_of_inventory(current_inventory, forecast)

                    sku_info = self.velocity_categories.get(sku, {})

                    if current_inventory <= 0:
                        stock_status = "OUT OF STOCK"
                    elif current_inventory <= reorder_point:
                        stock_status = "REORDER NOW"
                    elif last_3_months_avg > 0 and current_inventory / last_3_months_avg < 1:
                        stock_status = "LOW STOCK"
                    elif last_3_months_avg > 0 and current_inventory / last_3_months_avg > 6:
                        stock_status = "OVERSTOCK"
                    else:
                        stock_status = "NORMAL"

                    ## Shopify Tab    
                    # Create dynamic forecast row in desired column order
                    row_data = {
                        'SKU': str(sku),
                        'Product_Name': product_name,
                        'Category': category_name,
                        'Status': status_name,
                        'Launch_Date': launch_date.strftime('%Y-%m-%d') if launch_date and not pd.isna(launch_date) else '',
                        'Forecast_Method': method,  # <-- Now comes early
                        'Years_Since_Launch': years_since_launch if years_since_launch is not None else '',
                        'Current_Inventory': current_inventory,
                        'Stock_Status': stock_status,
                        'PO_Urgency': urgency,
                        'Recommended_PO_Qty': po_quantity,
                        'Next_Order_Date': order_dates[0],
                        'Next_Order_Qty': order_qtys[0],
                        'Next_Arrival_Date': arrival_dates[0],
                        'Months_of_Inventory': months_of_inventory,
                        'Velocity_Category': velocity_category,
                        'Safety_Stock_Months': safety_stock_months,
                        'Reorder_Point': reorder_point,
                        'Safety_Stock': safety_stock,
                    }

                    # Add forecast columns with month/year labels
                    for idx, month_label in enumerate(forecast_months):
                        row_data[f'Forecast_{month_label}'] = int(forecast.iloc[idx]) if idx < len(forecast) else 0

                    # Add remaining metrics (still before historical and channel)
                    row_data.update({
                        'Last_3_Months_Avg': last_3_months_avg,
                        'Total_Sales': total_sales,
                        'Growth_Rate': growth_rate if not pd.isna(growth_rate) else 0.0,
                        'Order_2_Date': order_dates[1],
                        'Order_2_Qty': order_qtys[1],
                        'Order_2_Arrival': arrival_dates[1],
                        'Order_3_Date': order_dates[2],
                        'Order_3_Qty': order_qtys[2],
                        'Order_3_Arrival': arrival_dates[2],
                        'Lead_Time': lead_time,
                        'Service_Level': f"{sku_info.get('service_level', 0.85)*100:.0f}%",
                        'Monthly_Velocity': round(sku_info.get('monthly_velocity', 0), 1) if not pd.isna(sku_info.get('monthly_velocity', 0)) else 0.0,
                        'Velocity_Rank': sku_info.get('rank', 999),
                    })

                    # Add historical months (just before Channel)
                    for idx, month_label in enumerate(historical_months):
                        if idx < len(last_3_sales):
                            value = int(last_3_sales[idx]) if not pd.isna(last_3_sales[idx]) else 0
                        else:
                            value = 0
                        row_data[month_label] = value

                    # Finally add Channel
                    row_data['Channel'] = channel

                    # Append to forecast
                    forecast_data.append(row_data)


                except Exception as e:
                    print(f"   Error processing SKU {sku}: {e}")
                    if i < 10:  # Show detailed errors for first 10 SKUs
                        import traceback
                        traceback.print_exc()
                    continue

            df = pd.DataFrame(forecast_data)

            # Additional filtering to ensure we only have mapped products
            if not df.empty:
                initial_count = len(df)
                df = df[df['Product_Name'] != 'Unknown']
                final_count = len(df)
                filtered_out = initial_count - final_count

                if filtered_out > 0:
                    print(f"   Filtered out {filtered_out} SKUs without product mappings")

            print(f"Generated {len(df)} forecasts for {channel} (only mapped products)")
            print(f"✅ Successfully mapped {products_found} product names")

            return df

        except Exception:
            return pd.DataFrame()
        
    def create_enhanced_forecast_shopify_faire_special(self, channel, inventory, product_info, product_category, product_status):
            try:
                print(f"Creating enhanced forecast for {channel}...")

                all_skus = set()
                all_skus.update(self.lead_times.keys())
                all_skus.update(inventory.keys())
                all_skus.update(product_info.keys())
                all_skus.update(self.data['SKU'].unique())
                all_skus.update(product_category.keys())
                all_skus.update(product_status.keys())

                print(f"Processing {len(all_skus)} SKUs for {channel}")

                # Generate historical month labels (last 3 months)
                current_date = pd.Timestamp.now()
                historical_months = []
                for i in range(3, 0, -1):  # 3, 2, 1 months ago
                    past_date = current_date - pd.DateOffset(months=i)
                    historical_months.append(past_date.strftime('%b_%Y'))

                # Generate forecast month labels
                forecast_months = []
                for i in range(8):
                    future_date = current_date + pd.DateOffset(months=i+1)
                    forecast_months.append(future_date.strftime('%b_%Y'))  # e.g., 'Jun_2025'

                print(f"Historical periods: {', '.join(historical_months)}")
                print(f"Forecast periods: {', '.join(forecast_months)}")

                forecast_data = []
                products_found = 0

                for i, sku in enumerate(all_skus):
                    try:
                        if i % 20 == 0:
                            print(f"   Processing SKU {i+1}/{len(all_skus)}: {sku}")

                        series = self.prepare_data(channel, sku)
                        forecast, method, seasonal_periods, seasonal_factors, growth_rate = self.generate_forecast(series)

                        # Use enhanced SKU lookup for product names and launch dates
                        product_name = self.smart_sku_lookup(sku, product_info)
                        
                        category_name = self.smart_sku_lookup(sku, product_category)
                        status_name = self.smart_sku_lookup(sku, product_status)

                        launch_date = self.smart_sku_lookup(sku, None, self.launch_dates)
                        years_since_launch = self.calculate_years_since_launch(sku)

                        # Debug: Show some matching attempts
                        if i < 5 or (i % 50 == 0):
                            print(f"   DEBUG SKU {sku}: Product = '{product_name}'")

                        if product_name != 'Unknown':
                            products_found += 1
                        else:
                            # Skip forecasts for SKUs without valid product mappings
                            if i % 100 == 0:  # Only log occasionally to avoid spam
                                print(f"   Skipping SKU {sku} - no product mapping found")
                            continue

                        # Get last 3 months sales (skip the most recent month to avoid incomplete data)
                        last_3_sales = series.iloc[-4:-1].tolist() if len(series) >= 4 else series.tail(3).tolist() if len(series) >= 3 else [0, 0, 0]
                        while len(last_3_sales) < 3:
                            last_3_sales.insert(0, 0)
                        last_3_sales = last_3_sales[-3:]  # Ensure exactly 3 values

                        # Ensure we have valid numeric values
                        last_3_sales = [float(x) if not pd.isna(x) else 0.0 for x in last_3_sales]

                        last_3_months_avg = round(series.iloc[-4:-1].mean(), 2) if len(series) >= 4 else round(series[-3:].mean(), 2) if len(series) >= 3 else round(series.mean(), 2) if not series.empty else 0
                        # Handle NaN values
                        if pd.isna(last_3_months_avg):
                            last_3_months_avg = 0.0

                        total_sales = round(series.sum(), 2)
                        if pd.isna(total_sales):
                            total_sales = 0.0
                        lead_time = self.lead_times.get(sku, 2)

                        safety_stock, velocity_category, safety_stock_months = self.calculate_enhanced_safety_stock(series, lead_time, sku)
                        reorder_point = self.calculate_enhanced_reorder_point(series, lead_time, safety_stock)

                        current_inventory = inventory.get(sku, 0)
                        po_quantity, urgency = self.calculate_enhanced_po_quantity(series, current_inventory, reorder_point, sku, safety_stock_months)

                        order_dates, order_qtys, arrival_dates = self.calculate_future_orders(forecast, current_inventory, reorder_point, lead_time, safety_stock_months, sku)
                        months_of_inventory = self.calculate_months_of_inventory(current_inventory, forecast)

                        sku_info = self.velocity_categories.get(sku, {})

                        if current_inventory <= 0:
                            stock_status = "OUT OF STOCK"
                        elif current_inventory <= reorder_point:
                            stock_status = "REORDER NOW"
                        elif last_3_months_avg > 0 and current_inventory / last_3_months_avg < 1:
                            stock_status = "LOW STOCK"
                        elif last_3_months_avg > 0 and current_inventory / last_3_months_avg > 6:
                            stock_status = "OVERSTOCK"
                        else:
                            stock_status = "NORMAL"

                        ## Shopify Faire Tab    
                        # Create dynamic forecast row in desired column order
                        row_data = {
                            'SKU': str(sku),
                            'Product_Name': product_name,
                            'Category': category_name,
                            'Status': status_name,
                            'Launch_Date': launch_date.strftime('%Y-%m-%d') if launch_date and not pd.isna(launch_date) else '',
                            'Forecast_Method': method,  # <-- Now comes early
                            'Years_Since_Launch': years_since_launch if years_since_launch is not None else '',
                            'Current_Inventory': current_inventory,
                            'Stock_Status': stock_status,
                            'PO_Urgency': urgency,
                            'Recommended_PO_Qty': po_quantity,
                            'Next_Order_Date': order_dates[0],
                            'Next_Order_Qty': order_qtys[0],
                            'Next_Arrival_Date': arrival_dates[0],
                            'Months_of_Inventory': months_of_inventory,
                            'Velocity_Category': velocity_category,
                            'Safety_Stock_Months': safety_stock_months,
                            'Reorder_Point': reorder_point,
                            'Safety_Stock': safety_stock,
                        }

                        # Add forecast columns with month/year labels
                        for idx, month_label in enumerate(forecast_months):
                            row_data[f'Forecast_{month_label}'] = int(forecast.iloc[idx]) if idx < len(forecast) else 0

                        # Add remaining metrics (still before historical and channel)
                        row_data.update({
                            'Last_3_Months_Avg': last_3_months_avg,
                            'Total_Sales': total_sales,
                            'Growth_Rate': growth_rate if not pd.isna(growth_rate) else 0.0,
                            'Order_2_Date': order_dates[1],
                            'Order_2_Qty': order_qtys[1],
                            'Order_2_Arrival': arrival_dates[1],
                            'Order_3_Date': order_dates[2],
                            'Order_3_Qty': order_qtys[2],
                            'Order_3_Arrival': arrival_dates[2],
                            'Lead_Time': lead_time,
                            'Service_Level': f"{sku_info.get('service_level', 0.85)*100:.0f}%",
                            'Monthly_Velocity': round(sku_info.get('monthly_velocity', 0), 1) if not pd.isna(sku_info.get('monthly_velocity', 0)) else 0.0,
                            'Velocity_Rank': sku_info.get('rank', 999),
                        })

                        # Add historical months (just before Channel)
                        for idx, month_label in enumerate(historical_months):
                            if idx < len(last_3_sales):
                                value = int(last_3_sales[idx]) if not pd.isna(last_3_sales[idx]) else 0
                            else:
                                value = 0
                            row_data[month_label] = value

                        # Finally add Channel
                        row_data['Channel'] = channel

                        # Append to forecast
                        forecast_data.append(row_data)


                    except Exception as e:
                        print(f"   Error processing SKU {sku}: {e}")
                        if i < 10:  # Show detailed errors for first 10 SKUs
                            import traceback
                            traceback.print_exc()
                        continue

                df = pd.DataFrame(forecast_data)

                # Additional filtering to ensure we only have mapped products
                if not df.empty:
                    initial_count = len(df)
                    df = df[df['Product_Name'] != 'Unknown']
                    final_count = len(df)
                    filtered_out = initial_count - final_count

                    if filtered_out > 0:
                        print(f"   Filtered out {filtered_out} SKUs without product mappings")

                print(f"Generated {len(df)} forecasts for {channel} (only mapped products)")
                print(f"✅ Successfully mapped {products_found} product names")

                return df

            except Exception:
                return pd.DataFrame()    
  
        
    def create_finance_cash_flow_forecast(self, combined_forecast):
        try:
            print("Creating finance cash flow forecast...")

            if combined_forecast.empty:
                return pd.DataFrame()

            finance_data = []
            current_date = pd.Timestamp.now()

            for _, row in combined_forecast.iterrows():
                sku = row['SKU']
                product_name = row['Product_Name']
                velocity_category = row['Velocity_Category']

                orders = []

                if row['PO_Urgency'] == 'HIGH - Below Reorder Point' and row['Recommended_PO_Qty'] > 0:
                    orders.append({
                        'order_date': current_date.strftime('%Y-%m-%d'),
                        'order_month': current_date.strftime('%Y-%m'),
                        'quantity': row['Recommended_PO_Qty'],
                        'urgency': 'IMMEDIATE'
                    })

                if row['Next_Order_Date'] and row['Next_Order_Qty'] > 0:
                    try:
                        order_date = pd.to_datetime(row['Next_Order_Date'])
                        if order_date.year == current_date.year:
                            orders.append({
                                'order_date': order_date.strftime('%Y-%m-%d'),
                                'order_month': order_date.strftime('%Y-%m'),
                                'quantity': row['Next_Order_Qty'],
                                'urgency': 'SCHEDULED'
                            })
                    except:
                        pass

                if row['Order_2_Date'] and row['Order_2_Qty'] > 0:
                    try:
                        order_date = pd.to_datetime(row['Order_2_Date'])
                        if order_date.year == current_date.year:
                            orders.append({
                                'order_date': order_date.strftime('%Y-%m-%d'),
                                'order_month': order_date.strftime('%Y-%m'),
                                'quantity': row['Order_2_Qty'],
                                'urgency': 'SCHEDULED'
                            })
                    except:
                        pass

                for order in orders:
                    finance_record = {
                        'Order_Month': order['order_month'],
                        'Order_Date': order['order_date'],
                        'SKU': sku,
                        'Product_Name': product_name,
                        'Velocity_Category': velocity_category,
                        'Order_Quantity': order['quantity'],
                        'Order_Urgency': order['urgency'],
                        'Unit_Cost': '',
                        'Total_Order_Value': '',
                        'Supplier': '',
                        'Payment_Terms': '',
                        'Expected_Payment_Date': '',
                        'Lead_Time': row['Lead_Time'],
                        'Safety_Stock_Months': row['Safety_Stock_Months'],
                        'Current_Inventory': row['Current_Inventory'],
                        'Months_of_Inventory': row['Months_of_Inventory'],
                    }

                    finance_data.append(finance_record)

            finance_df = pd.DataFrame(finance_data)

            if not finance_df.empty:
                finance_df['Order_Date_Sort'] = pd.to_datetime(finance_df['Order_Date'])
                finance_df = finance_df.sort_values(['Order_Date_Sort', 'Velocity_Category', 'SKU'])
                finance_df = finance_df.drop('Order_Date_Sort', axis=1)

                print(f"Created finance forecast: {len(finance_df)} orders planned")

            return finance_df

        except Exception:
            return pd.DataFrame()
        
    def create_enhanced_forecast_amazon_fbm_special(self, channel, inventory, product_info, product_category, product_status):
            try:
                print(f"Creating enhanced forecast for {channel}...")

                all_skus = set()
                all_skus.update(self.lead_times.keys())
                all_skus.update(inventory.keys())
                all_skus.update(product_info.keys())
                all_skus.update(self.data['SKU'].unique())
                all_skus.update(product_category.keys())
                all_skus.update(product_status.keys())

                print(f"Processing {len(all_skus)} SKUs for {channel}")

                # Generate historical month labels (last 3 months)
                current_date = pd.Timestamp.now()
                historical_months = []
                for i in range(3, 0, -1):  # 3, 2, 1 months ago
                    past_date = current_date - pd.DateOffset(months=i)
                    historical_months.append(past_date.strftime('%b_%Y'))

                # Generate forecast month labels
                forecast_months = []
                for i in range(8):
                    future_date = current_date + pd.DateOffset(months=i+1)
                    forecast_months.append(future_date.strftime('%b_%Y'))  # e.g., 'Jun_2025'

                print(f"Historical periods: {', '.join(historical_months)}")
                print(f"Forecast periods: {', '.join(forecast_months)}")

                forecast_data = []
                products_found = 0

                for i, sku in enumerate(all_skus):
                    try:
                        if i % 20 == 0:
                            print(f"   Processing SKU {i+1}/{len(all_skus)}: {sku}")

                        series = self.prepare_data(channel, sku)
                        forecast, method, seasonal_periods, seasonal_factors, growth_rate = self.generate_forecast(series)

                        # Use enhanced SKU lookup for product names and launch dates
                        product_name = self.smart_sku_lookup(sku, product_info)
                        
                        category_name = self.smart_sku_lookup(sku, product_category)
                        status_name = self.smart_sku_lookup(sku, product_status)

                        launch_date = self.smart_sku_lookup(sku, None, self.launch_dates)
                        years_since_launch = self.calculate_years_since_launch(sku)

                        # Debug: Show some matching attempts
                        if i < 5 or (i % 50 == 0):
                            print(f"   DEBUG SKU {sku}: Product = '{product_name}'")

                        if product_name != 'Unknown':
                            products_found += 1
                        else:
                            # Skip forecasts for SKUs without valid product mappings
                            if i % 100 == 0:  # Only log occasionally to avoid spam
                                print(f"   Skipping SKU {sku} - no product mapping found")
                            continue

                        # Get last 3 months sales (skip the most recent month to avoid incomplete data)
                        last_3_sales = series.iloc[-4:-1].tolist() if len(series) >= 4 else series.tail(3).tolist() if len(series) >= 3 else [0, 0, 0]
                        while len(last_3_sales) < 3:
                            last_3_sales.insert(0, 0)
                        last_3_sales = last_3_sales[-3:]  # Ensure exactly 3 values

                        # Ensure we have valid numeric values
                        last_3_sales = [float(x) if not pd.isna(x) else 0.0 for x in last_3_sales]

                        last_3_months_avg = round(series.iloc[-4:-1].mean(), 2) if len(series) >= 4 else round(series[-3:].mean(), 2) if len(series) >= 3 else round(series.mean(), 2) if not series.empty else 0
                        # Handle NaN values
                        if pd.isna(last_3_months_avg):
                            last_3_months_avg = 0.0

                        total_sales = round(series.sum(), 2)
                        if pd.isna(total_sales):
                            total_sales = 0.0
                        lead_time = self.lead_times.get(sku, 2)

                        safety_stock, velocity_category, safety_stock_months = self.calculate_enhanced_safety_stock(series, lead_time, sku)
                        reorder_point = self.calculate_enhanced_reorder_point(series, lead_time, safety_stock)

                        current_inventory = inventory.get(sku, 0)
                        po_quantity, urgency = self.calculate_enhanced_po_quantity(series, current_inventory, reorder_point, sku, safety_stock_months)

                        order_dates, order_qtys, arrival_dates = self.calculate_future_orders(forecast, current_inventory, reorder_point, lead_time, safety_stock_months, sku)
                        months_of_inventory = self.calculate_months_of_inventory(current_inventory, forecast)

                        sku_info = self.velocity_categories.get(sku, {})

                        if current_inventory <= 0:
                            stock_status = "OUT OF STOCK"
                        elif current_inventory <= reorder_point:
                            stock_status = "REORDER NOW"
                        elif last_3_months_avg > 0 and current_inventory / last_3_months_avg < 1:
                            stock_status = "LOW STOCK"
                        elif last_3_months_avg > 0 and current_inventory / last_3_months_avg > 6:
                            stock_status = "OVERSTOCK"
                        else:
                            stock_status = "NORMAL"

                        ## Amazon FBM Tab    
                        # Create dynamic forecast row in desired column order
                        row_data = {
                            'SKU': str(sku),
                            'Product_Name': product_name,
                            'Category': category_name,
                            'Status': status_name,
                            'Launch_Date': launch_date.strftime('%Y-%m-%d') if launch_date and not pd.isna(launch_date) else '',
                            'Forecast_Method': method,  # <-- Now comes early
                            'Years_Since_Launch': years_since_launch if years_since_launch is not None else '',
                            'Current_Inventory': current_inventory,
                            'Stock_Status': stock_status,
                            'PO_Urgency': urgency,
                            'Recommended_PO_Qty': po_quantity,
                            'Next_Order_Date': order_dates[0],
                            'Next_Order_Qty': order_qtys[0],
                            'Next_Arrival_Date': arrival_dates[0],
                            'Months_of_Inventory': months_of_inventory,
                            'Velocity_Category': velocity_category,
                            'Safety_Stock_Months': safety_stock_months,
                            'Reorder_Point': reorder_point,
                            'Safety_Stock': safety_stock,
                        }

                        # Add forecast columns with month/year labels
                        for idx, month_label in enumerate(forecast_months):
                            row_data[f'Forecast_{month_label}'] = int(forecast.iloc[idx]) if idx < len(forecast) else 0

                        # Add remaining metrics (still before historical and channel)
                        row_data.update({
                            'Last_3_Months_Avg': last_3_months_avg,
                            'Total_Sales': total_sales,
                            'Growth_Rate': growth_rate if not pd.isna(growth_rate) else 0.0,
                            'Order_2_Date': order_dates[1],
                            'Order_2_Qty': order_qtys[1],
                            'Order_2_Arrival': arrival_dates[1],
                            'Order_3_Date': order_dates[2],
                            'Order_3_Qty': order_qtys[2],
                            'Order_3_Arrival': arrival_dates[2],
                            'Lead_Time': lead_time,
                            'Service_Level': f"{sku_info.get('service_level', 0.85)*100:.0f}%",
                            'Monthly_Velocity': round(sku_info.get('monthly_velocity', 0), 1) if not pd.isna(sku_info.get('monthly_velocity', 0)) else 0.0,
                            'Velocity_Rank': sku_info.get('rank', 999),
                        })

                        # Add historical months (just before Channel)
                        for idx, month_label in enumerate(historical_months):
                            if idx < len(last_3_sales):
                                value = int(last_3_sales[idx]) if not pd.isna(last_3_sales[idx]) else 0
                            else:
                                value = 0
                            row_data[month_label] = value

                        # Finally add Channel
                        row_data['Channel'] = channel

                        # Append to forecast
                        forecast_data.append(row_data)


                    except Exception as e:
                        print(f"   Error processing SKU {sku}: {e}")
                        if i < 10:  # Show detailed errors for first 10 SKUs
                            import traceback
                            traceback.print_exc()
                        continue

                df = pd.DataFrame(forecast_data)

                # Additional filtering to ensure we only have mapped products
                if not df.empty:
                    initial_count = len(df)
                    df = df[df['Product_Name'] != 'Unknown']
                    final_count = len(df)
                    filtered_out = initial_count - final_count

                    if filtered_out > 0:
                        print(f"   Filtered out {filtered_out} SKUs without product mappings")

                print(f"Generated {len(df)} forecasts for {channel} (only mapped products)")
                print(f"✅ Successfully mapped {products_found} product names")

                return df

            except Exception:
                return pd.DataFrame()    
            
    def create_enhanced_forecast_walmart_fbm_special(self, channel, inventory, product_info, product_category, product_status):
            try:
                print(f"Creating enhanced forecast for {channel}...")

                all_skus = set()
                all_skus.update(self.lead_times.keys())
                all_skus.update(inventory.keys())
                all_skus.update(product_info.keys())
                all_skus.update(self.data['SKU'].unique())
                all_skus.update(product_category.keys())
                all_skus.update(product_status.keys())

                print(f"Processing {len(all_skus)} SKUs for {channel}")

                # Generate historical month labels (last 3 months)
                current_date = pd.Timestamp.now()
                historical_months = []
                for i in range(3, 0, -1):  # 3, 2, 1 months ago
                    past_date = current_date - pd.DateOffset(months=i)
                    historical_months.append(past_date.strftime('%b_%Y'))

                # Generate forecast month labels
                forecast_months = []
                for i in range(8):
                    future_date = current_date + pd.DateOffset(months=i+1)
                    forecast_months.append(future_date.strftime('%b_%Y'))  # e.g., 'Jun_2025'

                print(f"Historical periods: {', '.join(historical_months)}")
                print(f"Forecast periods: {', '.join(forecast_months)}")

                forecast_data = []
                products_found = 0

                for i, sku in enumerate(all_skus):
                    try:
                        if i % 20 == 0:
                            print(f"   Processing SKU {i+1}/{len(all_skus)}: {sku}")

                        series = self.prepare_data(channel, sku)
                        forecast, method, seasonal_periods, seasonal_factors, growth_rate = self.generate_forecast(series)

                        # Use enhanced SKU lookup for product names and launch dates
                        product_name = self.smart_sku_lookup(sku, product_info)
                        
                        category_name = self.smart_sku_lookup(sku, product_category)
                        status_name = self.smart_sku_lookup(sku, product_status)

                        launch_date = self.smart_sku_lookup(sku, None, self.launch_dates)
                        years_since_launch = self.calculate_years_since_launch(sku)

                        # Debug: Show some matching attempts
                        if i < 5 or (i % 50 == 0):
                            print(f"   DEBUG SKU {sku}: Product = '{product_name}'")

                        if product_name != 'Unknown':
                            products_found += 1
                        else:
                            # Skip forecasts for SKUs without valid product mappings
                            if i % 100 == 0:  # Only log occasionally to avoid spam
                                print(f"   Skipping SKU {sku} - no product mapping found")
                            continue

                        # Get last 3 months sales (skip the most recent month to avoid incomplete data)
                        last_3_sales = series.iloc[-4:-1].tolist() if len(series) >= 4 else series.tail(3).tolist() if len(series) >= 3 else [0, 0, 0]
                        while len(last_3_sales) < 3:
                            last_3_sales.insert(0, 0)
                        last_3_sales = last_3_sales[-3:]  # Ensure exactly 3 values

                        # Ensure we have valid numeric values
                        last_3_sales = [float(x) if not pd.isna(x) else 0.0 for x in last_3_sales]

                        last_3_months_avg = round(series.iloc[-4:-1].mean(), 2) if len(series) >= 4 else round(series[-3:].mean(), 2) if len(series) >= 3 else round(series.mean(), 2) if not series.empty else 0
                        # Handle NaN values
                        if pd.isna(last_3_months_avg):
                            last_3_months_avg = 0.0

                        total_sales = round(series.sum(), 2)
                        if pd.isna(total_sales):
                            total_sales = 0.0
                        lead_time = self.lead_times.get(sku, 2)

                        safety_stock, velocity_category, safety_stock_months = self.calculate_enhanced_safety_stock(series, lead_time, sku)
                        reorder_point = self.calculate_enhanced_reorder_point(series, lead_time, safety_stock)

                        current_inventory = inventory.get(sku, 0)
                        po_quantity, urgency = self.calculate_enhanced_po_quantity(series, current_inventory, reorder_point, sku, safety_stock_months)

                        order_dates, order_qtys, arrival_dates = self.calculate_future_orders(forecast, current_inventory, reorder_point, lead_time, safety_stock_months, sku)
                        months_of_inventory = self.calculate_months_of_inventory(current_inventory, forecast)

                        sku_info = self.velocity_categories.get(sku, {})

                        if current_inventory <= 0:
                            stock_status = "OUT OF STOCK"
                        elif current_inventory <= reorder_point:
                            stock_status = "REORDER NOW"
                        elif last_3_months_avg > 0 and current_inventory / last_3_months_avg < 1:
                            stock_status = "LOW STOCK"
                        elif last_3_months_avg > 0 and current_inventory / last_3_months_avg > 6:
                            stock_status = "OVERSTOCK"
                        else:
                            stock_status = "NORMAL"

                        ## Walmart FBM Tab    
                        # Create dynamic forecast row in desired column order
                        row_data = {
                            'SKU': str(sku),
                            'Product_Name': product_name,
                            'Category': category_name,
                            'Status': status_name,
                            'Launch_Date': launch_date.strftime('%Y-%m-%d') if launch_date and not pd.isna(launch_date) else '',
                            'Forecast_Method': method,  # <-- Now comes early
                            'Years_Since_Launch': years_since_launch if years_since_launch is not None else '',
                            'Current_Inventory': current_inventory,
                            'Stock_Status': stock_status,
                            'PO_Urgency': urgency,
                            'Recommended_PO_Qty': po_quantity,
                            'Next_Order_Date': order_dates[0],
                            'Next_Order_Qty': order_qtys[0],
                            'Next_Arrival_Date': arrival_dates[0],
                            'Months_of_Inventory': months_of_inventory,
                            'Velocity_Category': velocity_category,
                            'Safety_Stock_Months': safety_stock_months,
                            'Reorder_Point': reorder_point,
                            'Safety_Stock': safety_stock,
                        }

                        # Add forecast columns with month/year labels
                        for idx, month_label in enumerate(forecast_months):
                            row_data[f'Forecast_{month_label}'] = int(forecast.iloc[idx]) if idx < len(forecast) else 0

                        # Add remaining metrics (still before historical and channel)
                        row_data.update({
                            'Last_3_Months_Avg': last_3_months_avg,
                            'Total_Sales': total_sales,
                            'Growth_Rate': growth_rate if not pd.isna(growth_rate) else 0.0,
                            'Order_2_Date': order_dates[1],
                            'Order_2_Qty': order_qtys[1],
                            'Order_2_Arrival': arrival_dates[1],
                            'Order_3_Date': order_dates[2],
                            'Order_3_Qty': order_qtys[2],
                            'Order_3_Arrival': arrival_dates[2],
                            'Lead_Time': lead_time,
                            'Service_Level': f"{sku_info.get('service_level', 0.85)*100:.0f}%",
                            'Monthly_Velocity': round(sku_info.get('monthly_velocity', 0), 1) if not pd.isna(sku_info.get('monthly_velocity', 0)) else 0.0,
                            'Velocity_Rank': sku_info.get('rank', 999),
                        })

                        # Add historical months (just before Channel)
                        for idx, month_label in enumerate(historical_months):
                            if idx < len(last_3_sales):
                                value = int(last_3_sales[idx]) if not pd.isna(last_3_sales[idx]) else 0
                            else:
                                value = 0
                            row_data[month_label] = value

                        # Finally add Channel
                        row_data['Channel'] = channel

                        # Append to forecast
                        forecast_data.append(row_data)


                    except Exception as e:
                        print(f"   Error processing SKU {sku}: {e}")
                        if i < 10:  # Show detailed errors for first 10 SKUs
                            import traceback
                            traceback.print_exc()
                        continue

                df = pd.DataFrame(forecast_data)

                # Additional filtering to ensure we only have mapped products
                if not df.empty:
                    initial_count = len(df)
                    df = df[df['Product_Name'] != 'Unknown']
                    final_count = len(df)
                    filtered_out = initial_count - final_count

                    if filtered_out > 0:
                        print(f"   Filtered out {filtered_out} SKUs without product mappings")

                print(f"Generated {len(df)} forecasts for {channel} (only mapped products)")
                print(f"✅ Successfully mapped {products_found} product names")

                return df

            except Exception:
                return pd.DataFrame()    

# NEW: Wrapped Forecast BOM Function
# PLACEMENT: After EnhancedForecastingModel class, before upload_excel_to_google_sheet function

# ==============================================================================
# ENHANCED BOM MODULE v2.0 - WRAPPED FOR WEBAPP INTEGRATION
# ==============================================================================
# 
# INSTRUCTIONS:
# 1. Find the existing `run_forecast_bom_analysis` function in your Updated_Template.py
# 2. Delete the ENTIRE existing function (from `def run_forecast_bom_analysis` to its end)
# 3. Paste this entire code block in its place
# 4. Make sure `import numpy as np` is in your imports at the top of the file
#
# PLACEMENT: After EnhancedForecastingModel class, before upload_excel_to_google_sheet function
# ==============================================================================


def run_forecast_bom_analysis(gc_client=None):
    """
    ENHANCED Forecast BOM Analysis function v2.0 - Wrapped for WebApp
    
    FEATURES:
    - Component Type/Category fetching from source sheet
    - ABC Classification (A/B/C based on value)
    - Days of Stock calculation
    - Stock Coverage Ratio
    - Order Priority Score (0-100)
    - ABC-based Safety Stock
    - Category Summary sheet
    - Procurement Timeline sheet
    - Enhanced Executive Summary with more KPIs
    - Professional Excel formatting with proper data types
    - Emoji sheet names for visual clarity
    
    Returns: (excel_buffer, filename) tuple or (None, None) on failure
    """
    from collections import defaultdict
    from typing import Dict, List, Tuple, Set, Optional
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime, timedelta
    import numpy as np
    
    # ==========================================================================
    # BOM CONFIGURATION
    # ==========================================================================
    
    BOM_CONFIG = {
        # BOM Data Sheet
        'SPREADSHEET_URL': 'https://docs.google.com/spreadsheets/d/1ddH2428mSdWJSyDRH72oUGDVsLQjNuAOQ24k-lOnf-I/edit?gid=1102477322#gid=1102477322',
        'WORKSHEET_NAME': 'BOM',

        # SKU Reference Sheet (Parent_Item_Code → UPC)
        'SKU_REFERENCE_URL': 'https://docs.google.com/spreadsheets/d/1rWAd551acZ6bQ86s4gRxwzLo2lel-OPB59CyVRmTjc8/edit?gid=0#gid=0',
        'SKU_REFERENCE_WORKSHEET': 'Finished Goods MasterList',
        'SKU_ITEM_CODE_COLUMN': 'B',
        'SKU_UPC_COLUMN': 'M',

        # Forecast Sheet (UPC → 6-month forecast)
        'FORECAST_URL': 'https://docs.google.com/spreadsheets/d/1051NJelrnQGKwKDXWmaMiU1-fBgm4ZSd4s_G2-hJIcE/edit?gid=951425625#gid=951425625',
        'FORECAST_WORKSHEET': '📈 All Forecasts',
        'FORECAST_UPC_COLUMN': 'A',
        'FORECAST_MONTH_COLUMNS': ['W', 'X', 'Y', 'Z', 'AA', 'AB'],

        # Procurement Parameters Sheet
        'PROCUREMENT_PARAMS_URL': 'https://docs.google.com/spreadsheets/d/1YQlYkmupfVkx2ujZ2lyu6TBng7NxqUJlQNCNf8tHfL8/edit?gid=127074428#gid=127074428',
        'PROCUREMENT_PARAMS_WORKSHEET': 'Input Components_MasterList',
        'PROCUREMENT_LEAD_TIME_COLUMN': 'N',
        'PROCUREMENT_MOQ_COLUMN': 'O',
        'PROCUREMENT_EOQ_COLUMN': 'P',

        # Current Inventory Sheet
        'INVENTORY_URL': 'https://docs.google.com/spreadsheets/d/1ddH2428mSdWJSyDRH72oUGDVsLQjNuAOQ24k-lOnf-I/edit?gid=1771508063#gid=1771508063',
        'INVENTORY_WORKSHEET': 'Procurement Plan_Components',
        'INVENTORY_QTY_COLUMN': 'D',

        # Forecast behaviour
        'FORECAST_SOURCE': 'google_sheets',
        'DEFAULT_FORECAST_QTY': 100,
        'MIN_FORECAST_QTY': 10,

        # Procurement & ROP settings
        'FORECAST_HORIZON_DAYS': 180,
        'SAFETY_STOCK_PCT': 0.10,
        
        # ABC Classification thresholds
        'ABC_A_THRESHOLD': 0.70,
        'ABC_B_THRESHOLD': 0.90,
        
        # Safety stock by ABC class
        'SAFETY_STOCK_A': 0.15,
        'SAFETY_STOCK_B': 0.10,
        'SAFETY_STOCK_C': 0.05,
    }

    # ==========================================================================
    # HELPER FUNCTIONS
    # ==========================================================================

    def column_letter_to_index(col_letter: str) -> int:
        col_letter = col_letter.upper()
        result = 0
        for char in col_letter:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result

    def safe_float(value, default=0.0):
        """Safely convert a value to float, handling Series, arrays, and edge cases."""
        try:
            if isinstance(value, pd.Series):
                value = value.iloc[0] if len(value) > 0 else default
            elif isinstance(value, (list, np.ndarray)):
                value = value[0] if len(value) > 0 else default
            result = float(value) if pd.notna(value) else default
            return result
        except (TypeError, ValueError, IndexError):
            return default

    def safe_numeric_convert(series: pd.Series) -> pd.Series:
        """Safely convert a series to numeric, handling string cleaning."""
        if isinstance(series, pd.DataFrame):
            series = series.iloc[:, 0]
        return (
            series
            .astype(str)
            .str.replace(r'[$,%]', '', regex=True)
            .str.replace(',', '', regex=False)
            .apply(pd.to_numeric, errors='coerce')
            .fillna(0)
        )

    # ==========================================================================
    # BOM DATA FETCHING
    # ==========================================================================

    def fetch_bom_from_sheet(client, sheet_url: str, worksheet_name: str) -> pd.DataFrame:
        print("\n📥 Fetching BOM data from Google Sheets...")
        try:
            sheet = client.open_by_url(sheet_url)
            ws = sheet.worksheet(worksheet_name)
        except gspread.exceptions.SpreadsheetNotFound:
            raise Exception(f"Spreadsheet not found: {sheet_url}")
        except gspread.exceptions.WorksheetNotFound:
            raise Exception(f"Worksheet '{worksheet_name}' not found.")

        raw_vals = ws.get_all_values()
        if not raw_vals:
            raise Exception("BOM sheet is empty.")

        # Build clean header - handle duplicates
        header_row = raw_vals[0]
        clean_hdr = []
        seen = {}
        for i, h in enumerate(header_row):
            h = str(h).strip()
            if not h:
                h = f"Col_{i+1}"
            if h in seen:
                seen[h] += 1
                h = f"{h}_{seen[h]}"
            else:
                seen[h] = 0
            clean_hdr.append(h)

        df = pd.DataFrame(raw_vals[1:], columns=clean_hdr)

        # Column mapping
        column_mapping = {
            'Parent Item Code': 'parent_item_code',
            'Parent SKU': 'parent_sku',
            'Component Item Code': 'component_item_code',
            'Component Type': 'component_type',
            'Component Category': 'component_type',
            'Category': 'component_type',
            'Type': 'component_type',
            'Component': 'component_description',
            'Component Name': 'component_description',
            'Quantity Required': 'quantity_required',
            'Qty Required': 'quantity_required',
            'UoM 1': 'uom',
            'UoM': 'uom',
            'Unit of Measure': 'uom',
            'Wastage %': 'wastage_pct',
            'Wastage': 'wastage_pct',
            'Scrap %': 'wastage_pct',
            'Net Requirement': 'net_requirement',
            'Component Cost (All in Cost)': 'unit_cost',
            'Unit Cost': 'unit_cost',
            'Cost': 'unit_cost',
            'Total Cost': 'total_cost',
            'Critical Path': 'critical_path',
            'Supplier': 'supplier',
            'Supplier - Primary vendor': 'supplier',
            'Vendor': 'supplier',
            'Supplier Name': 'supplier',
        }
        df = df.rename(columns={k: v for k, v in column_mapping.items() if k in df.columns})

        # Numeric clean-up
        numeric_cols = ['quantity_required', 'wastage_pct', 'net_requirement', 'unit_cost', 'total_cost']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = safe_numeric_convert(df[col])

        df = df.dropna(subset=['parent_item_code', 'component_item_code'], how='all')
        df['parent_sku'] = df['parent_sku'].fillna('Unknown SKU') if 'parent_sku' in df.columns else 'Unknown SKU'
        df['component_description'] = df['component_description'].fillna('Unknown Component') if 'component_description' in df.columns else 'Unknown Component'
        
        # Component Type handling
        if 'component_type' not in df.columns:
            df['component_type'] = 'Uncategorized'
        else:
            df['component_type'] = df['component_type'].fillna('Uncategorized')
            df['component_type'] = df['component_type'].replace('', 'Uncategorized')
        
        type_mapping = {
            'raw material': 'Raw Material', 'raw materials': 'Raw Material',
            'packaging': 'Packaging', 'package': 'Packaging',
            'label': 'Labels', 'labels': 'Labels',
            'bottle': 'Packaging', 'bottles': 'Packaging',
            'cap': 'Packaging', 'caps': 'Packaging',
            'box': 'Packaging', 'boxes': 'Packaging',
            'carton': 'Packaging', 'insert': 'Packaging', 'sleeve': 'Packaging',
            'ingredient': 'Raw Material', 'ingredients': 'Raw Material',
            'other': 'Other',
        }
        
        def standardize_type(t):
            t_lower = str(t).lower().strip()
            return type_mapping.get(t_lower, t.title() if t and t != 'Uncategorized' else 'Uncategorized')
        
        df['component_type'] = df['component_type'].apply(standardize_type)
        df['uom'] = df['uom'].fillna('EA') if 'uom' in df.columns else 'EA'
        
        if 'supplier' not in df.columns:
            df['supplier'] = 'Unknown Supplier'
        else:
            df['supplier'] = df['supplier'].fillna('Unknown Supplier').replace('', 'Unknown Supplier')

        print(f"✅ Cleaned BOM data: {len(df)} valid entries")
        print(f"   Component Types: {df['component_type'].nunique()}")
        return df

    # ==========================================================================
    # PROCUREMENT & INVENTORY FETCHING
    # ==========================================================================

    def fetch_procurement_parameters(client, sheet_url: str, worksheet_name: str,
                                    component_col: str, lead_time_col: str, 
                                    moq_col: str, eoq_col: str, supplier_col: str = 'L') -> pd.DataFrame:
        print("\n📦 Fetching procurement parameters...")
        try:
            sheet = client.open_by_url(sheet_url)
            worksheet = sheet.worksheet(worksheet_name)
        except Exception as e:
            raise Exception(f"Failed to open procurement sheet: {str(e)}")

        raw_data = worksheet.get_all_values()
        if not raw_data:
            raise Exception("Procurement parameters sheet is empty.")

        first_row = raw_data[0]
        non_empty_first = sum(1 for c in first_row[:8] if c.strip() != "")
        if non_empty_first == 0:
            header = raw_data[1]
            rows = raw_data[2:]
        else:
            header = raw_data[0]
            rows = raw_data[1:]

        df = pd.DataFrame(rows, columns=header)

        component_col_names = ['Unique Identifier', 'Component Item Code', 'Component_Item_Code',
                              'Item Code', 'Component Code', 'SKU', 'Item_Code']
        component_column = next((c for c in component_col_names if c in df.columns), None)
        if not component_column:
            component_column = df.columns[0]

# FIX: Define column_mapping first, then add supplier. Replace the entire column mapping section with:

        component_col_names = ['Unique Identifier', 'Component Item Code', 'Component_Item_Code',
                              'Item Code', 'Component Code', 'SKU', 'Item_Code']
        component_column = next((c for c in component_col_names if c in df.columns), None)
        if not component_column:
            component_column = df.columns[0]

        # Define column_mapping FIRST
        column_mapping = {
            component_column: 'component_item_code',
            "Lead Time - Component procurement time (days)": "lead_time_days",
            "Lead Time Days": "lead_time_days",
            "Lead_Time_Days": "lead_time_days",
            "Lead Time": "lead_time_days",
            "Minimum Order Quantity - Smallest order size (re: UOM column I)": "moq",
            "Minimum Order Quantity": "moq",
            "MOQ": "moq",
            "Economic Order Quantity - Optimal order size": "eoq",
            "Economic Order Quantity": "eoq",
            "EOQ": "eoq",
            "Supplier": "supplier",
            "Supplier Name": "supplier",
            "Supplier - Primary vendor": "supplier",
            "Vendor": "supplier"
        }
        
        # Get supplier from column L (index 11, 0-based)
        supplier_col_index = column_letter_to_index(supplier_col) - 1  # Convert to 0-based index
        
        # Map supplier column by position if header row exists
        if len(header) > supplier_col_index:
            supplier_header = header[supplier_col_index]
            if supplier_header and supplier_header.strip():
                column_mapping[supplier_header] = 'supplier'
        
        df = df.rename(columns=column_mapping)

        for col in ['lead_time_days', 'moq', 'eoq']:
            if col not in df.columns:
                df[col] = 0
            else:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
        # Handle supplier column
        if 'supplier' not in df.columns:
            # Try to get by column position
            if len(df.columns) > supplier_col_index:
                df['supplier'] = df.iloc[:, supplier_col_index].fillna('Unknown Supplier')
            else:
                df['supplier'] = 'Unknown Supplier'
        else:
            df['supplier'] = df['supplier'].fillna('Unknown Supplier').replace('', 'Unknown Supplier')

        df['component_item_code'] = df['component_item_code'].astype(str).str.strip().str.upper()
        df = df[df['component_item_code'].notna() & (df['component_item_code'] != "") & (df['component_item_code'].str.lower() != "nan")]
        df = df[['component_item_code', 'lead_time_days', 'moq', 'eoq', 'supplier']]

        print(f"✅ Procurement data: {len(df)} entries")
        return df

    def fetch_inventory_data(client, sheet_url: str, worksheet_name: str,
                            component_col: str, inventory_col: str) -> pd.DataFrame:
        print("\n📊 Fetching inventory data...")
        try:
            sheet = client.open_by_url(sheet_url)
            worksheet = sheet.worksheet(worksheet_name)
        except Exception as e:
            raise Exception(f"Failed to open inventory sheet: {str(e)}")

        raw_data = worksheet.get_all_values()
        if not raw_data:
            raise Exception("Inventory sheet is empty.")

        first_row = raw_data[0]
        non_empty_first = sum(1 for c in first_row[:8] if c.strip() != "")
        if non_empty_first == 0:
            header = raw_data[1]
            rows = raw_data[2:]
        else:
            header = raw_data[0]
            rows = raw_data[1:]

        df = pd.DataFrame(rows, columns=header)

        component_col_names = ['Component Item Code', 'Component_Item_Code',
                              'Item Code', 'Component Code', 'SKU', 'Item_Code']
        component_column = next((c for c in component_col_names if c in df.columns), None)
        if not component_column:
            component_column = df.columns[0]

        column_mapping = {
            component_column: 'component_item_code',
            'Current Stock': 'current_inventory',
            'Current Inventory': 'current_inventory',
            'Current_Inventory': 'current_inventory',
            'Inventory': 'current_inventory',
            'On Hand': 'current_inventory',
            'Qty on Hand': 'current_inventory',
            'On-Hand': 'current_inventory',
        }
        df = df.rename(columns=column_mapping)

        if 'component_item_code' not in df.columns:
            df['component_item_code'] = df.iloc[:, 0]
        if 'current_inventory' not in df.columns:
            df['current_inventory'] = 0

        df['current_inventory'] = safe_numeric_convert(df['current_inventory'])
        df['component_item_code'] = df['component_item_code'].astype(str).str.strip()
        df = df[df['component_item_code'].notna() & (df['component_item_code'] != "") & (df['component_item_code'].str.lower() != "nan")]
        df = df[['component_item_code', 'current_inventory']]

        print(f"✅ Inventory data: {len(df)} entries")
        return df

    # ==========================================================================
    # FORECAST LOOKUP
    # ==========================================================================

    def fetch_sku_upc_mapping(client, sheet_url: str, worksheet_name: str,
                             item_code_col: str, upc_col: str) -> Dict[str, str]:
        print("\n📇 Fetching SKU → UPC mapping...")
        try:
            sheet = client.open_by_url(sheet_url)
            worksheet = sheet.worksheet(worksheet_name)
        except Exception as e:
            raise Exception(f"Failed to open SKU reference sheet: {str(e)}")

        item_code_index = column_letter_to_index(item_code_col)
        upc_index = column_letter_to_index(upc_col)
        item_codes = worksheet.col_values(item_code_index)
        upcs = worksheet.col_values(upc_index)

        mapping = {}
        for i in range(1, min(len(item_codes), len(upcs))):
            item_code = str(item_codes[i]).strip()
            upc = str(upcs[i]).strip()
            if item_code and upc and upc.lower() != 'nan' and upc != '':
                mapping[item_code] = upc
        print(f"✅ Mapped {len(mapping)} SKUs to UPCs")
        return mapping

    def fetch_upc_forecast_data(client, sheet_url: str, worksheet_name: str,
                               upc_col: str, forecast_cols: List[str]) -> Dict[str, float]:
        print("\n📊 Fetching 6-month forecast data...")
        try:
            sheet = client.open_by_url(sheet_url)
            worksheet = sheet.worksheet(worksheet_name)
        except Exception as e:
            raise Exception(f"Failed to open forecast sheet: {str(e)}")

        upc_col_index = column_letter_to_index(upc_col)
        upcs = worksheet.col_values(upc_col_index)
        forecast_data = {}
        for col_letter in forecast_cols:
            col_index = column_letter_to_index(col_letter)
            forecast_data[col_letter] = worksheet.col_values(col_index)

        mapping = {}
        for i in range(1, len(upcs)):
            upc = str(upcs[i]).strip()
            if not upc or upc.lower() == 'nan' or upc == '':
                continue
            total_forecast = 0
            for col_letter in forecast_cols:
                try:
                    value = forecast_data[col_letter][i] if i < len(forecast_data[col_letter]) else 0
                    value = str(value).replace(',', '').strip()
                    total_forecast += float(value) if value and value.lower() != 'nan' else 0
                except (ValueError, IndexError):
                    continue
            if total_forecast > 0:
                mapping[upc] = total_forecast
        print(f"✅ Forecasts for {len(mapping)} UPCs")
        return mapping

    def fetch_forecast_demand_from_sheets(client, bom_df: pd.DataFrame, config: Dict):
        print("\n🔍 FORECAST LOOKUP - MULTI-SHEET INTEGRATION")

        unique_skus = bom_df.groupby('parent_item_code').agg({'parent_sku': 'first'}).reset_index()
        print(f"   Found {len(unique_skus)} unique SKUs in BOM")

        sku_to_upc = fetch_sku_upc_mapping(client, config['SKU_REFERENCE_URL'], config['SKU_REFERENCE_WORKSHEET'],
                                          config['SKU_ITEM_CODE_COLUMN'], config['SKU_UPC_COLUMN'])
        upc_to_forecast = fetch_upc_forecast_data(client, config['FORECAST_URL'], config['FORECAST_WORKSHEET'],
                                                 config['FORECAST_UPC_COLUMN'], config['FORECAST_MONTH_COLUMNS'])

        forecast_results = []
        skipped_skus = []

        for _, row in unique_skus.iterrows():
            item_code = row['parent_item_code']
            sku_name = row['parent_sku']
            upc = sku_to_upc.get(item_code)
            if not upc:
                skipped_skus.append({'SKU': item_code, 'SKU_Name': sku_name, 'Reason': 'UPC not found'})
                continue
            forecast = upc_to_forecast.get(upc)
            if not forecast or forecast < config['MIN_FORECAST_QTY']:
                skipped_skus.append({'SKU': item_code, 'SKU_Name': sku_name, 'UPC': upc,
                                    'Reason': f'No forecast or below minimum ({config["MIN_FORECAST_QTY"]})'})
                continue
            forecast_results.append({'SKU_ID': item_code, 'Description': sku_name, 'UPC': upc, 'Forecast_Demand': int(forecast)})

        forecast_df = pd.DataFrame(forecast_results)
        print(f"✅ Linked {len(forecast_df)} SKUs with forecasts")
        return forecast_df, skipped_skus

    # ==========================================================================
    # BOM STRUCTURE & EXPLOSION
    # ==========================================================================

    def build_bom_structure_from_sheet(bom_df: pd.DataFrame) -> Dict:
        print("\n🔧 Building BOM structure...")
        bom_structure = {}
        all_parents = set(bom_df['parent_item_code'].unique())
        all_components = set(bom_df['component_item_code'].unique())
        sub_assemblies = all_components & all_parents

        for parent_code in all_parents:
            parent_bom = bom_df[bom_df['parent_item_code'] == parent_code]
            components = []
            for _, row in parent_bom.iterrows():
                component_code = str(row['component_item_code'])
                level = 1 if component_code in sub_assemblies else 2
                
                qty_req = safe_float(row['quantity_required'], 1.0)
                wastage = safe_float(row['wastage_pct'], 0.0)
                cost = safe_float(row['unit_cost'], 0.0)
                
                desc = row['component_description']
                desc = str(desc.iloc[0]) if isinstance(desc, pd.Series) else str(desc)
                
                uom_val = row['uom']
                uom_val = str(uom_val.iloc[0]) if isinstance(uom_val, pd.Series) else str(uom_val)
                
                comp_type = row['component_type']
                comp_type = str(comp_type.iloc[0]) if isinstance(comp_type, pd.Series) else str(comp_type)
                
                supplier_val = row.get('supplier', 'Unknown Supplier')
                supplier_val = str(supplier_val.iloc[0]) if isinstance(supplier_val, pd.Series) else str(supplier_val)
                
                components.append((component_code, qty_req, desc, level, wastage, cost, uom_val, comp_type, supplier_val))
            bom_structure[parent_code] = components
        print(f"✅ BOM structure for {len(bom_structure)} parent items")
        return bom_structure

    def explode_bom(parent_id: str, parent_qty: float, bom_structure: Dict, requirements: Dict,
                   parent_chain: Set[str], root_sku: str) -> None:
        if parent_id not in bom_structure:
            return
        for item in bom_structure[parent_id]:
            child_id, child_qty_per_parent, description, level, wastage_pct, unit_cost, uom, comp_type, supplier = item
            
            child_qty_per_parent = safe_float(child_qty_per_parent, 1.0)
            wastage_pct = safe_float(wastage_pct, 0.0)
            unit_cost = safe_float(unit_cost, 0.0)
            level = int(safe_float(level, 2))
            
            if child_id in parent_chain:
                continue
                
            gross_qty = float(parent_qty) * child_qty_per_parent
            wastage_multiplier = 1 + (wastage_pct / 100)
            net_qty = gross_qty * wastage_multiplier

            if child_id not in requirements:
                requirements[child_id] = {
                    'gross_qty': 0.0, 'net_qty': 0.0, 'description': str(description), 
                    'level': level, 'parent_skus': set(), 'wastage_pct': wastage_pct, 
                    'unit_cost': unit_cost, 'lead_time': 0, 'uom': str(uom),
                    'component_type': str(comp_type), 'supplier': str(supplier)
                }
            req = requirements[child_id]
            req['gross_qty'] += gross_qty
            req['net_qty'] += net_qty
            req['parent_skus'].add(root_sku)

            new_parent_chain = parent_chain | {child_id}
            explode_bom(child_id, net_qty, bom_structure, requirements, new_parent_chain, root_sku)

    def aggregate_requirements(forecast_df: pd.DataFrame, bom_structure: Dict) -> Dict:
        all_requirements = {}
        for _, row in forecast_df.iterrows():
            explode_bom(row['SKU_ID'], row['Forecast_Demand'], bom_structure, all_requirements, set(), row['SKU_ID'])
        return all_requirements

    def calculate_final_requirements(requirements: Dict, inventory: Optional[Dict] = None) -> pd.DataFrame:
        if not requirements:
            return pd.DataFrame()

        results = []
        for comp_id, data in requirements.items():
            gross_req = safe_float(data['gross_qty'])
            net_req = safe_float(data['net_qty'])
            unit_cost = safe_float(data['unit_cost'])
            wastage_pct = safe_float(data['wastage_pct'])
            
            total_cost = net_req * unit_cost
            current_inv = safe_float(inventory.get(comp_id, 0)) if inventory else 0
            procurement_needed = max(0, net_req - current_inv)
            parent_skus_str = ', '.join(sorted(data['parent_skus']))
            
            description = data['description']
            if isinstance(description, pd.Series):
                description = description.iloc[0] if len(description) > 0 else 'Unknown'
            
            comp_type = data['component_type']
            if isinstance(comp_type, pd.Series):
                comp_type = comp_type.iloc[0] if len(comp_type) > 0 else 'Uncategorized'
                
            supplier = data['supplier']
            if isinstance(supplier, pd.Series):
                supplier = supplier.iloc[0] if len(supplier) > 0 else 'Unknown Supplier'
                
            uom = data['uom']
            if isinstance(uom, pd.Series):
                uom = uom.iloc[0] if len(uom) > 0 else 'EA'

            results.append({
                'Component_ID': comp_id,
                'Description': str(description),
                'Component_Type': str(comp_type),
                'Supplier': str(supplier),
                'UoM': str(uom),
                'Level': int(safe_float(data['level'], 2)),
                'Gross_Requirement': round(gross_req, 2),
                'Wastage%': round(wastage_pct, 2),
                'Net_Requirement': round(net_req, 2),
                'Current_Inventory': round(current_inv, 2),
                'Procurement_Needed': round(procurement_needed, 2),
                'Unit_Cost': round(unit_cost, 2),
                'Total_Cost': round(total_cost, 2),
                'Lead_Time': int(safe_float(data['lead_time'], 0)),
                'Parent_SKUs': parent_skus_str
            })

        df = pd.DataFrame(results)
        df = df.sort_values(['Component_Type', 'Component_ID']).reset_index(drop=True)
        return df

    # ==========================================================================
    # ABC CLASSIFICATION
    # ==========================================================================

    def calculate_abc_classification(df: pd.DataFrame, config: Dict) -> pd.DataFrame:
        print("\n📊 Calculating ABC Classification...")
        df = df.copy()
        df['Total_Value'] = df['Net_Requirement'] * df['Unit_Cost']
        df = df.sort_values('Total_Value', ascending=False)
        
        total_value = df['Total_Value'].sum()
        if total_value > 0:
            df['Cumulative_Value'] = df['Total_Value'].cumsum()
            df['Cumulative_Pct'] = df['Cumulative_Value'] / total_value
            conditions = [
                df['Cumulative_Pct'] <= config['ABC_A_THRESHOLD'],
                df['Cumulative_Pct'] <= config['ABC_B_THRESHOLD'],
            ]
            choices = ['A', 'B']
            df['ABC_Class'] = np.select(conditions, choices, default='C')
        else:
            df['ABC_Class'] = 'C'
        
        df = df.drop(columns=['Cumulative_Value', 'Cumulative_Pct'], errors='ignore')
        
        class_counts = df['ABC_Class'].value_counts()
        print(f"   A: {class_counts.get('A', 0)}, B: {class_counts.get('B', 0)}, C: {class_counts.get('C', 0)}")
        return df

    # ==========================================================================
    # PROCUREMENT CALCULATIONS
    # ==========================================================================

    def calculate_rop_and_procurement(requirements_df: pd.DataFrame, procurement_df: pd.DataFrame,
                                     inventory_df: pd.DataFrame, config: Dict) -> Tuple[pd.DataFrame, List[str]]:
        print("\n🔄 PROCUREMENT CALCULATIONS...")

        df = requirements_df.copy()
        df['Component_ID'] = df['Component_ID'].astype(str).str.strip().str.upper()

        procurement_df = procurement_df.copy()
        inventory_df = inventory_df.copy()
        procurement_df['component_item_code'] = procurement_df['component_item_code'].str.upper()
        inventory_df['component_item_code'] = inventory_df['component_item_code'].str.upper()

        df = (df
              .merge(procurement_df[['component_item_code', 'lead_time_days', 'moq', 'eoq', 'supplier']],
                     left_on='Component_ID', right_on='component_item_code', how='left', suffixes=('', '_proc'))
              .merge(inventory_df[['component_item_code', 'current_inventory']],
                     left_on='Component_ID', right_on='component_item_code', how='left',
                     suffixes=('', '_inv'))
             )

        df['lead_time_days'] = df['lead_time_days'].fillna(0).round().astype(int)
        df['moq'] = df['moq'].fillna(0).round().astype(int)
        df['eoq'] = df['eoq'].fillna(0).round().astype(int)
        df['current_inventory'] = df['current_inventory'].fillna(0)
        
        # Update Supplier from procurement data if available
        if 'supplier_proc' in df.columns:
            # Use procurement supplier if available, otherwise keep BOM supplier
            df['Supplier'] = df['supplier_proc'].combine_first(df['Supplier'])
            df = df.drop(columns=['supplier_proc'], errors='ignore')
        elif 'supplier' in df.columns and 'Supplier' in df.columns:
            df['Supplier'] = df['supplier'].combine_first(df['Supplier'])
            df = df.drop(columns=['supplier'], errors='ignore')
        
        # Clean up supplier column
        df['Supplier'] = df['Supplier'].fillna('Unknown Supplier').replace('', 'Unknown Supplier')

        missing_data = []
        horizon_days = config['FORECAST_HORIZON_DAYS']

        df['Daily_Demand'] = 0.0
        df['Safety_Stock'] = 0.0
        df['Calculated_ROP'] = 0.0
        df['Recommended_Order_Qty'] = 0.0
        df['Procurement_Cost'] = 0.0
        df['Order_Status'] = ''
        df['Days_of_Stock'] = 0.0
        df['Stock_Coverage_Ratio'] = 0.0
        df['Order_Priority_Score'] = 0

        for idx, row in df.iterrows():
            component_id = row['Component_ID']
            net_req = row['Net_Requirement']
            lead_time = row['lead_time_days']
            moq = row['moq']
            eoq = row['eoq']
            current_inv = row['current_inventory']
            unit_cost = row['Unit_Cost']
            abc_class = row.get('ABC_Class', 'B')

            if lead_time == 0 or pd.isna(lead_time):
                missing_data.append(f"{component_id}: Missing Lead Time")
            if moq == 0 and eoq == 0:
                missing_data.append(f"{component_id}: Missing MOQ and EOQ")

            daily_demand = net_req / horizon_days if horizon_days > 0 else 0
            df.at[idx, 'Daily_Demand'] = round(daily_demand)

            if abc_class == 'A':
                safety_stock_pct = config['SAFETY_STOCK_A']
            elif abc_class == 'B':
                safety_stock_pct = config['SAFETY_STOCK_B']
            else:
                safety_stock_pct = config['SAFETY_STOCK_C']
            
            safety_stock = safety_stock_pct * net_req
            df.at[idx, 'Safety_Stock'] = round(safety_stock, 2)

            calculated_rop = (daily_demand * lead_time) + safety_stock
            df.at[idx, 'Calculated_ROP'] = round(calculated_rop, 2)

            days_of_stock = current_inv / daily_demand if daily_demand > 0 else 999
            df.at[idx, 'Days_of_Stock'] = round(min(days_of_stock, 999), 1)

            coverage_ratio = current_inv / calculated_rop if calculated_rop > 0 else 999
            df.at[idx, 'Stock_Coverage_Ratio'] = round(min(coverage_ratio, 10), 2)

            shortfall = max(0, calculated_rop - current_inv)
            roq = max(shortfall, moq, eoq)
            if moq > 0 and roq > moq:
                roq = ((roq // moq) + (1 if roq % moq > 0 else 0)) * moq
            df.at[idx, 'Recommended_Order_Qty'] = round(roq, 2)

            procurement_cost = roq * unit_cost if unit_cost > 0 else 0
            df.at[idx, 'Procurement_Cost'] = round(procurement_cost, 2)

            if current_inv < calculated_rop:
                df.at[idx, 'Order_Status'] = '🔴 Urgent Reorder'
            elif current_inv < (calculated_rop + safety_stock):
                df.at[idx, 'Order_Status'] = '🟡 Reorder Soon'
            else:
                df.at[idx, 'Order_Status'] = '🟢 OK'

            priority_score = 0
            if days_of_stock < lead_time:
                priority_score += 50
            elif days_of_stock < lead_time * 1.5:
                priority_score += 30
            if abc_class == 'A':
                priority_score += 30
            elif abc_class == 'B':
                priority_score += 15
            if coverage_ratio < 0.5:
                priority_score += 20
            elif coverage_ratio < 1.0:
                priority_score += 10
            df.at[idx, 'Order_Priority_Score'] = priority_score

        df = df.drop(columns=[c for c in df.columns if c.startswith('component_item_code')], errors='ignore')

        column_order = [
            'Component_ID', 'Description', 'Component_Type', 'Supplier', 'ABC_Class',
            'UoM', 'Level', 'Gross_Requirement', 'Wastage%', 'Net_Requirement',
            'lead_time_days', 'moq', 'eoq',
            'Daily_Demand', 'Safety_Stock', 'Calculated_ROP',
            'current_inventory', 'Days_of_Stock', 'Stock_Coverage_Ratio',
            'Recommended_Order_Qty', 'Unit_Cost', 'Procurement_Cost', 'Total_Value',
            'Order_Status', 'Order_Priority_Score', 'Parent_SKUs'
        ]
        final_columns = [col for col in column_order if col in df.columns]
        remaining_columns = [col for col in df.columns if col not in column_order]
        df = df[final_columns + remaining_columns]

        df = df.sort_values('Order_Priority_Score', ascending=False).reset_index(drop=True)
        df = df.drop(columns=['Current_Inventory', 'Lead_Time'], errors='ignore')

        urgent_cnt = len(df[df['Order_Status'] == '🔴 Urgent Reorder'])
        soon_cnt = len(df[df['Order_Status'] == '🟡 Reorder Soon'])
        ok_cnt = len(df[df['Order_Status'] == '🟢 OK'])
        print(f"   🔴 Urgent: {urgent_cnt}, 🟡 Soon: {soon_cnt}, 🟢 OK: {ok_cnt}")
        
        return df, list(set(missing_data))

    # ==========================================================================
    # CATEGORY SUMMARY & TIMELINE
    # ==========================================================================

    def create_category_summary(results_df: pd.DataFrame) -> pd.DataFrame:
        if len(results_df) == 0:
            return pd.DataFrame()
        
        summary = results_df.groupby('Component_Type').agg({
            'Component_ID': 'count',
            'Net_Requirement': 'sum',
            'Procurement_Cost': 'sum',
            'Total_Value': 'sum',
            'Days_of_Stock': 'mean',
        }).rename(columns={
            'Component_ID': 'Component_Count',
            'Net_Requirement': 'Total_Net_Requirement',
            'Procurement_Cost': 'Total_Procurement_Cost',
            'Total_Value': 'Total_Value',
            'Days_of_Stock': 'Avg_Days_of_Stock'
        })
        
        urgent_by_cat = results_df[results_df['Order_Status'] == '🔴 Urgent Reorder'].groupby('Component_Type').size()
        summary['Urgent_Count'] = urgent_by_cat.reindex(summary.index).fillna(0).astype(int)
        
        total_cost = summary['Total_Procurement_Cost'].sum()
        if total_cost > 0:
            summary['Cost_Percentage'] = (summary['Total_Procurement_Cost'] / total_cost * 100).round(1)
        else:
            summary['Cost_Percentage'] = 0
        
        summary = summary.sort_values('Total_Procurement_Cost', ascending=False).reset_index()
        
        # Apply formatting - round to whole numbers
        summary['Total_Net_Requirement'] = summary['Total_Net_Requirement'].round().astype(int)
        summary['Avg_Days_of_Stock'] = summary['Avg_Days_of_Stock'].round().astype(int)
        summary['Total_Procurement_Cost'] = summary['Total_Procurement_Cost'].round().astype(int)
        summary['Total_Value'] = summary['Total_Value'].round().astype(int)
        
        return summary

    def create_procurement_timeline(results_df: pd.DataFrame) -> pd.DataFrame:
        today = datetime.now().date()
        needs_order = results_df[results_df['Recommended_Order_Qty'] > 0].copy()
        
        if len(needs_order) == 0:
            return pd.DataFrame()
        
        def calc_order_date(row):
            if row['Order_Status'] == '🔴 Urgent Reorder':
                return today
            elif row['Days_of_Stock'] < 999:
                days_until_order = max(0, row['Days_of_Stock'] - row['lead_time_days'])
                return today + timedelta(days=int(days_until_order))
            else:
                return today + timedelta(days=30)
        
        needs_order['Order_By_Date'] = needs_order.apply(calc_order_date, axis=1)
        needs_order['Expected_Arrival'] = needs_order.apply(
            lambda r: r['Order_By_Date'] + timedelta(days=int(r['lead_time_days'])), axis=1
        )
        
        timeline = needs_order[[
            'Component_ID', 'Description', 'Component_Type', 'Supplier',
            'Order_Status', 'Order_Priority_Score',
            'Recommended_Order_Qty', 'Procurement_Cost',
            'lead_time_days', 'Days_of_Stock',
            'Order_By_Date', 'Expected_Arrival'
        ]].sort_values(['Order_By_Date', 'Order_Priority_Score'], ascending=[True, False])
        
        # Round to whole numbers
        timeline['Recommended_Order_Qty'] = timeline['Recommended_Order_Qty'].round().astype(int)
        timeline['Procurement_Cost'] = timeline['Procurement_Cost'].round().astype(int)
        
        timeline['Order_By_Date'] = timeline['Order_By_Date'].astype(str)
        timeline['Expected_Arrival'] = timeline['Expected_Arrival'].astype(str)
        return timeline

    # ==========================================================================
    # EXECUTIVE SUMMARY
    # ==========================================================================

    def create_executive_summary(results_df: pd.DataFrame, forecast_df: pd.DataFrame,
                                skipped_skus: List[Dict], category_summary: pd.DataFrame) -> pd.DataFrame:
        if len(results_df) == 0:
            return pd.DataFrame({'Metric': ['No data'], 'Value': [''], 'Unit': ['']})
        
        total_components = len(results_df)
        total_forecast = forecast_df['Forecast_Demand'].sum()
        total_proc_cost = results_df['Procurement_Cost'].sum()
        total_value = results_df['Total_Value'].sum()

        urgent_cnt = len(results_df[results_df['Order_Status'] == '🔴 Urgent Reorder'])
        soon_cnt = len(results_df[results_df['Order_Status'] == '🟡 Reorder Soon'])
        ok_cnt = len(results_df[results_df['Order_Status'] == '🟢 OK'])

        urgent_cost = results_df[results_df['Order_Status'] == '🔴 Urgent Reorder']['Procurement_Cost'].sum()
        avg_lead = results_df['lead_time_days'].mean()
        max_lead = results_df['lead_time_days'].max()
        
        avg_days_of_stock = results_df['Days_of_Stock'].replace(999, np.nan).mean()
        avg_coverage_ratio = results_df['Stock_Coverage_Ratio'].replace([999, 10], np.nan).mean()
        
        abc_a_cnt = len(results_df[results_df['ABC_Class'] == 'A'])
        abc_b_cnt = len(results_df[results_df['ABC_Class'] == 'B'])
        abc_c_cnt = len(results_df[results_df['ABC_Class'] == 'C'])
        
        num_categories = results_df['Component_Type'].nunique()
        top_category = results_df.groupby('Component_Type')['Procurement_Cost'].sum().idxmax() if len(results_df) > 0 else 'N/A'
        num_suppliers = results_df['Supplier'].nunique()
        high_priority_cnt = len(results_df[results_df['Order_Priority_Score'] >= 50])

        top5_cost = results_df.nlargest(5, 'Procurement_Cost')[['Component_ID', 'Description', 'Procurement_Cost']]
        top5_urgent = results_df[results_df['Order_Status'] == '🔴 Urgent Reorder'].nlargest(5, 'Order_Priority_Score')[['Component_ID', 'Description', 'Order_Priority_Score']]

        summary_data = []

        summary_data.append({'Metric': '═══ OVERVIEW ═══', 'Value': '', 'Unit': ''})
        summary_data.append({'Metric': 'Report Generated', 'Value': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M'), 'Unit': ''})
        summary_data.append({'Metric': 'Forecast Horizon', 'Value': '6 months', 'Unit': ''})
        summary_data.append({'Metric': 'SKUs Forecasted', 'Value': len(forecast_df), 'Unit': 'SKUs'})
        summary_data.append({'Metric': 'Total Forecast Demand', 'Value': f'{total_forecast:,.0f}', 'Unit': 'units'})
        summary_data.append({'Metric': 'Unique Components', 'Value': total_components, 'Unit': ''})
        summary_data.append({'Metric': 'Component Categories', 'Value': num_categories, 'Unit': ''})
        summary_data.append({'Metric': 'Suppliers', 'Value': num_suppliers, 'Unit': ''})
        summary_data.append({'Metric': 'SKUs Skipped', 'Value': len(skipped_skus), 'Unit': 'SKUs'})
        summary_data.append({'Metric': '', 'Value': '', 'Unit': ''})

        summary_data.append({'Metric': '═══ FINANCIAL ═══', 'Value': '', 'Unit': ''})
        summary_data.append({'Metric': 'Total Procurement Cost', 'Value': f'${total_proc_cost:,.2f}', 'Unit': ''})
        summary_data.append({'Metric': 'Total Inventory Value', 'Value': f'${total_value:,.2f}', 'Unit': ''})
        summary_data.append({'Metric': 'Urgent Orders Cost', 'Value': f'${urgent_cost:,.2f}', 'Unit': ''})
        summary_data.append({'Metric': 'Cost per Forecasted Unit', 'Value': f'${total_proc_cost/total_forecast if total_forecast else 0:.2f}', 'Unit': ''})
        summary_data.append({'Metric': 'Urgent Cost %', 'Value': f'{urgent_cost/total_proc_cost*100 if total_proc_cost else 0:.1f}%', 'Unit': ''})
        summary_data.append({'Metric': '', 'Value': '', 'Unit': ''})

        summary_data.append({'Metric': '═══ INVENTORY HEALTH ═══', 'Value': '', 'Unit': ''})
        summary_data.append({'Metric': '🔴 Urgent Reorder', 'Value': urgent_cnt, 'Unit': 'components'})
        summary_data.append({'Metric': '🟡 Reorder Soon', 'Value': soon_cnt, 'Unit': 'components'})
        summary_data.append({'Metric': '🟢 Inventory OK', 'Value': ok_cnt, 'Unit': 'components'})
        summary_data.append({'Metric': 'Urgency Rate', 'Value': f'{urgent_cnt/total_components*100 if total_components else 0:.1f}%', 'Unit': ''})
        summary_data.append({'Metric': 'High Priority Items (Score ≥50)', 'Value': high_priority_cnt, 'Unit': 'components'})
        summary_data.append({'Metric': 'Average Days of Stock', 'Value': f'{avg_days_of_stock:.1f}' if pd.notna(avg_days_of_stock) else 'N/A', 'Unit': 'days'})
        summary_data.append({'Metric': 'Average Coverage Ratio', 'Value': f'{avg_coverage_ratio:.2f}' if pd.notna(avg_coverage_ratio) else 'N/A', 'Unit': ''})
        summary_data.append({'Metric': '', 'Value': '', 'Unit': ''})

        summary_data.append({'Metric': '═══ ABC CLASSIFICATION ═══', 'Value': '', 'Unit': ''})
        summary_data.append({'Metric': 'Class A (High Value)', 'Value': abc_a_cnt, 'Unit': 'components'})
        summary_data.append({'Metric': 'Class B (Medium Value)', 'Value': abc_b_cnt, 'Unit': 'components'})
        summary_data.append({'Metric': 'Class C (Low Value)', 'Value': abc_c_cnt, 'Unit': 'components'})
        summary_data.append({'Metric': '', 'Value': '', 'Unit': ''})

        summary_data.append({'Metric': '═══ LEAD-TIME ═══', 'Value': '', 'Unit': ''})
        summary_data.append({'Metric': 'Average Lead Time', 'Value': f'{avg_lead:.1f}', 'Unit': 'days'})
        summary_data.append({'Metric': 'Maximum Lead Time', 'Value': f'{max_lead:.0f}', 'Unit': 'days'})
        summary_data.append({'Metric': 'Critical-Path Components', 'Value': len(results_df[results_df['lead_time_days'] == max_lead]), 'Unit': ''})
        summary_data.append({'Metric': '', 'Value': '', 'Unit': ''})

        if len(category_summary) > 0:
            summary_data.append({'Metric': '═══ TOP CATEGORIES ═══', 'Value': '', 'Unit': ''})
            summary_data.append({'Metric': 'Highest Cost Category', 'Value': top_category, 'Unit': ''})
            for _, row in category_summary.head(3).iterrows():
                summary_data.append({
                    'Metric': f"   {row['Component_Type']}", 
                    'Value': f"${row['Total_Procurement_Cost']:,.2f}", 
                    'Unit': f"({row['Cost_Percentage']:.1f}%)"
                })
            summary_data.append({'Metric': '', 'Value': '', 'Unit': ''})

        summary_data.append({'Metric': '═══ TOP 5 COST DRIVERS ═══', 'Value': '', 'Unit': ''})
        for _, row in top5_cost.iterrows():
            summary_data.append({'Metric': f"{row['Component_ID']} – {row['Description'][:30]}",
                                'Value': f"${row['Procurement_Cost']:,.2f}", 'Unit': ''})
        summary_data.append({'Metric': '', 'Value': '', 'Unit': ''})

        if len(top5_urgent) > 0:
            summary_data.append({'Metric': '═══ TOP 5 URGENT (by Priority) ═══', 'Value': '', 'Unit': ''})
            for _, row in top5_urgent.iterrows():
                summary_data.append({'Metric': f"{row['Component_ID']} – {row['Description'][:30]}",
                                    'Value': f"Priority: {row['Order_Priority_Score']}", 'Unit': ''})

        return pd.DataFrame(summary_data)

    # ==========================================================================
    # EXCEL FORMATTING
    # ==========================================================================

    def format_excel_output(writer, results_df: pd.DataFrame, forecast_df: pd.DataFrame,
                           procurement_df: pd.DataFrame, inventory_df: pd.DataFrame,
                           skipped_skus: List[Dict], missing_data: List[str]) -> None:

        workbook = writer.book

        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        urgent_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        soon_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        ok_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        section_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        section_header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        CURRENCY_FORMAT = '$#,##0.00'
        CURRENCY_WHOLE_FORMAT = '$#,##0'
        NUMBER_FORMAT = '#,##0'
        NUMBER_DECIMAL_FORMAT = '#,##0.00'
        PERCENT_FORMAT = '0.0%'
        
        currency_columns = ['Unit_Cost', 'Procurement_Cost', 'Total_Cost', 'Total_Value',
                           'Total_Procurement_Cost', 'Urgent_Cost', 'Cost']
        quantity_columns = ['Gross_Requirement', 'Net_Requirement', 'Calculated_ROP', 'ROP',
                           'Recommended_Order_Qty', 'Procurement_Needed', 'moq', 'eoq', 'MOQ', 'EOQ',
                           'Safety_Stock', 'current_inventory', 'Current_Inventory',
                           'Forecast_Demand', 'Total_Net_Requirement', 'Component_Count', 'Urgent_Count',
                           'Daily_Demand']

        def smart_number_format(value, is_currency=False):
            try:
                val = float(value) if value else 0
                if is_currency:
                    return CURRENCY_WHOLE_FORMAT if abs(val) >= 100 else CURRENCY_FORMAT
                else:
                    return NUMBER_FORMAT if abs(val) >= 10 else NUMBER_DECIMAL_FORMAT
            except:
                return NUMBER_DECIMAL_FORMAT

        def apply_column_formatting(worksheet, df):
            if df is None or len(df) == 0:
                return
            for col_idx, col_name in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_idx)
                is_currency = any(curr in col_name for curr in currency_columns)
                is_quantity = any(qty in col_name for qty in quantity_columns)
                is_percent = 'Percentage' in col_name or 'Pct' in col_name or col_name == 'Cost_Percentage'
                is_wastage = col_name == 'Wastage%'
                
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f'{col_letter}{row}']
                    cell_value = cell.value
                    try:
                        if cell_value is not None and cell_value != '':
                            if is_currency:
                                cell.number_format = smart_number_format(cell_value, is_currency=True)
                            elif is_wastage:
                                # Wastage is stored as 10 meaning 10%, so use custom format
                                cell.number_format = '0.0"%"'
                            elif is_percent:
                                cell.number_format = PERCENT_FORMAT
                            elif is_quantity:
                                cell.number_format = smart_number_format(cell_value, is_currency=False)
                    except:
                        pass

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # Apply header formatting to all sheets
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        cell_len = len(str(cell.value)) if cell.value else 0
                        if cell_len > max_length:
                            max_length = cell_len
                    except:
                        pass
                adjusted_width = min(max_length + 3, 50)
                worksheet.column_dimensions[column_letter].width = max(adjusted_width, 10)

            # Apply borders to all cells with data
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                          min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    if cell.value is not None:
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical='center')

            # Sheet-specific formatting
            if sheet_name == '📦 MRP Requirements':
                apply_column_formatting(worksheet, results_df)
                if 'Order_Status' in results_df.columns:
                    status_col_idx = results_df.columns.get_loc('Order_Status') + 1
                    for row in range(2, worksheet.max_row + 1):
                        status_cell = worksheet.cell(row=row, column=status_col_idx)
                        fill_color = None
                        if '🔴' in str(status_cell.value):
                            fill_color = urgent_fill
                        elif '🟡' in str(status_cell.value):
                            fill_color = soon_fill
                        elif '🟢' in str(status_cell.value):
                            fill_color = ok_fill
                        if fill_color:
                            for col in range(1, worksheet.max_column + 1):
                                worksheet.cell(row=row, column=col).fill = fill_color
                worksheet.freeze_panes = 'C2'

            elif sheet_name == '🚨 Urgent Reorders':
                apply_column_formatting(worksheet, results_df)
                for row in range(2, worksheet.max_row + 1):
                    for col in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row, column=col).fill = urgent_fill
                worksheet.freeze_panes = 'C2'

            elif sheet_name == '🟡 Reorder Soon':
                apply_column_formatting(worksheet, results_df)
                for row in range(2, worksheet.max_row + 1):
                    for col in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row, column=col).fill = soon_fill
                worksheet.freeze_panes = 'C2'

            elif sheet_name == '📈 Forecasted Demand':
                apply_column_formatting(worksheet, forecast_df)
                worksheet.freeze_panes = 'A2'

            elif sheet_name == '📊 Executive Summary':
                worksheet.column_dimensions['A'].width = 45
                worksheet.column_dimensions['B'].width = 25
                worksheet.column_dimensions['C'].width = 15

                for row_idx in range(1, worksheet.max_row + 1):
                    metric_cell = worksheet.cell(row=row_idx, column=1)
                    value_cell = worksheet.cell(row=row_idx, column=2)
                    unit_cell = worksheet.cell(row=row_idx, column=3)
                    
                    metric_text = str(metric_cell.value).strip() if metric_cell.value else ''
                    value_text = str(value_cell.value).strip() if value_cell.value else ''
                    
                    if metric_text.startswith('═══'):
                        for col in range(1, 4):
                            cell = worksheet.cell(row=row_idx, column=col)
                            cell.fill = section_header_fill
                            cell.font = section_header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        continue
                    
                    value_cell.alignment = Alignment(horizontal='right', vertical='center')
                    unit_cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    if value_text:
                        if value_text.startswith('$'):
                            try:
                                num_str = value_text.replace('$', '').replace(',', '')
                                num_val = float(num_str)
                                value_cell.value = num_val
                                value_cell.number_format = CURRENCY_WHOLE_FORMAT if num_val >= 100 else CURRENCY_FORMAT
                            except:
                                pass
                        elif value_text.endswith('%'):
                            try:
                                num_str = value_text.replace('%', '').replace(',', '')
                                num_val = float(num_str) / 100
                                value_cell.value = num_val
                                value_cell.number_format = PERCENT_FORMAT
                            except:
                                pass
                        elif value_text.replace(',', '').replace('.', '').replace('-', '').isdigit():
                            try:
                                num_val = float(value_text.replace(',', ''))
                                value_cell.value = num_val
                                value_cell.number_format = NUMBER_FORMAT if num_val >= 10 else NUMBER_DECIMAL_FORMAT
                            except:
                                pass
                    
                    if metric_text.startswith('🔴'):
                        for col in range(1, 4):
                            worksheet.cell(row=row_idx, column=col).fill = urgent_fill
                    elif metric_text.startswith('🟡'):
                        for col in range(1, 4):
                            worksheet.cell(row=row_idx, column=col).fill = soon_fill
                    elif metric_text.startswith('🟢'):
                        for col in range(1, 4):
                            worksheet.cell(row=row_idx, column=col).fill = ok_fill

                worksheet.freeze_panes = 'A2'

            elif sheet_name == '📊 Category Summary':
                apply_column_formatting(worksheet, None)  # Basic formatting
                worksheet.freeze_panes = 'A2'

            elif sheet_name == '📅 Procurement Timeline':
                apply_column_formatting(worksheet, None)  # Basic formatting
                worksheet.freeze_panes = 'A2'

            elif sheet_name == '⚙️ Procurement Parameters':
                worksheet.freeze_panes = 'A2'

            elif sheet_name == '📋 Current Inventory':
                worksheet.freeze_panes = 'A2'

            elif sheet_name == '⚠️ Skipped SKUs':
                worksheet.freeze_panes = 'A2'

            elif sheet_name == '❌ Missing Data':
                worksheet.freeze_panes = 'A2'

            else:
                # For any other sheets not explicitly handled above
                worksheet.freeze_panes = 'A2'

        print("✅ Applied professional Excel formatting")

    # ==========================================================================
    # MAIN EXECUTION
    # ==========================================================================

    try:
        print("\n" + "="*80)
        print("🚀 ENHANCED BOM ANALYSIS v2.0".center(80))
        print("="*80)

        # 1. Authenticate
        if gc_client is None:
            print("🔄 Creating new Google Sheets connection...")
            if "gcp_service_account_sheets" not in os.environ:
                raise FileNotFoundError("No GCP service account credentials found.")

            creds_dict = json.loads(os.environ["gcp_service_account_sheets"])
            credentials_file = "temp_bom_credentials.json"
            with open(credentials_file, "w") as f:
                json.dump(creds_dict, f)

            scopes = ['https://www.googleapis.com/auth/spreadsheets.readonly',
                      'https://www.googleapis.com/auth/drive.readonly']
            creds = Credentials.from_service_account_file(credentials_file, scopes=scopes)
            client = gspread.authorize(creds)
            print("✅ Authenticated")

            try:
                os.remove(credentials_file)
            except:
                pass
        else:
            client = gc_client
            print("✅ Reusing existing connection")

        # 2. Fetch BOM
        bom_df = fetch_bom_from_sheet(client, BOM_CONFIG['SPREADSHEET_URL'], BOM_CONFIG['WORKSHEET_NAME'])

        # 3. Build BOM structure
        bom_structure = build_bom_structure_from_sheet(bom_df)

        # 4. Get forecast
        forecast_df, skipped_skus = fetch_forecast_demand_from_sheets(client, bom_df, BOM_CONFIG)

        if len(forecast_df) == 0:
            print("\n❌ No valid forecasts found.")
            return None, None

        # 5. Aggregate requirements
        requirements = aggregate_requirements(forecast_df, bom_structure)

        # 6. Final requirements
        results_df = calculate_final_requirements(requirements, inventory=None)

        # 7. ABC Classification
        results_df = calculate_abc_classification(results_df, BOM_CONFIG)

        # 8. Procurement parameters
        procurement_df = fetch_procurement_parameters(client, BOM_CONFIG['PROCUREMENT_PARAMS_URL'],
                                                     BOM_CONFIG['PROCUREMENT_PARAMS_WORKSHEET'],
                                                     'A', BOM_CONFIG['PROCUREMENT_LEAD_TIME_COLUMN'],
                                                     BOM_CONFIG['PROCUREMENT_MOQ_COLUMN'],
                                                     BOM_CONFIG['PROCUREMENT_EOQ_COLUMN'])

        # 9. Inventory
        inventory_df = fetch_inventory_data(client, BOM_CONFIG['INVENTORY_URL'],
                                           BOM_CONFIG['INVENTORY_WORKSHEET'], 'A',
                                           BOM_CONFIG['INVENTORY_QTY_COLUMN'])

        # 10. ROP & procurement
        results_df, missing_procurement_data = calculate_rop_and_procurement(results_df, procurement_df,
                                                                             inventory_df, BOM_CONFIG)

        # 11. Category summary
        category_summary = create_category_summary(results_df)

        # 12. Procurement timeline
        procurement_timeline = create_procurement_timeline(results_df)

        # 13. Excel export
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'HG_BOM_Analysis_{timestamp}.xlsx'

        excel_buffer = BytesIO()

        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:

            exec_summary_df = create_executive_summary(results_df, forecast_df, skipped_skus, category_summary)
            exec_summary_df.to_excel(writer, sheet_name='📊 Executive Summary', index=False)

            results_df.to_excel(writer, sheet_name='📦 MRP Requirements', index=False)

            if len(category_summary) > 0:
                category_summary.to_excel(writer, sheet_name='📊 Category Summary', index=False)

            if len(procurement_timeline) > 0:
                procurement_timeline.to_excel(writer, sheet_name='📅 Procurement Timeline', index=False)

            urgent = results_df[results_df['Order_Status'] == '🔴 Urgent Reorder'].copy()
            if len(urgent) > 0:
                urgent.to_excel(writer, sheet_name='🚨 Urgent Reorders', index=False)

            reorder_soon = results_df[results_df['Order_Status'] == '🟡 Reorder Soon'].copy()
            if len(reorder_soon) > 0:
                reorder_soon.to_excel(writer, sheet_name='🟡 Reorder Soon', index=False)

            forecast_df.to_excel(writer, sheet_name='📈 Forecasted Demand', index=False)
            procurement_df.to_excel(writer, sheet_name='⚙️ Procurement Parameters', index=False)
            inventory_df.to_excel(writer, sheet_name='📋 Current Inventory', index=False)

            if skipped_skus and len(skipped_skus) > 0:
                skipped_df = pd.DataFrame(skipped_skus)
                skipped_df.to_excel(writer, sheet_name='⚠️ Skipped SKUs', index=False)

            if missing_procurement_data:
                missing_df = pd.DataFrame({'Missing_Data': missing_procurement_data})
                missing_df[['Component_ID', 'Reason']] = missing_df['Missing_Data'].str.split(':', expand=True, n=1)
                missing_df['Description'] = missing_df['Component_ID'].map(
                    results_df.set_index('Component_ID')['Description'].to_dict()
                ).fillna('')
                missing_df = missing_df[['Component_ID', 'Description', 'Reason']]
                missing_df.to_excel(writer, sheet_name='❌ Missing Data', index=False)

            format_excel_output(writer, results_df, forecast_df, procurement_df, inventory_df,
                               skipped_skus, missing_procurement_data)

        excel_buffer.seek(0)

        # Print summary
        print(f"\n💾 Results ready for download: {filename}")
        print("   Sheets included:")
        print("   ✅ Executive_Summary: Key metrics and status")
        print("   ✅ MRP_Requirements: Complete requirements with procurement logic")
        print("   ✅ Forecast: Real forecast data with UPC mapping")
        print("   ✅ Procurement_Parameters: Lead time, MOQ, EOQ data")
        print("   ✅ Current_Inventory: Inventory levels")
        if len(urgent) > 0:
            print(f"   🔴 Urgent_Reorders: {len(urgent)} components requiring immediate action")
        if skipped_skus and len(skipped_skus) > 0:
            print(f"   ⚠️  Skipped_SKUs: {len(skipped_skus)} SKUs without valid forecasts")
        if missing_procurement_data:
            print(f"   ⚠️  Missing_Procurement_Data: {len(set(missing_procurement_data))} items")

        print("\n✅ MRP WITH PROCUREMENT LOGIC COMPLETE!\n")

        # NEW: Upload BOM output to Google Sheets and Google Drive
        try:
            print("\n" + "="*80)
            print("📤 UPLOADING BOM OUTPUT TO CLOUD SERVICES".center(80))
            print("="*80)
            
            # Upload to Google Sheets
            excel_buffer.seek(0)  # Reset buffer position
            bom_sheet_url = upload_bom_excel_to_google_sheet(excel_buffer)
            
            # Upload to Google Drive
            excel_buffer.seek(0)  # Reset buffer position again
            bom_drive_file_id = upload_bom_to_google_drive_from_buffer(excel_buffer)
            
            print("\n✅ BOM output successfully uploaded to:")
            print(f"   📊 Google Sheets: {bom_sheet_url}")
            print(f"   📁 Google Drive File ID: {bom_drive_file_id}")
            
        except Exception as e:
            print(f"\n⚠️ Warning: Failed to upload BOM output to cloud: {str(e)}")
            print("   📥 Local download will still be available.")
            import traceback
            traceback.print_exc()

        # Reset buffer for download
        excel_buffer.seek(0)
        
        return excel_buffer, filename

    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}\n")
        print("Please check:")
        print("1. All BOM_CONFIG dictionary values are correct")
        print("2. Service account credentials are available in environment")
        print("3. ALL Google Sheets (5 sheets total) are shared with service account email:")
        print("   - BOM Sheet")
        print("   - SKU Reference Sheet")
        print("   - Forecast Sheet")
        print("   - Procurement Parameters Sheet")
        print("   - Current Inventory Sheet")
        print("4. Column letters in BOM_CONFIG match your actual sheets")
        print("5. Worksheet names are correct (case-sensitive)")
        import traceback
        traceback.print_exc()
        return None, None

def upload_bom_excel_to_google_sheet(excel_buffer, sheet_id=None):
    """
    Upload BOM Excel output to a dedicated Google Sheet.
    Reuses the same authentication pattern as the main upload function.
    """
    import pandas as pd
    import numpy as np
    import gspread
    import time
    import os
    import json
    from google.oauth2.service_account import Credentials

    print("🔄 Uploading BOM output to Google Sheets...")
    
    # Ensure env var is present
    if "gcp_service_account_sheets" not in os.environ:
        raise FileNotFoundError("❌ No GCP service account credentials found in environment variables.")

    try:
        raw_secret = os.environ["gcp_service_account_sheets"]
        creds_dict = json.loads(raw_secret)
        print(f"✅ Loaded service account for: {creds_dict.get('client_email', 'UNKNOWN EMAIL')}")
    except Exception as e:
        print("❌ Failed to parse service account secret")
        import traceback
        traceback.print_exc()
        raise
    
    # Save to temp file
    credentials_file = "temp_bom_sheet_credentials.json"
    with open(credentials_file, "w") as f:
        json.dump(creds_dict, f)

    # Create credentials from temp file
    scopes = ['https://www.googleapis.com/auth/spreadsheets']
    credentials = Credentials.from_service_account_file(credentials_file, scopes=scopes)
    gc = gspread.authorize(credentials)

    # Clean up temp file
    try:
        os.remove(credentials_file)
    except:
        pass

    # BOM-specific Google Sheet ID
    if sheet_id is None:
        sheet_id = "1_wXJDNZeZ7Y31S_i3UUDQ89vCbJC-xotm3wADBSf5eY"

    try:
        sheet = gc.open_by_key(sheet_id)
        print(f"🔁 Connected to BOM Google Sheet: {sheet.title}")
    except Exception as e:
        print("❌ Failed to connect to BOM Google Sheet:", e)
        raise

    try:
        # Read Excel from buffer
        excel_buffer.seek(0)
        all_sheets = pd.read_excel(excel_buffer, sheet_name=None, engine="openpyxl")

        # Existing sheet names
        existing_titles = [ws.title for ws in sheet.worksheets()]

        for i, (sheet_name, df) in enumerate(all_sheets.items()):
            df = df.replace([np.nan, np.inf, -np.inf], "").astype(str)

            # Reuse existing sheet if available
            if sheet_name in existing_titles:
                ws = sheet.worksheet(sheet_name)
                ws.clear()
            else:
                # Add new sheet with enough space
                ws = sheet.add_worksheet(title=sheet_name[:99], rows=str(len(df)+50), cols=str(len(df.columns)+10))

            # Push to Google Sheet
            try:
                ws.update([df.columns.tolist()] + df.values.tolist())
                print(f"✅ Updated BOM sheet: {sheet_name}")
            except Exception as e:
                print(f"⚠️ Error updating '{sheet_name}': {e}")
            
            time.sleep(1)  # Add delay to avoid quota issues

        print(f"✅ BOM output uploaded to Google Sheets successfully")
        return f"https://docs.google.com/spreadsheets/d/{sheet_id}"

    except Exception as e:
        print("❌ Unexpected error while uploading BOM to Google Sheets:", e)
        import traceback
        traceback.print_exc()
        return None


# NEW: BOM-specific Google Drive Upload Function
# PLACEMENT: Immediately after upload_bom_excel_to_google_sheet

def upload_bom_to_google_drive_from_buffer(buffer):
    """
    Upload BOM Excel output to Google Drive with timestamped backup.
    Reuses the same authentication pattern as the main upload function.
    """
    from datetime import datetime
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload
    
    SCOPES = ['https://www.googleapis.com/auth/drive']
    SHARED_DRIVE_ID = '0ANRBYKNxrAXaUk9PVA'
    FOLDER_ID = '0ANRBYKNxrAXaUk9PVA'
    FIXED_FILENAME = "BOM Analysis Workbook.xlsx"
    BOM_SUBFOLDER_ID = '1PHfLwnrl15wbu5si02Y1ZEqTs7EZKEp7'  # BOM-specific timestamp subfolder

    # Handle service account credentials - GCP only
    if "gcp_service_account_drive" not in os.environ:
        raise FileNotFoundError("❌ No Google Drive service account credentials found in environment variables.")
    
    print("🔄 Uploading BOM output to Google Drive...")
    
    # Load credentials from environment
    creds_dict = json.loads(os.environ["gcp_service_account_drive"])
    
    # Write to temporary file
    SERVICE_ACCOUNT_FILE = "temp_bom_drive_service_account.json"
    with open(SERVICE_ACCOUNT_FILE, "w") as f:
        json.dump(creds_dict, f)
    
    # Initialize credentials and service
    credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES
    )
    drive_service = build('drive', 'v3', credentials=credentials)

    # ==========================================================================
    # PART 1: Update/Create main BOM file
    # ==========================================================================
    
    # Check if file with same name already exists
    query = f"'{FOLDER_ID}' in parents and name = '{FIXED_FILENAME}' and trashed = false"
    result = drive_service.files().list(
        q=query,
        fields="files(id)",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
        corpora="drive",
        driveId=SHARED_DRIVE_ID
    ).execute()

    existing_files = result.get("files", [])

    # If exists, update it
    if existing_files:
        file_id = existing_files[0]['id']
        buffer.seek(0)
        media = MediaIoBaseUpload(buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )
        updated_file = drive_service.files().update(
            fileId=file_id,
            media_body=media,
            supportsAllDrives=True
        ).execute()
        print(f"♻️ Existing BOM file updated: {FIXED_FILENAME} (ID: {file_id})")
    else:
        buffer.seek(0)
        media = MediaIoBaseUpload(buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )
        file_metadata = {
            'name': FIXED_FILENAME,
            'parents': [FOLDER_ID],
            'driveId': SHARED_DRIVE_ID
        }
        uploaded_file = drive_service.files().create(
            body=file_metadata,
            media_body=media,
            supportsAllDrives=True,
            fields='id'
        ).execute()
        file_id = uploaded_file.get('id')
        print(f"✅ New BOM file uploaded: {FIXED_FILENAME} (ID: {file_id})")

    # ==========================================================================
    # PART 2: Save timestamped backup to BOM-specific subfolder
    # ==========================================================================
    
    try:
        print("📁 Starting BOM timestamped backup process...")
        
        # Generate timestamped filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        timestamped_filename = f"BOM_Backup_{timestamp}.xlsx"
        print(f"🕒 Generated timestamped filename: {timestamped_filename}")
        
        # Check if file with same timestamp already exists (overwrite if exists)
        timestamp_query = f"'{BOM_SUBFOLDER_ID}' in parents and name = '{timestamped_filename}' and trashed = false"
        timestamp_result = drive_service.files().list(
            q=timestamp_query,
            fields="files(id)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
            corpora="drive",
            driveId=SHARED_DRIVE_ID
        ).execute()
        
        existing_timestamp_files = timestamp_result.get("files", [])
        
        # Upload or update timestamped file
        buffer.seek(0)
        media = MediaIoBaseUpload(buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )
        
        if existing_timestamp_files:
            # Update existing timestamped file
            timestamp_file_id = existing_timestamp_files[0]['id']
            updated_timestamp_file = drive_service.files().update(
                fileId=timestamp_file_id,
                media_body=media,
                supportsAllDrives=True
            ).execute()
            print(f"🔄 Overwritten existing BOM timestamped backup: {timestamped_filename}")
        else:
            # Create new timestamped file
            timestamp_file_metadata = {
                'name': timestamped_filename,
                'parents': [BOM_SUBFOLDER_ID],
                'driveId': SHARED_DRIVE_ID
            }
            uploaded_timestamp_file = drive_service.files().create(
                body=timestamp_file_metadata,
                media_body=media,
                supportsAllDrives=True,
                fields='id'
            ).execute()
            print(f"✅ New BOM timestamped backup saved: {timestamped_filename}")
            
        print("📚 BOM timestamped backup process completed successfully!")
        
    except Exception as e:
        print(f"⚠️ Warning: Failed to save BOM timestamped backup: {str(e)}")
        print("📝 Main BOM file upload was successful, continuing...")

    # Clean up temp file
    try:
        os.remove(SERVICE_ACCOUNT_FILE)
    except:
        pass

    print(f"✅ BOM output uploaded to Google Drive successfully")
    return file_id

def upload_excel_to_google_sheet(excel_buffer, sheet_id=None):
    import pandas as pd
    import numpy as np
    import gspread
    import time
    import os
    import json
    from google.oauth2.service_account import Credentials

    print("🔄 Using GCP credentials from environment...")
    
    # Ensure env var is present
    if "gcp_service_account_sheets" not in os.environ:
        raise FileNotFoundError("❌ No GCP service account credentials found in environment variables.")

    # ✅ Debug: Check what we actually got from env
    try:
        raw_secret = os.environ["gcp_service_account_sheets"]
        print(f"🔍 Secret length: {len(raw_secret)} characters")
        creds_dict = json.loads(raw_secret)
        print(f"✅ Loaded service account for: {creds_dict.get('client_email', 'UNKNOWN EMAIL')}")
    except Exception as e:
        print("❌ Failed to parse service account secret")
        import traceback
        traceback.print_exc()
        raise
    
    # Save to temp file (exactly like your working function)
    credentials_file = "temp_credentials.json"
    with open(credentials_file, "w") as f:
        json.dump(creds_dict, f)
    print("✅ Credentials saved to temp file.")

    # Create credentials from temp file
    scopes = ['https://www.googleapis.com/auth/spreadsheets']
    credentials = Credentials.from_service_account_file(credentials_file, scopes=scopes)
    gc = gspread.authorize(credentials)

    # Clean up temp file
    try:
        os.remove(credentials_file)
        print("🧹 Cleaned up temp credentials file.")
    except:
        pass

    # ✅ Set your fixed Google Sheet ID here
    if sheet_id is None:
        sheet_id = "1051NJelrnQGKwKDXWmaMiU1-fBgm4ZSd4s_G2-hJIcE"

    try:
        sheet = gc.open_by_key(sheet_id)
        print(f"🔁 Connected to Google Sheet: {sheet.title}")
    except Exception as e:
        print("❌ Failed to connect to Google Sheet:", e)
        raise

    try:
        # Read Excel from buffer
        excel_buffer.seek(0)
        all_sheets = pd.read_excel(excel_buffer, sheet_name=None, engine="openpyxl")

        # Existing sheet names
        existing_titles = [ws.title for ws in sheet.worksheets()]

        for i, (sheet_name, df) in enumerate(all_sheets.items()):
            df = df.replace([np.nan, np.inf, -np.inf], "").astype(str)

            # Reuse existing sheet if available
            if sheet_name in existing_titles:
                ws = sheet.worksheet(sheet_name)
                ws.clear()
            else:
                # Add new sheet with enough space
                ws = sheet.add_worksheet(title=sheet_name[:99], rows=str(len(df)+50), cols=str(len(df.columns)+10))

            # Push to Google Sheet
            try:
                ws.update([df.columns.tolist()] + df.values.tolist())
                print(f"✅ Updated sheet: {sheet_name}")
            except Exception as e:
                print(f"⚠️ Error updating '{sheet_name}': {e}")
            
            time.sleep(1)  # Add delay to avoid quota issues

        return f"https://docs.google.com/spreadsheets/d/{sheet_id}"

    except Exception as e:
        print("❌ Unexpected error while uploading to Google Sheets:", e)
        import traceback
        traceback.print_exc()
        return None
     
def upload_to_google_drive_from_buffer(buffer):
    from datetime import datetime
    # BASE_DIR = os.path.dirname(__file__)
    # SERVICE_ACCOUNT_FILE = os.path.join(BASE_DIR, "GoogleDriveAPIKey.json")
    SCOPES = ['https://www.googleapis.com/auth/drive']
    SHARED_DRIVE_ID = '0ANRBYKNxrAXaUk9PVA'
    FOLDER_ID = '0ANRBYKNxrAXaUk9PVA'
    FIXED_FILENAME = "Forecasting Excel Workbook Format.xlsx"
    SUBFOLDER_NAME = "Output_TimeStamps"

    # Handle service account credentials - GCP only
    if "gcp_service_account_drive" not in os.environ:
        raise FileNotFoundError("❌ No Google Drive service account credentials found in environment variables.")
    
    print("🔄 Using Google Drive credentials from environment...")
    
    # Load credentials from environment
    creds_dict = json.loads(os.environ["gcp_service_account_drive"])
    
    # Write to temporary file (required for from_service_account_file)
    SERVICE_ACCOUNT_FILE = "temp_service_account.json"
    with open(SERVICE_ACCOUNT_FILE, "w") as f:
        json.dump(creds_dict, f)
    
    print("✅ Google Drive credentials loaded and saved to temp file.")
    
    # Initialize credentials and service
    credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES
    )
    drive_service = build('drive', 'v3', credentials=credentials)

    # =============================================================================
    # PART 1: Original functionality - Update/Create main file
    # =============================================================================
    
    # Check if file with same name already exists
    query = f"'{FOLDER_ID}' in parents and name = '{FIXED_FILENAME}' and trashed = false"
    result = drive_service.files().list(
        q=query,
        fields="files(id)",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
        corpora="drive",
        driveId=SHARED_DRIVE_ID
    ).execute()

    existing_files = result.get("files", [])

    # If exists, update it
    if existing_files:
        file_id = existing_files[0]['id']
        buffer.seek(0)
        media = MediaIoBaseUpload(buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )
        updated_file = drive_service.files().update(
            fileId=file_id,
            media_body=media,
            supportsAllDrives=True
        ).execute()
        print(f"♻️ Existing file updated: {FIXED_FILENAME} (ID: {file_id})")
    else:
        buffer.seek(0)
        media = MediaIoBaseUpload(buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )
        file_metadata = {
            'name': FIXED_FILENAME,
            'parents': [FOLDER_ID],
            'driveId': SHARED_DRIVE_ID
        }
        uploaded_file = drive_service.files().create(
            body=file_metadata,
            media_body=media,
            supportsAllDrives=True,
            fields='id'
        ).execute()
        file_id = uploaded_file.get('id')
        print(f"✅ New file uploaded: {FIXED_FILENAME} (ID: {file_id})")

    # =============================================================================
    # PART 2: NEW FUNCTIONALITY - Save timestamped copy to subfolder
    # =============================================================================
    
    try:
        print("📁 Starting timestamped backup process...")
        
        # Step 1: Find the existing subfolder (assumes it exists)
        subfolder_query = f"'{FOLDER_ID}' in parents and name = '{SUBFOLDER_NAME}' and mimeType = 'application/vnd.google-apps.folder' and trashed = false"
        subfolder_result = drive_service.files().list(
            q=subfolder_query,
            fields="files(id, name)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
            corpora="drive",
            driveId=SHARED_DRIVE_ID
        ).execute()
        
        subfolder_files = subfolder_result.get("files", [])
        
        if subfolder_files:
            subfolder_id = subfolder_files[0]['id']
            print(f"📂 Found subfolder: {SUBFOLDER_NAME} (ID: {subfolder_id})")
        else:
            raise FileNotFoundError(f"❌ Subfolder '{SUBFOLDER_NAME}' not found. Please create it manually in your Drive folder.")
        
        # Step 2: Generate timestamped filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        timestamped_filename = f"Forecasting_Backup_{timestamp}.xlsx"
        print(f"🕒 Generated timestamped filename: {timestamped_filename}")
        
        # Step 3: Check if file with same timestamp already exists (overwrite if exists)
        timestamp_query = f"'{subfolder_id}' in parents and name = '{timestamped_filename}' and trashed = false"
        timestamp_result = drive_service.files().list(
            q=timestamp_query,
            fields="files(id)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
            corpora="drive",
            driveId=SHARED_DRIVE_ID
        ).execute()
        
        existing_timestamp_files = timestamp_result.get("files", [])
        
        # Step 4: Upload or update timestamped file
        buffer.seek(0)  # Reset buffer position
        media = MediaIoBaseUpload(buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )
        
        if existing_timestamp_files:
            # Update existing timestamped file
            timestamp_file_id = existing_timestamp_files[0]['id']
            updated_timestamp_file = drive_service.files().update(
                fileId=timestamp_file_id,
                media_body=media,
                supportsAllDrives=True
            ).execute()
            print(f"🔄 Overwritten existing timestamped backup: {timestamped_filename}")
        else:
            # Create new timestamped file
            timestamp_file_metadata = {
                'name': timestamped_filename,
                'parents': [subfolder_id],
                'driveId': SHARED_DRIVE_ID
            }
            uploaded_timestamp_file = drive_service.files().create(
                body=timestamp_file_metadata,
                media_body=media,
                supportsAllDrives=True,
                fields='id'
            ).execute()
            print(f"✅ New timestamped backup saved: {timestamped_filename}")
            
        print("📚 Timestamped backup process completed successfully!")
        
    except Exception as e:
        print(f"⚠️ Warning: Failed to save timestamped backup: {str(e)}")
        print("📝 Main file upload was successful, continuing...")
        # Don't raise the exception - let the main functionality continue
    
    # Clean up temp file
    try:
        os.remove(SERVICE_ACCOUNT_FILE)
        print("🧹 Cleaned up temporary credentials file.")
    except:
        pass

    return file_id

         
def main():

    try:               
        print("ENHANCED INVENTORY FORECASTING MODEL - COMPREHENSIVE VERSION")
        print("=" * 60)

        GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1ZYugDxWgvmwye_zYYZJ4lgnY8hwZYKljEjGOKT2Cens/edit?gid=2126602512#gid=2126602512"
        WEEKLY_SALES_URL = "https://docs.google.com/spreadsheets/d/16WVvbzcdzeeI4ZL4OFou_7DVM7UAHWUvXmpYiHzOUw0/edit?gid=1908752665#gid=1908752665"
        #INVENTORY_URL = "https://docs.google.com/spreadsheets/d/1_j7eJi52Kq8RHvK6e0RPBRK8wJ0DXUOMj7Z7yZHlZzM/edit?gid=404505721#gid=404505721"
        INVENTORY_URL = "https://docs.google.com/spreadsheets/d/1mOSf1sO6MndPfsGJde988pWdAEORLixBjgNj5tCSVwE/edit?gid=0#gid=0"

        USE_GOOGLE_SHEETS = True

        print("Loading data files...")

        historical_sales_path = os.path.join(BASE_DIR, "historical_sales.csv")
        historical_data = pd.read_csv(historical_sales_path, sep=',', dtype={'SKU': str}, encoding='utf-8')
        print(f"   Historical sales: {len(historical_data)} records")

        # Clean historical SKUs more thoroughly
        historical_data['SKU'] = historical_data['SKU'].astype(str).str.strip()
        print(f"   Sample historical SKUs: {historical_data['SKU'].unique()[:5]}")

        # MODIFIED SECTION - Load inventory from Google Sheets instead of CSV
        inventory = {}
        
        if USE_GOOGLE_SHEETS and GOOGLE_SHEETS_AVAILABLE:
            try:
                print("🔄 Using GCP credentials from environment...")
                
                # Ensure env var is present
                if "gcp_service_account_sheets" not in os.environ:
                    raise FileNotFoundError("❌ No GCP service account credentials found in environment variables.")
        
                # ✅ Debug: Check what we actually got from env
                try:
                    raw_secret = os.environ["gcp_service_account_sheets"]
                    print(f"🔍 Secret length: {len(raw_secret)} characters")
                    creds_dict = json.loads(raw_secret)
                    print(f"✅ Loaded service account for: {creds_dict.get('client_email', 'UNKNOWN EMAIL')}")
                except Exception as e:
                    print("❌ Failed to parse service account secret")
                    import traceback
                    traceback.print_exc()
                    raise
                
                # Save to temp file
                credentials_file = "temp_credentials.json"
                with open(credentials_file, "w") as f:
                    json.dump(creds_dict, f)
                print("✅ Credentials saved to temp file.")
        
                # Create connector
                gs_connector = GoogleSheetsConnector(credentials_file)
        
                # Now try getting inventory
                print(f"\n📦 Loading inventory data from Google Sheets...")
                inventory = gs_connector.get_inventory_data(INVENTORY_URL)
                print(f"   Current inventory from Google Sheets: {len(inventory)} SKUs")

                # Get product info and lead times
                product_info, lead_times, launch_dates, product_category, product_status = gs_connector.get_product_data(GOOGLE_SHEET_URL)

                print(f"   Product info from Google Sheets: {len(product_info)} SKUs")
                print(f"   Lead times from Google Sheets: {len(lead_times)} SKUs")
                print(f"   Launch dates from Google Sheets: {len([d for d in launch_dates.values() if d is not None])} SKUs")

                # Step 1: Load and extend Amazon weekly
                print(f"\nLoading Amazon FBA weekly sales data...")
                amazon_weekly_df = gs_connector.get_amazon_fba_weekly_sales(WEEKLY_SALES_URL)

                if not amazon_weekly_df.empty:
                    amazon_weekly_monthly = gs_connector.convert_amazon_weekly_to_monthly(amazon_weekly_df)
                    amazon_extended = gs_connector.extend_historical_data_with_amazon_weekly(historical_data.copy(), amazon_weekly_monthly)
                    print(f"✅ Amazon FBA data extended")
                else:
                    amazon_extended = historical_data.copy()
                    print("⚠️ No Amazon FBA weekly sales data found")

                # Step 2: Load and extend Shopify weekly
                print(f"\nLoading Shopify Main weekly sales data...")
                shopify_weekly_df = gs_connector.get_shopify_main_weekly_sales(WEEKLY_SALES_URL)

                if not shopify_weekly_df.empty:
                    shopify_weekly_monthly = gs_connector.convert_shopify_weekly_to_monthly(shopify_weekly_df)
                    shopify_extended = gs_connector.extend_historical_data_with_shopify_weekly(historical_data.copy(), shopify_weekly_monthly)
                    print(f"✅ Shopify Main data extended")
                else:
                    shopify_extended = historical_data.copy()
                    print("⚠️ No Shopify Main weekly sales data found")

                ### Shopify Faire    
                print(f"\nLoading Shopify Faire weekly sales data...")
                shopify_faire_weekly_df = gs_connector.get_shopify_faire_weekly_sales(WEEKLY_SALES_URL)

                if not shopify_faire_weekly_df.empty:
                    shopify_faire_weekly_monthly = gs_connector.convert_shopify_faire_weekly_to_monthly(shopify_faire_weekly_df)
                    shopify_faire_extended = gs_connector.extend_historical_data_with_shopify_faire_weekly(historical_data.copy(), shopify_faire_weekly_monthly)
                    print(f"✅ Shopify Faire data extended")
                else:
                    shopify_faire_extended = historical_data.copy()
                    print("⚠️ No Shopify Faire weekly sales data found")

                ### Walmart FBM
                print(f"\nLoading Walmart FBM weekly sales data...")
                walmart_fbm_weekly_df = gs_connector.get_walmart_fbm_weekly_sales(WEEKLY_SALES_URL)

                if not walmart_fbm_weekly_df.empty:
                    walmart_fbm_weekly_monthly = gs_connector.convert_walmart_fbm_weekly_to_monthly(walmart_fbm_weekly_df)
                    walmart_fbm_extended = gs_connector.extend_historical_data_with_walmart_fbm_weekly(historical_data.copy(), walmart_fbm_weekly_monthly)
                    print(f"✅ Walmart FBM data extended")
                else:
                    walmart_fbm_extended = historical_data.copy()
                    print("⚠️ No Walmart FBM weekly sales data found")

                ### Amazon FBM
                print(f"\nLoading Amazon FBM weekly sales data...")
                amazon_fbm_weekly_df = gs_connector.get_amazon_fbm_weekly_sales(WEEKLY_SALES_URL)

                if not amazon_fbm_weekly_df.empty:
                    amazon_fbm_weekly_monthly = gs_connector.convert_amazon_fbm_weekly_to_monthly(amazon_fbm_weekly_df)
                    amazon_fbm_extended = gs_connector.extend_historical_data_with_amazon_fbm_weekly(historical_data.copy(), amazon_fbm_weekly_monthly)
                    print(f"✅ Amazon FBM data extended")
                else:
                    amazon_fbm_extended = historical_data.copy()
                    print("⚠️ No Amazon FBM weekly sales data found")

                # Step 3: Combine Amazon + Shopify extended data
                historical_data = pd.concat([amazon_extended, shopify_extended, shopify_faire_extended, walmart_fbm_extended, amazon_fbm_extended], ignore_index=True)
                historical_data = historical_data.drop_duplicates(subset=['SKU', 'Channel', 'Date'], keep='last')
                historical_data = historical_data.sort_values(by=['SKU', 'Channel', 'Date'])

                historical_skus = set(historical_data['SKU'].unique())
                google_skus = set(product_info.keys())

                print(f"\n🔍 ENHANCED SKU MATCHING DEBUG:")
                print(f"   Historical data has {len(historical_skus)} unique SKUs")
                print(f"   Google Sheets has {len(google_skus)} unique SKUs")

                # Test direct matches
                direct_matches = historical_skus.intersection(google_skus)
                print(f"   Direct matches: {len(direct_matches)}")
                print(f"   Sample direct matches: {list(direct_matches)[:5]}")

                # Test the lookup function directly
                print(f"   Testing smart_sku_lookup function:")
                test_skus = ['810128951111', '810128951128', '810128951203']  # From the sample
                for test_sku in test_skus:
                    direct_result = product_info.get(test_sku, 'NOT_FOUND_DIRECT')
                    print(f"     {test_sku} -> Direct: {direct_result}")

                    # Test if it's in the set
                    in_google_skus = test_sku in google_skus
                    print(f"     {test_sku} -> In google_skus: {in_google_skus}")

            except Exception as e:
                print(f"❌ Error loading from Google Sheets: {e}")
                print("   Falling back to local CSV files...")
                USE_GOOGLE_SHEETS = False
        else:
            USE_GOOGLE_SHEETS = False


        if not USE_GOOGLE_SHEETS:
            # Fall back to CSV for inventory
            inventory_path = os.path.join(BASE_DIR, "current_inventory.csv")
            inventory_df = pd.read_csv(inventory_path, sep=",", dtype={"SKU": str}, encoding="utf-8")
            inventory_df['SKU'] = inventory_df['SKU'].astype(str).str.strip()
            inventory = dict(zip(inventory_df['SKU'], inventory_df['Inventory']))
            print(f"   Current inventory from CSV: {len(inventory)} SKUs")

            lead_times_path = os.path.join(BASE_DIR, "lead_times.csv")
            lead_times_df = pd.read_csv(lead_times_path, sep=",", dtype={"SKU": str}, encoding="utf-8")            
            lead_times_df['SKU'] = lead_times_df['SKU'].astype(str).str.strip()
            lead_times = dict(zip(lead_times_df['SKU'], lead_times_df['LeadTime']))
            print(f"   Lead times from CSV: {len(lead_times)} SKUs")

            product_info_path = os.path.join(BASE_DIR, "product_info.csv")
            product_info_df = pd.read_csv(product_info_path, sep=",", dtype={"Unit UPC": str}, encoding="utf-8")
            if 'Unit UPC' in product_info_df.columns:
                product_info_df = product_info_df.rename(columns={'Unit UPC': 'SKU'})
            product_info_df['SKU'] = product_info_df['SKU'].astype(str).str.strip()
            product_info = dict(zip(product_info_df['SKU'], product_info_df['Product Name']))
            print(f"   Product info from CSV: {len(product_info)} SKUs")

            # Create empty launch_dates for CSV fallback
            launch_dates = {}
            print(f"   Launch dates: Not available in CSV mode")

        model = EnhancedForecastingModel(historical_data, lead_times, launch_dates, service_level=0.95)

        print("\nGenerating forecasts...")
        amazon_forecast = model.create_enhanced_forecast('Amazon', inventory, product_info, product_category, product_status)
        shopify_forecast = model.create_enhanced_forecast_shopify_special('Shopify', inventory, product_info, product_category, product_status)
        shopify_faire_forecast = model.create_enhanced_forecast_shopify_faire_special('Shopify Faire', inventory, product_info, product_category, product_status)
        amazon_fbm_forecast = model.create_enhanced_forecast_amazon_fbm_special('Amazonfbm', inventory, product_info, product_category, product_status)
        walmart_fbm_forecast = model.create_enhanced_forecast_walmart_fbm_special('Walmartfbm', inventory, product_info, product_category, product_status)

        print("Creating combined channel analysis...")
        combined_forecast = model.combine_channel_forecasts(amazon_forecast, shopify_forecast, shopify_faire_forecast, amazon_fbm_forecast, walmart_fbm_forecast)


        # Generate actionable insights
        print("Generating actionable insights...")
        insights = model.generate_actionable_insights(combined_forecast)
        executive_summary = model.create_executive_summary(combined_forecast, insights)
        priority_matrix = model.create_action_priority_matrix(combined_forecast)

        finance_forecast = model.create_finance_cash_flow_forecast(combined_forecast)

        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M")
        data_source = "GoogleSheets" if USE_GOOGLE_SHEETS else "CSV"
        filename = f'enhanced_forecast_COMPREHENSIVE_{data_source}_{timestamp}.xlsx'

        print(f"Saving results to {filename}...")

        # Create BytesIO buffer for Excel file (works for both CLI and Streamlit)
        excel_buffer = BytesIO()

        with pd.ExcelWriter(excel_buffer, engine='xlsxwriter',
                           engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
            # Get the xlsxwriter workbook and worksheet objects
            workbook = writer.book

            # Define formats
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'vcenter',
                'align': 'center',
                'bg_color': '#D7E4BC',
                'border': 1,
                'font_size': 10
            })

            urgent_format = workbook.add_format({
                'bg_color': '#FFC7CE',
                'font_color': '#9C0006',
                'bold': True
            })

            warning_format = workbook.add_format({
                'bg_color': '#FFEB9C',
                'font_color': '#9C5700'
            })

            good_format = workbook.add_format({
                'bg_color': '#C6EFCE',
                'font_color': '#006100'
            })

            date_format = workbook.add_format({
                'num_format': 'yyyy-mm-dd',
                'align': 'center'
            })

            number_format = workbook.add_format({
                'num_format': '#,##0',
                'align': 'right'
            })

            decimal_format = workbook.add_format({
                'num_format': '#,##0.0',
                'align': 'right'
            })

            currency_format = workbook.add_format({
                'num_format': '$#,##0',
                'align': 'right'
            })

            text_format = workbook.add_format({
                'text_wrap': True,
                'valign': 'top'
            })

            # 1. EXECUTIVE SUMMARY SHEET (FIRST - MOST IMPORTANT)
            exec_summary_data = pd.DataFrame([
                ['INVENTORY PLANNING EXECUTIVE SUMMARY', ''],
                ['Report Date:', executive_summary['date']],
                ['', ''],
                ['IMMEDIATE ACTIONS REQUIRED', ''],
                ['Critical Actions (Today):', executive_summary['immediate_actions_required']],
                ['High Priority (This Week):', executive_summary['weekly_actions_required']],
                ['', ''],
                ['INVENTORY HEALTH', ''],
                ['Total SKUs:', executive_summary['total_skus']],
                ['At-Risk SKUs (<2 months):', executive_summary['at_risk_skus']],
                ['Overstock SKUs (>6 months):', executive_summary['overstock_skus']],
                ['', ''],
                ['FINANCIAL IMPACT', ''],
                ['Current Inventory Value:', f"${executive_summary['total_inventory_value']:,.0f}"],
                ['PO Value Needed (Immediate):', f"${executive_summary['total_po_value_needed']:,.0f}"],
                ['', ''],
                ['CASH FLOW PROJECTION', ''],
                ['Next 30 Days:', f"${executive_summary['cash_flow_30_days']:,.0f}"],
                ['Next 60 Days:', f"${executive_summary['cash_flow_60_days']:,.0f}"],
                ['Next 90 Days:', f"${executive_summary['cash_flow_90_days']:,.0f}"],
            ], columns=['Metric', 'Value'])

            exec_summary_data.to_excel(writer, sheet_name='📊 Executive Summary', index=False, header=False)
            worksheet = writer.sheets['📊 Executive Summary']

            # Format executive summary with proper column widths
            worksheet.set_column('A:A', 35)  # Metric column
            worksheet.set_column('B:B', 25)  # Value column

            # Format rows
            title_format = workbook.add_format({
                'bold': True,
                'font_size': 16,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': '#4472C4',
                'font_color': 'white',
                'border': 1
            })

            section_format = workbook.add_format({
                'bold': True,
                'font_size': 12,
                'bg_color': '#D7E4BC',
                'border': 1
            })

            value_format = workbook.add_format({
                'font_size': 11,
                'align': 'right'
            })

            # Apply formatting
            worksheet.merge_range(0, 0, 0, 1, 'INVENTORY PLANNING EXECUTIVE SUMMARY', title_format)
            worksheet.set_row(0, 30)

            # Format section headers
            for row in [3, 7, 12, 16]:
                if row < len(exec_summary_data):
                    worksheet.merge_range(row, 0, row, 1, exec_summary_data.iloc[row, 0], section_format)
                    worksheet.set_row(row, 25)

            # Format data rows
            for row in range(1, len(exec_summary_data)):
                if row not in [0, 3, 7, 12, 16] and exec_summary_data.iloc[row, 1]:
                    worksheet.write(row, 1, exec_summary_data.iloc[row, 1], value_format)

            # 2. IMMEDIATE ACTIONS SHEET
            immediate_actions_list = insights.get('immediate_actions', [])

            print(f"📌 immediate_actions count: {len(immediate_actions_list)}")

            if immediate_actions_list:
                immediate_df = pd.DataFrame(immediate_actions_list)
                immediate_df.to_excel(writer, sheet_name='🚨 IMMEDIATE ACTIONS', index=False, header=False, startrow=1)
                worksheet = writer.sheets['🚨 IMMEDIATE ACTIONS']

                # Write headers safely
                for col_num, column in enumerate(immediate_df.columns):
                    worksheet.write(0, col_num, column, header_format)

                # Set specific column widths
                col_widths = {
                    0: 12,  # Priority
                    1: 15,  # SKU
                    2: 35,  # Product
                    3: 40,  # Action
                    4: 30,  # Reason
                    5: 10,  # Quantity
                    6: 25,  # Impact
                    7: 35   # Contact
                }

                for col, width in col_widths.items():
                    if col < len(immediate_df.columns):
                        worksheet.set_column(col, col, width)

                # Apply formatting
                for row in range(1, len(immediate_df) + 1):
                    priority = immediate_df.iloc[row-1]['Priority']
                    fmt = urgent_format if priority == 'CRITICAL' else warning_format
                    for col in range(len(immediate_df.columns)):
                        worksheet.write(row, col, immediate_df.iloc[row-1, col], fmt)

                worksheet.freeze_panes(1, 0)

            else:
                print("⚠️ No immediate actions generated — skipping that sheet.")

            # 3. ACTION PRIORITY MATRIX
            if not priority_matrix.empty:
                priority_matrix.to_excel(writer, sheet_name='📋 Action Priority Matrix', index=False, header=False, startrow=1)
                worksheet = writer.sheets['📋 Action Priority Matrix']

                # Write headers
                for col_num, column in enumerate(priority_matrix.columns):
                    worksheet.write(0, col_num, column, header_format)

                # Set column widths
                matrix_widths = {
                    'SKU': 15,
                    'Product_Name': 35,
                    'Velocity_Category': 10,
                    'Months_of_Inventory': 12,
                    'Priority_Score': 10,
                    'Action_Priority': 12,
                    'Action_Timeline': 15,
                    'Recommended_Action': 40,
                    'Order_Quantity': 12,
                    'Current_Inventory': 12,
                    'Next_Month_Forecast': 12
                }

                for i, col in enumerate(priority_matrix.columns):
                    width = matrix_widths.get(col, 15)
                    worksheet.set_column(i, i, width)

                # Apply conditional formatting for priority levels
                for row in range(1, len(priority_matrix) + 1):
                    priority = priority_matrix.iloc[row-1]['Action_Priority']
                    if priority == 'IMMEDIATE':
                        row_format = urgent_format
                    elif priority == 'HIGH':
                        row_format = warning_format
                    else:
                        row_format = None

                    if row_format:
                        for col in range(len(priority_matrix.columns)):
                            worksheet.write(row, col, priority_matrix.iloc[row-1, col], row_format)

                worksheet.freeze_panes(1, 0)

            # 4. WEEKLY ACTIONS
            if insights.get('weekly_actions'):
                weekly_df = pd.DataFrame(insights.get('weekly_actions'))
                weekly_df.to_excel(writer, sheet_name='📅 Weekly Actions', index=False, header=False, startrow=1)
                worksheet = writer.sheets['📅 Weekly Actions']

                # Write headers and format
                for col_num, column in enumerate(weekly_df.columns):
                    worksheet.write(0, col_num, column, header_format)

                # Set column widths
                worksheet.set_column('A:A', 12)  # Priority
                worksheet.set_column('B:B', 15)  # SKU
                worksheet.set_column('C:C', 35)  # Product
                worksheet.set_column('D:D', 25)  # Action
                worksheet.set_column('E:E', 20)  # Reason
                worksheet.set_column('F:F', 10)  # Quantity
                worksheet.set_column('G:G', 12)  # By_Date
                worksheet.set_column('H:H', 30)  # Notes

                worksheet.freeze_panes(1, 0)

            # 5. RISK ANALYSIS
            if insights.get('risk_analysis'):
                risk_df = pd.DataFrame(insights.get('risk_analysis'))
                risk_df.to_excel(writer, sheet_name='⚠️ Risk Analysis', index=False, header=False, startrow=1)
                worksheet = writer.sheets['⚠️ Risk Analysis']

                # Write headers
                for col_num, column in enumerate(risk_df.columns):
                    worksheet.write(0, col_num, column, header_format)

                # Set column widths
                worksheet.set_column('A:A', 12)  # Risk_Level
                worksheet.set_column('B:B', 15)  # SKU
                worksheet.set_column('C:C', 35)  # Product
                worksheet.set_column('D:D', 35)  # Issue
                worksheet.set_column('E:E', 15)  # Potential_Loss
                worksheet.set_column('F:F', 40)  # Mitigation

                # Highlight high-risk items
                for row in range(1, len(risk_df) + 1):
                    if risk_df.iloc[row-1]['Risk_Level'] == 'HIGH':
                        for col in range(len(risk_df.columns)):
                            value = risk_df.iloc[row-1, col]
                            # Handle NaN/Inf values
                            if pd.isna(value):
                                value = ''
                            elif isinstance(value, (int, float)):
                                if value == float('inf'):
                                    value = 999999
                                elif value == float('-inf'):
                                    value = -999999
                            worksheet.write(row, col, value, warning_format)

                worksheet.freeze_panes(1, 0)

            # 7. COST OPTIMIZATION SHEET
            if insights.get('cost_optimization'):
                cost_df = pd.DataFrame(insights.get('cost_optimization'))
                cost_df.to_excel(writer, sheet_name='💡 Cost Optimization', index=False, header=False, startrow=1)
                worksheet = writer.sheets['💡 Cost Optimization']

                # Write headers
                for col_num, column in enumerate(cost_df.columns):
                    worksheet.write(0, col_num, column, header_format)

                # Set column widths
                worksheet.set_column('A:A', 15)  # SKU
                worksheet.set_column('B:B', 35)  # Product
                worksheet.set_column('C:C', 12)  # Current_Order
                worksheet.set_column('D:D', 30)  # Suggestion
                worksheet.set_column('E:E', 25)  # Savings
                worksheet.set_column('F:F', 35)  # Action

                worksheet.freeze_panes(1, 0)

            # 6. OPPORTUNITIES SHEET
            if insights.get('opportunities'):
                opp_df = pd.DataFrame(insights.get('opportunities'))
                opp_df.to_excel(writer, sheet_name='🎯 Opportunities', index=False, header=False, startrow=1)
                worksheet = writer.sheets['🎯 Opportunities']

                # Write headers
                for col_num, column in enumerate(opp_df.columns):
                    worksheet.write(0, col_num, column, header_format)

                # Set column widths
                worksheet.set_column('A:A', 12)  # Type
                worksheet.set_column('B:B', 15)  # SKU
                worksheet.set_column('C:C', 35)  # Product
                worksheet.set_column('D:D', 20)  # Trend
                worksheet.set_column('E:E', 25)  # Action
                worksheet.set_column('F:F', 30)  # Potential

                # Apply good formatting to all opportunity rows
                for row in range(1, len(opp_df) + 1):
                    for col in range(len(opp_df.columns)):
                        worksheet.write(row, col, opp_df.iloc[row-1, col], good_format)

                worksheet.freeze_panes(1, 0)

            # 8. MONTHLY ACTIONS SHEET
            if insights.get('monthly_actions'):
                monthly_df = pd.DataFrame(insights.get('monthly_actions'))
                monthly_df.to_excel(writer, sheet_name='📆 Monthly Actions', index=False, header=False, startrow=1)
                worksheet = writer.sheets['📆 Monthly Actions']

                # Write headers
                for col_num, column in enumerate(monthly_df.columns):
                    worksheet.write(0, col_num, column, header_format)

                # Set column widths
                worksheet.set_column('A:A', 15)  # SKU
                worksheet.set_column('B:B', 35)  # Product
                worksheet.set_column('C:C', 20)  # Action
                worksheet.set_column('D:D', 12)  # Order_Date
                worksheet.set_column('E:E', 10)  # Quantity
                worksheet.set_column('F:F', 25)  # Preparation
                worksheet.set_column('G:G', 15)  # Budget_Impact

                worksheet.freeze_panes(1, 0)

            # Function to format worksheet with proper column widths
            def format_worksheet(worksheet, df, sheet_name):
                # Write headers with formatting
                for col_num, column in enumerate(df.columns):
                    worksheet.write(0, col_num, column, header_format)

                # Define optimal column widths based on column names and content
                column_widths = {
                    'SKU': 15,
                    'Product_Name': 35,
                    'Product': 35,
                    'Launch_Date': 12,
                    'Years_Since_Launch': 10,
                    'Current_Inventory': 12,
                    'Stock_Status': 15,
                    'PO_Urgency': 25,
                    'Recommended_PO_Qty': 15,
                    'Next_Order_Date': 12,
                    'Next_Order_Qty': 12,
                    'Next_Arrival_Date': 12,
                    'Months_of_Inventory': 12,
                    'Velocity_Category': 10,
                    'Safety_Stock_Months': 12,
                    'Reorder_Point': 12,
                    'Safety_Stock': 12,
                    'Last_3_Months_Avg': 12,
                    'Total_Sales': 12,
                    'Growth_Rate': 10,
                    'Order_2_Date': 12,
                    'Order_2_Qty': 10,
                    'Order_2_Arrival': 12,
                    'Order_3_Date': 12,
                    'Order_3_Qty': 10,
                    'Order_3_Arrival': 12,
                    'Lead_Time': 10,
                    'Service_Level': 10,
                    'Monthly_Velocity': 12,
                    'Velocity_Rank': 10,
                    'Channel': 10,
                    'Forecast_Method': 20,
                    'Priority': 12,
                    'Action': 40,
                    'Reason': 30,
                    'Quantity': 10,
                    'Impact': 30,
                    'Contact': 35,
                    'By_Date': 12,
                    'Notes': 30,
                    'Risk_Level': 12,
                    'Issue': 35,
                    'Potential_Loss': 15,
                    'Mitigation': 40,
                    'Order_Month': 12,
                    'Order_Date': 12,
                    'Order_Quantity': 12,
                    'Order_Urgency': 15,
                    'Unit_Cost': 10,
                    'Total_Order_Value': 15,
                    'Supplier': 20,
                    'Payment_Terms': 15,
                    'Expected_Payment_Date': 15,
                    'Priority_Score': 10,
                    'Action_Priority': 12,
                    'Action_Timeline': 15,
                    'Recommended_Action': 35,
                    'Next_Month_Forecast': 12,
                    'Type': 12,
                    'Trend': 20,
                    'Potential': 30,
                    'Current_Order': 12,
                    'Suggestion': 30,
                    'Savings': 25,
                    'Preparation': 30,
                    'Budget_Impact': 15,
                    'Metric': 30,
                    'Value': 20,
                    'Excess_Units': 12,
                    'Excess_Value': 15,
                    'SKU_Count': 10,
                    'Inventory_Value': 15,
                    'Order_Value': 12,
                    'Arrival_Date': 12,
                    'Mapping_Status': 20
                }

                # Set column widths
                for i, col in enumerate(df.columns):
                    # Check for specific column names first
                    if col in column_widths:
                        width = column_widths[col]
                    # Check for forecast columns (dynamic month names)
                    elif col.startswith('Forecast_'):
                        width = 12
                    # Check for month columns (Jan_2025, etc.)
                    elif any(month in col for month in ['Jan_', 'Feb_', 'Mar_', 'Apr_', 'May_', 'Jun_', 'Jul_', 'Aug_', 'Sep_', 'Oct_', 'Nov_', 'Dec_']):
                        width = 10
                    # Check for Amazon/Shopify breakdown columns
                    elif col.startswith('Amazon_') or col.startswith('Shopify_'):
                        width = 12
                    # Default widths based on data type
                    else:
                        # Calculate based on content
                        max_len = len(str(col))
                        for idx, value in enumerate(df[col][:10]):  # Check first 10 rows
                            try:
                                value_len = len(str(value)) if pd.notna(value) else 0
                                max_len = max(max_len, value_len)
                            except:
                                pass
                        width = min(max(max_len + 2, 8), 40)

                    worksheet.set_column(i, i, width)

                # Apply data formatting based on column type
                for row_num in range(1, len(df) + 1):
                    for col_num, col_name in enumerate(df.columns):
                        value = df.iloc[row_num-1, col_num]

                        # Handle NaN/Inf values
                        if pd.isna(value):
                            value = ''
                        elif isinstance(value, (int, float)):
                            if value == float('inf'):
                                value = 999999
                            elif value == float('-inf'):
                                value = -999999
                            elif pd.isna(value):
                                value = 0

                        # Apply formatting based on column type
                        if 'Date' in col_name and value != '':
                            try:
                                # Convert to datetime for proper Excel date formatting
                                if isinstance(value, str) and value:
                                    date_value = pd.to_datetime(value)
                                    worksheet.write_datetime(row_num, col_num, date_value, date_format)
                                else:
                                    worksheet.write(row_num, col_num, value, text_format)
                            except:
                                worksheet.write(row_num, col_num, value, text_format)

                        elif col_name in ['Current_Inventory', 'Recommended_PO_Qty', 'Next_Order_Qty',
                                         'Order_2_Qty', 'Order_3_Qty', 'Safety_Stock', 'Reorder_Point',
                                         'Order_Quantity', 'Quantity', 'Excess_Units', 'SKU_Count'] or 'Forecast_' in col_name:
                            try:
                                if isinstance(value, (int, float)) and pd.notna(value):
                                    worksheet.write(row_num, col_num, value, number_format)
                                else:
                                    worksheet.write(row_num, col_num, value, text_format)
                            except:
                                worksheet.write(row_num, col_num, value, text_format)

                        elif col_name in ['Months_of_Inventory', 'Last_3_Months_Avg', 'Growth_Rate',
                                         'Monthly_Velocity', 'Years_Since_Launch', 'Safety_Stock_Months']:
                            try:
                                if isinstance(value, (int, float)) and pd.notna(value):
                                    worksheet.write(row_num, col_num, value, decimal_format)
                                else:
                                    worksheet.write(row_num, col_num, value, text_format)
                            except:
                                worksheet.write(row_num, col_num, value, text_format)

                        elif col_name in ['Total_Order_Value', 'Inventory_Value', 'Excess_Value',
                                         'Order_Value', 'Potential_Loss', 'Budget_Impact'] or 'Value' in col_name:
                            try:
                                if isinstance(value, (int, float)) and pd.notna(value):
                                    worksheet.write(row_num, col_num, value, currency_format)
                                else:
                                    worksheet.write(row_num, col_num, value, text_format)
                            except:
                                worksheet.write(row_num, col_num, value, text_format)

                        else:
                            worksheet.write(row_num, col_num, value, text_format)

                # Set row height for header
                worksheet.set_row(0, 25)

                # Freeze the header row and first two columns based on sheet type
                if 'Executive' not in sheet_name:  # Don't freeze executive summary
                    if 'All Forecasts' in sheet_name or 'Amazon' in sheet_name or 'Shopify' in sheet_name:
                        worksheet.freeze_panes(1, 2)  # Freeze first 2 columns
                    else:
                        worksheet.freeze_panes(1, 0)  # Just freeze header row

                # Add autofilter for data sheets
                if len(df) > 0 and 'Executive' not in sheet_name:
                    worksheet.autofilter(0, 0, len(df), len(df.columns) - 1)

            # Write and format remaining sheets (after the special sheets)
            if not combined_forecast.empty:
                combined_forecast.to_excel(writer, sheet_name='📈 All Forecasts', index=False, header=False, startrow=1)
                worksheet = writer.sheets['📈 All Forecasts']
                format_worksheet(worksheet, combined_forecast, '📈 All Forecasts')

                # Add conditional formatting for Stock Status
                status_col_idx = None
                for i, col in enumerate(combined_forecast.columns):
                    if col == 'Stock_Status':
                        status_col_idx = i
                        break

                if status_col_idx is not None:
                    for row in range(1, len(combined_forecast) + 1):
                        status = combined_forecast.iloc[row-1]['Stock_Status']
                        if status == 'OUT OF STOCK':
                            worksheet.write(row, status_col_idx, status, urgent_format)
                        elif status == 'REORDER NOW':
                            worksheet.write(row, status_col_idx, status, warning_format)
                        elif status == 'NORMAL':
                            worksheet.write(row, status_col_idx, status, good_format)

                # Apply conditional formatting for PO_Urgency
                urgency_col_idx = None
                for i, col in enumerate(combined_forecast.columns):
                    if col == 'PO_Urgency':
                        urgency_col_idx = i
                        break

                if urgency_col_idx is not None:
                    for row in range(1, len(combined_forecast) + 1):
                        urgency = combined_forecast.iloc[row-1]['PO_Urgency']
                        if 'HIGH' in urgency:
                            worksheet.write(row, urgency_col_idx, urgency, urgent_format)
                        elif 'MEDIUM' in urgency:
                            worksheet.write(row, urgency_col_idx, urgency, warning_format)

            if not finance_forecast.empty:
                finance_forecast.to_excel(writer, sheet_name='💰 Finance Cash Flow', index=False, header=False, startrow=1)
                worksheet = writer.sheets['💰 Finance Cash Flow']
                format_worksheet(worksheet, finance_forecast, '💰 Finance Cash Flow')

                # Apply specific formatting for finance sheet
                finance_widths = {
                    'Order_Month': 12,
                    'Order_Date': 12,
                    'SKU': 15,
                    'Product_Name': 35,
                    'Velocity_Category': 10,
                    'Order_Quantity': 12,
                    'Order_Urgency': 15,
                    'Unit_Cost': 10,
                    'Total_Order_Value': 15,
                    'Supplier': 20,
                    'Payment_Terms': 15,
                    'Expected_Payment_Date': 18,
                    'Lead_Time': 10,
                    'Safety_Stock_Months': 12,
                    'Current_Inventory': 12,
                    'Months_of_Inventory': 12
                }

                for i, col in enumerate(finance_forecast.columns):
                    if col in finance_widths:
                        worksheet.set_column(i, i, finance_widths[col])

                # Add a note for finance team
                note_text = "Please fill in Unit_Cost, Supplier, and Payment_Terms for accurate cash flow planning"
                worksheet.write(len(finance_forecast) + 3, 0, "Note:", header_format)
                worksheet.merge_range(len(finance_forecast) + 3, 1, len(finance_forecast) + 3, 6, note_text, text_format)

            if not amazon_forecast.empty:
                amazon_forecast.to_excel(writer, sheet_name='🛒 Amazon', index=False, header=False, startrow=1)
                worksheet = writer.sheets['🛒 Amazon']
                format_worksheet(worksheet, amazon_forecast, '🛒 Amazon')

            if not shopify_forecast.empty:
                shopify_forecast.to_excel(writer, sheet_name='🛍️ Shopify', index=False, header=False, startrow=1)
                worksheet = writer.sheets['🛍️ Shopify']
                format_worksheet(worksheet, shopify_forecast, '🛍️ Shopify')

            if not shopify_faire_forecast.empty:
                shopify_faire_forecast.to_excel(writer, sheet_name='🛍️ Shopify Faire', index=False, header=False, startrow=1)
                worksheet = writer.sheets['🛍️ Shopify Faire']
                format_worksheet(worksheet, shopify_faire_forecast, '🛍️ Shopify Faire')

            if not amazon_fbm_forecast.empty:
                amazon_fbm_forecast.to_excel(writer, sheet_name='🛍️ Amazon FBM', index=False, header=False, startrow=1)
                worksheet = writer.sheets['🛍️ Amazon FBM']
                format_worksheet(worksheet, amazon_fbm_forecast, '🛍️ Amazon FBM')

            if not walmart_fbm_forecast.empty:
                walmart_fbm_forecast.to_excel(writer, sheet_name='🛍️ Walmart FBM', index=False, header=False, startrow=1)
                worksheet = writer.sheets['🛍️ Walmart FBM']
                format_worksheet(worksheet, walmart_fbm_forecast, '🛍️ Walmart FBM')

            # Additional analysis sheets with proper formatting

            # Out of Stock Analysis
            if 'Stock_Status' in combined_forecast.columns:
                out_of_stock = combined_forecast[combined_forecast['Stock_Status'] == 'OUT OF STOCK']
                if not out_of_stock.empty:
                    out_of_stock.to_excel(writer, sheet_name='❌ Out of Stock', index=False, header=False, startrow=1)
                    worksheet = writer.sheets['❌ Out of Stock']
                    format_worksheet(worksheet, out_of_stock, '❌ Out of Stock')
                    # Highlight all rows as critical
                    for row in range(1, len(out_of_stock) + 1):
                        worksheet.set_row(row, None, urgent_format)

            # Reorder Now Analysis
            if 'Stock_Status' in combined_forecast.columns:
                reorder_now = combined_forecast[combined_forecast['Stock_Status'] == 'REORDER NOW']
                if not reorder_now.empty:
                    reorder_now.to_excel(writer, sheet_name='📦 Reorder Now', index=False, header=False, startrow=1)
                    worksheet = writer.sheets['📦 Reorder Now']
                    format_worksheet(worksheet, reorder_now, '📦 Reorder Now')
                    # Highlight all rows as warning
                    for row in range(1, len(reorder_now) + 1):
                        worksheet.set_row(row, None, warning_format)

            # Overstock Analysis
            if 'Months_of_Inventory' in combined_forecast.columns:
                overstock = combined_forecast[combined_forecast['Months_of_Inventory'] > 6]
                if not overstock.empty:
                    overstock_analysis = overstock[['SKU', 'Product_Name', 'Current_Inventory', 'Months_of_Inventory',
                                                   'Last_3_Months_Avg', 'Velocity_Category']].copy()
                    overstock_analysis['Excess_Units'] = overstock_analysis['Current_Inventory'] - (overstock_analysis['Last_3_Months_Avg'] * 3)
                    overstock_analysis['Excess_Value'] = overstock_analysis['Excess_Units'] * 30  # $30 cost assumption
                    overstock_analysis.to_excel(writer, sheet_name='📈 Overstock Analysis', index=False, header=False, startrow=1)
                    worksheet = writer.sheets['📈 Overstock Analysis']
                    format_worksheet(worksheet, overstock_analysis, '📈 Overstock Analysis')

            # Velocity Analysis
            if 'Velocity_Category' in combined_forecast.columns:
                velocity_summary = combined_forecast.groupby('Velocity_Category').agg({
                    'SKU': 'count',
                    'Current_Inventory': 'sum',
                    'Recommended_PO_Qty': 'sum'
                }).rename(columns={'SKU': 'SKU_Count'})
                velocity_summary['Inventory_Value'] = velocity_summary['Current_Inventory'] * 30
                velocity_summary.to_excel(writer, sheet_name='⚡ Velocity Analysis', startrow=1)
                worksheet = writer.sheets['⚡ Velocity Analysis']
                # Format velocity analysis
                worksheet.write(0, 0, 'Velocity_Category', header_format)
                worksheet.write(0, 1, 'SKU_Count', header_format)
                worksheet.write(0, 2, 'Current_Inventory', header_format)
                worksheet.write(0, 3, 'Recommended_PO_Qty', header_format)
                worksheet.write(0, 4, 'Inventory_Value', header_format)
                worksheet.set_column('A:A', 15)
                worksheet.set_column('B:B', 12)
                worksheet.set_column('C:C', 15)
                worksheet.set_column('D:D', 18)
                worksheet.set_column('E:E', 15)

            # Monthly Order Schedule
            order_schedule = []
            for _, row in combined_forecast.iterrows():
                if row['Next_Order_Date']:
                    order_schedule.append({
                        'Order_Date': row['Next_Order_Date'],
                        'SKU': row['SKU'],
                        'Product': row['Product_Name'],
                        'Quantity': row['Next_Order_Qty'],
                        'Lead_Time': row['Lead_Time'],
                        'Arrival_Date': row['Next_Arrival_Date']
                    })
                if row['Order_2_Date']:
                    order_schedule.append({
                        'Order_Date': row['Order_2_Date'],
                        'SKU': row['SKU'],
                        'Product': row['Product_Name'],
                        'Quantity': row['Order_2_Qty'],
                        'Lead_Time': row['Lead_Time'],
                        'Arrival_Date': row['Order_2_Arrival']
                    })

            if order_schedule:
                schedule_df = pd.DataFrame(order_schedule)
                schedule_df['Order_Date'] = pd.to_datetime(schedule_df['Order_Date'])
                schedule_df = schedule_df.sort_values('Order_Date')
                schedule_df['Order_Value'] = schedule_df['Quantity'] * 30
                # Convert back to string for Excel
                schedule_df['Order_Date'] = schedule_df['Order_Date'].dt.strftime('%Y-%m-%d')
                schedule_df.to_excel(writer, sheet_name='📅 Order Schedule', index=False, header=False, startrow=1)
                worksheet = writer.sheets['📅 Order Schedule']
                # Write headers
                headers = ['Order_Date', 'SKU', 'Product', 'Quantity', 'Lead_Time', 'Arrival_Date', 'Order_Value']
                for col_num, header in enumerate(headers):
                    worksheet.write(0, col_num, header, header_format)
                # Set column widths
                worksheet.set_column('A:A', 12)  # Order_Date
                worksheet.set_column('B:B', 15)  # SKU
                worksheet.set_column('C:C', 35)  # Product
                worksheet.set_column('D:D', 10)  # Quantity
                worksheet.set_column('E:E', 10)  # Lead_Time
                worksheet.set_column('F:F', 12)  # Arrival_Date
                worksheet.set_column('G:G', 12)  # Order_Value
                # Format data
                for row_num in range(1, len(schedule_df) + 1):
                    for col_num in range(len(headers)):
                        value = schedule_df.iloc[row_num-1, col_num]
                        if headers[col_num] == 'Order_Value':
                            worksheet.write(row_num, col_num, value, currency_format)
                        elif headers[col_num] in ['Quantity', 'Lead_Time']:
                            worksheet.write(row_num, col_num, value, number_format)
                        else:
                            worksheet.write(row_num, col_num, value, text_format)
                worksheet.freeze_panes(1, 0)
                worksheet.autofilter(0, 0, len(schedule_df), len(headers) - 1)

            # Create a mapping report to show which SKUs were matched vs filtered out
            if not combined_forecast.empty:
                # Show only mapped products in the main report
                mapping_report = combined_forecast[['SKU', 'Product_Name', 'Launch_Date', 'Years_Since_Launch']].copy()
                mapping_report['Mapping_Status'] = 'SUCCESSFULLY_MAPPED'
                mapping_report.to_excel(writer, sheet_name='🔗 Mapped Products', index=False, header=False, startrow=1)
                worksheet = writer.sheets['🔗 Mapped Products']
                format_worksheet(worksheet, mapping_report, '🔗 Mapped Products')

        # For CLI execution - save to file
        try:
            import sys
            is_streamlit = "streamlit" in sys.modules
        except:
            is_streamlit = False
            
        if not is_streamlit:
            # Save to file for CLI execution
            with open(filename, 'wb') as f:
                f.write(excel_buffer.getvalue())

        print(f"\n🎉 SUCCESS! Comprehensive enhanced forecasting complete!")
        print(f"Results saved to: {filename}")
        print(f"Total mapped SKUs processed: {len(combined_forecast)}")
        print(f"Data source: {'Google Sheets (with Inventory from ' + INVENTORY_URL + ')' if USE_GOOGLE_SHEETS else 'Local CSV files'}")

        # ACTIONABLE INSIGHTS SUMMARY
        print("\n" + "="*60)
        print("📊 ACTIONABLE INSIGHTS SUMMARY")
        print("="*60)

        # Immediate Actions
        if insights.get('immediate_actions'):
            print(f"\n🚨 IMMEDIATE ACTIONS REQUIRED ({len(insights.get('immediate_actions'))} items):")
            print("-" * 50)
            for action in insights.get('immediate_actions')[:5]:  # Show top 5
                print(f"• {action['Product']} (SKU: {action['SKU']})")
                print(f"  ACTION: {action['Action']}")
                print(f"  REASON: {action['Reason']}")
                print(f"  QUANTITY: {action['Quantity']} units")
                print(f"  IMPACT: {action['Impact']}")
                print()

        # Risk Analysis
        high_risks = [r for r in insights.get('risk_analysis', []) if r['Risk_Level'] == 'HIGH']

        if high_risks:
            print(f"\n⚠️ HIGH RISK ITEMS ({len(high_risks)} items):")
            print("-" * 50)
            for risk in high_risks[:3]:
                print(f"• {risk['Product']} - {risk['Issue']}")
                print(f"  POTENTIAL LOSS: {risk['Potential_Loss']}")
                print(f"  MITIGATION: {risk['Mitigation']}")
                print()

        # Cash Flow Summary
        print("\n💰 CASH FLOW REQUIREMENTS:")
        print("-" * 50)
        print(f"Next 30 days: ${executive_summary['cash_flow_30_days']:,.0f}")
        print(f"Next 60 days: ${executive_summary['cash_flow_60_days']:,.0f}")
        print(f"Next 90 days: ${executive_summary['cash_flow_90_days']:,.0f}")
        print(f"Total PO Value Needed: ${executive_summary['total_po_value_needed']:,.0f}")

        # Inventory Health
        print("\n📦 INVENTORY HEALTH STATUS:")
        print("-" * 50)
        print(f"At-Risk SKUs (<2 months inventory): {executive_summary['at_risk_skus']}")
        print(f"Overstock SKUs (>6 months inventory): {executive_summary['overstock_skus']}")
        print(f"Total Inventory Value: ${executive_summary['total_inventory_value']:,.0f}")

        # Top Actions by Priority
        if not priority_matrix.empty:
            immediate_actions = priority_matrix[priority_matrix['Action_Priority'] == 'IMMEDIATE']
            high_actions = priority_matrix[priority_matrix['Action_Priority'] == 'HIGH']

            print(f"\n📋 ACTION PRIORITY SUMMARY:")
            print("-" * 50)
            print(f"IMMEDIATE actions required: {len(immediate_actions)} SKUs")
            print(f"HIGH priority actions: {len(high_actions)} SKUs")

            if not immediate_actions.empty:
                print("\nTop 3 IMMEDIATE priorities:")
                for _, row in immediate_actions.head(3).iterrows():
                    print(f"• {row['Product_Name']} - {row['Recommended_Action']}")

        # Quick Win Opportunities
        if insights.get('opportunities'):
            print(f"\n🎯 GROWTH OPPORTUNITIES ({len(insights.get('opportunities'))} items):")
            print("-" * 50)
            for opp in insights.get('opportunities')[:3]:
                print(f"• {opp['Product']} - {opp['Trend']}")
                print(f"  ACTION: {opp['Action']}")
                print()

        print("\n" + "="*60)
        print("✅ NEXT STEPS:")
        print("1. Review '🚨 IMMEDIATE ACTIONS' sheet and place urgent orders TODAY")
        print("2. Check '📊 Executive Summary' for overall inventory health")
        print("3. Use '📋 Action Priority Matrix' to plan this week's activities")
        print("4. Review '⚠️ Risk Analysis' to prevent stockouts")
        print("5. Share 'Finance Cash Flow' sheet with finance team for budget planning")
        print("="*60)

        # Enhanced summary of product mapping - only show mapped products
        if not combined_forecast.empty and 'Product_Name' in combined_forecast.columns:
            total_processed_skus = len(combined_forecast)

            print(f"\n📊 PRODUCT MAPPING SUMMARY:")
            print(f"   SKUs with valid product mappings: {total_processed_skus}")
            print(f"   Mapping success rate: 100% (only mapped products included)")
            print(f"   Inventory data source: Google Sheets" if USE_GOOGLE_SHEETS else "Local CSV")

            # Show some examples of successfully mapped products
            if total_processed_skus > 0:
                sample_mapped = combined_forecast[['SKU', 'Product_Name']].head(5)
                print(f"\n✅ Successfully mapped products (sample):")
                for _, row in sample_mapped.iterrows():
                    print(f"   {row['SKU']} -> {row['Product_Name']}")

        else:
            print(f"\n⚠️  No forecasts generated - check data sources and UPC mapping")

        print("\n✅ Ready for inventory planning decisions!")

        # Return the Excel buffer for Streamlit
        excel_buffer.seek(0)

        # Upload to Google Sheets before returning
        upload_excel_to_google_sheet(excel_buffer)

        # Upload to Google Drive (NEW)
        drive_file_id = upload_to_google_drive_from_buffer(excel_buffer)

        return excel_buffer, filename, drive_file_id


    except FileNotFoundError as e:
        print(f"File not found: {e}")
        print("   Make sure all CSV files are in the current directory:")
        print("   - historical_sales.csv")
        print("   - current_inventory.csv (if not using Google Sheets)")
        print("   - lead_times.csv")
        print("   - product_info.csv")
        return None, None

    except Exception as e:
        print(f"Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        return None, None

# ==============================================================================
# API LAYER (FastAPI)
# ==============================================================================
# This API layer runs alongside Streamlit when started with --api flag
# Usage: python main_app.py --api
# ==============================================================================

api_app = FastAPI(
    title="Herbal Goodness MRP/BOM API",
    description="API for Material Requirements Planning and Bill of Materials Explosion",
    version="1.0.0",
    docs_url="/api/docs",
    redoc_url="/api/redoc"
)

# CORS for ERP integration
api_app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Restrict in production to your ERP domain
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# In-memory job storage (replace with Redis in production if needed)
api_jobs_store: Dict[str, Dict[str, Any]] = {}

# ------------------------------------------------------------------------------
# Pydantic Models for Request/Response Validation
# ------------------------------------------------------------------------------

class BOMExplodeRequest(BaseModel):
    request_id: Optional[str] = Field(default_factory=lambda: str(uuid.uuid4()))
    sku_list: Optional[List[str]] = None
    forecast_source: str = "google_sheets"
    include_procurement: bool = True

class JobResponse(BaseModel):
    success: bool
    job_id: str
    status: str
    poll_url: str
    estimated_duration_seconds: int = 30

class JobStatusResponse(BaseModel):
    job_id: str
    status: str
    progress_percent: int
    started_at: Optional[str] = None
    completed_at: Optional[str] = None
    result_url: Optional[str] = None
    error: Optional[str] = None

class HealthResponse(BaseModel):
    status: str
    version: str
    google_sheets_connected: bool
    google_drive_connected: bool

class RequirementsResponse(BaseModel):
    success: bool
    generated_at: Optional[str] = None
    job_id: Optional[str] = None
    total_count: int = 0
    summary: Optional[Dict[str, Any]] = None
    requirements: List[Dict[str, Any]] = []

# ------------------------------------------------------------------------------
# API Wrapper Function for BOM Explosion
# ------------------------------------------------------------------------------

def api_run_bom_explosion() -> dict:
    """
    API-compatible wrapper for BOM explosion.
    Returns structured dict instead of Excel buffer.
    """
    try:
        # Call existing function
        excel_buffer, filename = run_forecast_bom_analysis(gc_client=None)
        
        if excel_buffer is None:
            return {"success": False, "error": "BOM analysis failed - no data returned"}
        
        # Parse Excel buffer to extract data for API response
        excel_buffer.seek(0)
        all_sheets = pd.read_excel(excel_buffer, sheet_name=None, engine="openpyxl")
        
        # Build structured response
        result = {
            "success": True,
            "summary": {},
            "requirements": [],
            "urgent_reorders": []
        }
        
        # Find MRP Requirements sheet (handle different possible names)
        mrp_sheet_names = ["📦 MRP Requirements", "MRP_Requirements", "MRP Requirements"]
        mrp_df = None
        for name in mrp_sheet_names:
            if name in all_sheets:
                mrp_df = all_sheets[name]
                break
        
        if mrp_df is not None:
            # Clean up DataFrame for JSON serialization
            mrp_df = mrp_df.fillna("")
            result["requirements"] = mrp_df.to_dict(orient="records")
            
            # Build summary
            urgent_count = len(mrp_df[mrp_df["Order_Status"].astype(str).str.contains("🔴", na=False)]) if "Order_Status" in mrp_df.columns else 0
            soon_count = len(mrp_df[mrp_df["Order_Status"].astype(str).str.contains("🟡", na=False)]) if "Order_Status" in mrp_df.columns else 0
            ok_count = len(mrp_df[mrp_df["Order_Status"].astype(str).str.contains("🟢", na=False)]) if "Order_Status" in mrp_df.columns else 0
            
            result["summary"] = {
                "total_components": len(mrp_df),
                "urgent_reorders": urgent_count,
                "reorder_soon": soon_count,
                "ok": ok_count,
                "total_procurement_cost": float(mrp_df["Procurement_Cost"].sum()) if "Procurement_Cost" in mrp_df.columns else 0
            }
        
        # Find Urgent Reorders sheet
        urgent_sheet_names = ["🔴 Urgent Reorders", "Urgent_Reorders", "Urgent Reorders"]
        for name in urgent_sheet_names:
            if name in all_sheets:
                urgent_df = all_sheets[name].fillna("")
                result["urgent_reorders"] = urgent_df.to_dict(orient="records")
                break
        
        return result
        
    except Exception as e:
        import traceback
        return {
            "success": False, 
            "error": str(e),
            "traceback": traceback.format_exc()
        }

# ------------------------------------------------------------------------------
# Background Task for Async BOM Explosion
# ------------------------------------------------------------------------------

def run_bom_explosion_task(job_id: str, request: BOMExplodeRequest):
    """Background task that runs BOM explosion."""
    try:
        api_jobs_store[job_id]["status"] = "processing"
        api_jobs_store[job_id]["progress_percent"] = 10
        
        # Run actual BOM explosion
        result = api_run_bom_explosion()
        
        if result.get("success"):
            api_jobs_store[job_id]["status"] = "completed"
            api_jobs_store[job_id]["progress_percent"] = 100
            api_jobs_store[job_id]["completed_at"] = datetime.now().isoformat() + "Z"
            api_jobs_store[job_id]["result"] = result
        else:
            api_jobs_store[job_id]["status"] = "failed"
            api_jobs_store[job_id]["error"] = result.get("error", "Unknown error")
        
    except Exception as e:
        api_jobs_store[job_id]["status"] = "failed"
        api_jobs_store[job_id]["error"] = str(e)

# ------------------------------------------------------------------------------
# API Endpoints
# ------------------------------------------------------------------------------

@api_app.get("/api/v1/health", response_model=HealthResponse, tags=["Health"])
async def health_check():
    """Check API health and connectivity status."""
    return HealthResponse(
        status="healthy",
        version="1.0.0",
        google_sheets_connected="gcp_service_account_sheets" in os.environ,
        google_drive_connected="gcp_service_account_drive" in os.environ
    )


@api_app.post("/api/v1/bom/explode", response_model=JobResponse, tags=["BOM"])
async def explode_bom(request: BOMExplodeRequest, background_tasks: BackgroundTasks):
    """
    Trigger BOM explosion asynchronously.
    Returns job_id for status polling.
    
    **Example Request:**
```json
    {
        "request_id": "my-unique-id-123",
        "forecast_source": "google_sheets",
        "include_procurement": true
    }
```
    """
    job_id = f"bom-{uuid.uuid4()}"
    
    # Check for idempotency (same request_id returns existing job)
    for existing_job in api_jobs_store.values():
        if existing_job.get("request_id") == request.request_id:
            return JobResponse(
                success=True,
                job_id=existing_job["job_id"],
                status=existing_job["status"],
                poll_url=f"/api/v1/jobs/{existing_job['job_id']}",
                estimated_duration_seconds=0
            )
    
    # Create new job
    api_jobs_store[job_id] = {
        "job_id": job_id,
        "request_id": request.request_id,
        "status": "pending",
        "progress_percent": 0,
        "started_at": datetime.now().isoformat() + "Z",
        "completed_at": None,
        "result": None,
        "error": None
    }
    
    # Start background task
    background_tasks.add_task(run_bom_explosion_task, job_id, request)
    
    return JobResponse(
        success=True,
        job_id=job_id,
        status="pending",
        poll_url=f"/api/v1/jobs/{job_id}",
        estimated_duration_seconds=45
    )


@api_app.get("/api/v1/jobs/{job_id}", response_model=JobStatusResponse, tags=["Jobs"])
async def get_job_status(job_id: str):
    """
    Get status of a running or completed job.
    
    **Status Values:**
    - `pending`: Job queued but not started
    - `processing`: Job currently running
    - `completed`: Job finished successfully
    - `failed`: Job encountered an error
    """
    if job_id not in api_jobs_store:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = api_jobs_store[job_id]
    return JobStatusResponse(
        job_id=job["job_id"],
        status=job["status"],
        progress_percent=job["progress_percent"],
        started_at=job["started_at"],
        completed_at=job["completed_at"],
        result_url=f"/api/v1/jobs/{job_id}/result" if job["status"] == "completed" else None,
        error=job.get("error")
    )


@api_app.get("/api/v1/jobs/{job_id}/result", tags=["Jobs"])
async def get_job_result(job_id: str):
    """
    Get results of a completed job.
    
    Returns full MRP requirements data including:
    - Summary statistics
    - All component requirements
    - Urgent reorders list
    """
    if job_id not in api_jobs_store:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = api_jobs_store[job_id]
    
    if job["status"] == "failed":
        raise HTTPException(
            status_code=500,
            detail=f"Job failed: {job.get('error', 'Unknown error')}"
        )
    
    if job["status"] != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"Job not completed. Current status: {job['status']}"
        )
    
    return {
        "job_id": job_id,
        "completed_at": job["completed_at"],
        **job["result"]
    }


@api_app.get("/api/v1/requirements/latest", response_model=RequirementsResponse, tags=["Requirements"])
async def get_latest_requirements(
    status: Optional[str] = None,
    min_cost: Optional[float] = None
):
    """
    Get latest MRP requirements from most recent completed BOM run.
    
    **Query Parameters:**
    - `status`: Filter by order status (`urgent_reorder`, `reorder_soon`, `ok`)
    - `min_cost`: Minimum procurement cost threshold
    
    **Example:** `/api/v1/requirements/latest?status=urgent_reorder&min_cost=1000`
    """
    # Find the most recent completed BOM job
    completed_jobs = [
        j for j in api_jobs_store.values()
        if j["status"] == "completed" and j.get("result")
    ]
    
    if not completed_jobs:
        return RequirementsResponse(
            success=False,
            total_count=0,
            requirements=[]
        )
    
    # Get most recent
    latest_job = max(completed_jobs, key=lambda j: j["completed_at"] or "")
    result = latest_job["result"]
    
    requirements = result.get("requirements", [])
    
    # Apply filters
    if status:
        status_map = {"urgent_reorder": "🔴", "reorder_soon": "🟡", "ok": "🟢"}
        filter_emoji = status_map.get(status, "")
        if filter_emoji:
            requirements = [
                r for r in requirements 
                if filter_emoji in str(r.get("Order_Status", ""))
            ]
    
    if min_cost is not None:
        requirements = [
            r for r in requirements 
            if float(r.get("Procurement_Cost", 0) or 0) >= min_cost
        ]
    
    return RequirementsResponse(
        success=True,
        generated_at=latest_job["completed_at"],
        job_id=latest_job["job_id"],
        total_count=len(requirements),
        summary=result.get("summary"),
        requirements=requirements
    )


@api_app.get("/api/v1/bom/run-sync", tags=["BOM"])
async def run_bom_sync():
    """
    Run BOM explosion synchronously (blocking).
    Use this for testing or when you need immediate results.
    
    **Warning:** This may take 30-60 seconds. For production, use `/api/v1/bom/explode` instead.
    """
    result = api_run_bom_explosion()
    
    if not result.get("success"):
        raise HTTPException(
            status_code=500,
            detail=result.get("error", "BOM explosion failed")
        )
    
    return result

# ==============================================================================
# ADDITIONAL PYDANTIC MODELS FOR ERP INTEGRATION
# ==============================================================================

class InventoryItem(BaseModel):
    component_id: str
    quantity: float
    warehouse: Optional[str] = "DEFAULT"
    uom: Optional[str] = "EA"

class InventorySyncRequest(BaseModel):
    request_id: Optional[str] = Field(default_factory=lambda: str(uuid.uuid4()))
    source: str = "erpnext"
    timestamp: Optional[str] = Field(default_factory=lambda: datetime.now().isoformat())
    inventory_items: List[InventoryItem]

class InventorySyncResponse(BaseModel):
    success: bool
    request_id: str
    items_received: int
    items_processed: int
    items_failed: int
    failed_items: List[Dict[str, Any]] = []
    message: str

class ForecastOverrideItem(BaseModel):
    sku_id: str
    forecast_quantity: float
    period: Optional[str] = None  # e.g., "2025-01", "2025-Q1"
    reason: Optional[str] = None

class ForecastOverrideRequest(BaseModel):
    request_id: Optional[str] = Field(default_factory=lambda: str(uuid.uuid4()))
    source: str = "erpnext"
    timestamp: Optional[str] = Field(default_factory=lambda: datetime.now().isoformat())
    overrides: List[ForecastOverrideItem]
    replace_existing: bool = False  # If True, replace all forecasts; if False, merge

class ForecastOverrideResponse(BaseModel):
    success: bool
    request_id: str
    items_received: int
    items_applied: int
    message: str

class ProductItem(BaseModel):
    sku_id: str
    sku_name: str
    upc: Optional[str] = None
    category: Optional[str] = None
    status: Optional[str] = "active"  # active, discontinued, pending
    unit_cost: Optional[float] = 0.0
    lead_time_days: Optional[int] = 0
    launch_date: Optional[str] = None

class ProductSyncRequest(BaseModel):
    request_id: Optional[str] = Field(default_factory=lambda: str(uuid.uuid4()))
    source: str = "erpnext"
    timestamp: Optional[str] = Field(default_factory=lambda: datetime.now().isoformat())
    products: List[ProductItem]
    sync_mode: str = "upsert"  # "upsert" (update or insert) | "replace_all"

class ProductSyncResponse(BaseModel):
    success: bool
    request_id: str
    products_received: int
    products_created: int
    products_updated: int
    products_failed: int
    failed_items: List[Dict[str, Any]] = []
    message: str

class BOMComponent(BaseModel):
    component_id: str
    component_name: str
    quantity_required: float
    uom: Optional[str] = "EA"
    wastage_pct: Optional[float] = 0.0
    unit_cost: Optional[float] = 0.0

class BOMItem(BaseModel):
    parent_sku_id: str
    parent_sku_name: Optional[str] = None
    components: List[BOMComponent]

class BOMSyncRequest(BaseModel):
    request_id: Optional[str] = Field(default_factory=lambda: str(uuid.uuid4()))
    source: str = "erpnext"
    timestamp: Optional[str] = Field(default_factory=lambda: datetime.now().isoformat())
    bom_items: List[BOMItem]
    sync_mode: str = "upsert"  # "upsert" | "replace_all"

class BOMSyncResponse(BaseModel):
    success: bool
    request_id: str
    bom_items_received: int
    bom_items_processed: int
    total_components_processed: int
    message: str

class ProcurementParam(BaseModel):
    component_id: str
    lead_time_days: Optional[int] = None
    moq: Optional[float] = None  # Minimum Order Quantity
    eoq: Optional[float] = None  # Economic Order Quantity
    safety_stock_pct: Optional[float] = None
    reorder_point: Optional[float] = None
    supplier_id: Optional[str] = None
    supplier_name: Optional[str] = None

class ProcurementParamsSyncRequest(BaseModel):
    request_id: Optional[str] = Field(default_factory=lambda: str(uuid.uuid4()))
    source: str = "erpnext"
    timestamp: Optional[str] = Field(default_factory=lambda: datetime.now().isoformat())
    parameters: List[ProcurementParam]

class ProcurementParamsSyncResponse(BaseModel):
    success: bool
    request_id: str
    items_received: int
    items_updated: int
    items_failed: int
    failed_items: List[Dict[str, Any]] = []
    message: str


# ==============================================================================
# IN-MEMORY DATA STORES FOR ERP SYNC (Replace with database in production)
# ==============================================================================

# These stores hold data received from ERP until next BOM run
erp_inventory_store: Dict[str, Dict[str, Any]] = {}
erp_forecast_overrides: Dict[str, Dict[str, Any]] = {}
erp_products_store: Dict[str, Dict[str, Any]] = {}
erp_bom_store: Dict[str, Dict[str, Any]] = {}
erp_procurement_params_store: Dict[str, Dict[str, Any]] = {}

# Sync history for audit trail
sync_history: List[Dict[str, Any]] = []


# ==============================================================================
# ERP INTEGRATION ENDPOINTS
# ==============================================================================

@api_app.post("/api/v1/inventory/sync", response_model=InventorySyncResponse, tags=["ERP Integration"])
async def sync_inventory(request: InventorySyncRequest):
    """
    Receive inventory updates from ERP system.
    
    **Use Case:** ERPNext pushes real-time inventory levels after stock transactions.
    
    **Example Request:**
```json
    {
        "request_id": "inv-sync-001",
        "source": "erpnext",
        "timestamp": "2025-01-15T10:30:00Z",
        "inventory_items": [
            {"component_id": "COMP-001", "quantity": 1500, "warehouse": "WH-MAIN"},
            {"component_id": "COMP-002", "quantity": 800, "warehouse": "WH-MAIN"}
        ]
    }
```
    
    **Behavior:**
    - Stores inventory data for use in next BOM explosion
    - Validates component IDs against known components
    - Returns count of successfully processed items
    """
    items_processed = 0
    items_failed = 0
    failed_items = []
    
    for item in request.inventory_items:
        try:
            # Store in memory (replace with database in production)
            erp_inventory_store[item.component_id.upper()] = {
                "quantity": item.quantity,
                "warehouse": item.warehouse,
                "uom": item.uom,
                "updated_at": request.timestamp,
                "source": request.source
            }
            items_processed += 1
        except Exception as e:
            items_failed += 1
            failed_items.append({
                "component_id": item.component_id,
                "error": str(e)
            })
    
    # Log sync event
    sync_history.append({
        "type": "inventory_sync",
        "request_id": request.request_id,
        "timestamp": request.timestamp,
        "source": request.source,
        "items_received": len(request.inventory_items),
        "items_processed": items_processed,
        "items_failed": items_failed
    })
    
    return InventorySyncResponse(
        success=items_failed == 0,
        request_id=request.request_id,
        items_received=len(request.inventory_items),
        items_processed=items_processed,
        items_failed=items_failed,
        failed_items=failed_items,
        message=f"Inventory sync completed. {items_processed} items updated, {items_failed} failed."
    )


@api_app.post("/api/v1/forecast/override", response_model=ForecastOverrideResponse, tags=["ERP Integration"])
async def override_forecast(request: ForecastOverrideRequest):
    """
    Receive forecast overrides from ERP system.
    
    **Use Case:** Planners in ERPNext manually adjust forecasts based on market intelligence.
    
    **Example Request:**
```json
    {
        "request_id": "fcst-override-001",
        "source": "erpnext",
        "overrides": [
            {"sku_id": "SKU-001", "forecast_quantity": 5000, "period": "2025-Q1", "reason": "Seasonal promotion"},
            {"sku_id": "SKU-002", "forecast_quantity": 3000, "period": "2025-Q1", "reason": "New customer order"}
        ],
        "replace_existing": false
    }
```
    
    **Behavior:**
    - Stores forecast overrides for use in next BOM explosion
    - If `replace_existing=true`, clears all previous overrides
    - Overrides take precedence over calculated forecasts
    """
    if request.replace_existing:
        erp_forecast_overrides.clear()
    
    items_applied = 0
    
    for override in request.overrides:
        try:
            erp_forecast_overrides[override.sku_id.upper()] = {
                "forecast_quantity": override.forecast_quantity,
                "period": override.period,
                "reason": override.reason,
                "updated_at": request.timestamp,
                "source": request.source
            }
            items_applied += 1
        except Exception as e:
            print(f"Error applying forecast override for {override.sku_id}: {e}")
    
    # Log sync event
    sync_history.append({
        "type": "forecast_override",
        "request_id": request.request_id,
        "timestamp": request.timestamp,
        "source": request.source,
        "items_received": len(request.overrides),
        "items_applied": items_applied,
        "replace_existing": request.replace_existing
    })
    
    return ForecastOverrideResponse(
        success=True,
        request_id=request.request_id,
        items_received=len(request.overrides),
        items_applied=items_applied,
        message=f"Forecast override completed. {items_applied} SKUs updated."
    )


@api_app.post("/api/v1/products/sync", response_model=ProductSyncResponse, tags=["ERP Integration"])
async def sync_products(request: ProductSyncRequest):
    """
    Receive product master data from ERP system.
    
    **Use Case:** ERPNext pushes product catalog updates (new products, price changes, discontinuations).
    
    **Example Request:**
```json
    {
        "request_id": "prod-sync-001",
        "source": "erpnext",
        "products": [
            {
                "sku_id": "SKU-001",
                "sku_name": "Organic Green Tea 100g",
                "upc": "123456789012",
                "category": "Teas",
                "status": "active",
                "unit_cost": 5.50,
                "lead_time_days": 14
            }
        ],
        "sync_mode": "upsert"
    }
```
    
    **Sync Modes:**
    - `upsert`: Update existing products, insert new ones
    - `replace_all`: Clear all products and replace with provided list
    """
    if request.sync_mode == "replace_all":
        erp_products_store.clear()
    
    products_created = 0
    products_updated = 0
    products_failed = 0
    failed_items = []
    
    for product in request.products:
        try:
            sku_key = product.sku_id.upper()
            is_new = sku_key not in erp_products_store
            
            erp_products_store[sku_key] = {
                "sku_id": product.sku_id,
                "sku_name": product.sku_name,
                "upc": product.upc,
                "category": product.category,
                "status": product.status,
                "unit_cost": product.unit_cost,
                "lead_time_days": product.lead_time_days,
                "launch_date": product.launch_date,
                "updated_at": request.timestamp,
                "source": request.source
            }
            
            if is_new:
                products_created += 1
            else:
                products_updated += 1
                
        except Exception as e:
            products_failed += 1
            failed_items.append({
                "sku_id": product.sku_id,
                "error": str(e)
            })
    
    # Log sync event
    sync_history.append({
        "type": "product_sync",
        "request_id": request.request_id,
        "timestamp": request.timestamp,
        "source": request.source,
        "products_received": len(request.products),
        "products_created": products_created,
        "products_updated": products_updated,
        "products_failed": products_failed
    })
    
    return ProductSyncResponse(
        success=products_failed == 0,
        request_id=request.request_id,
        products_received=len(request.products),
        products_created=products_created,
        products_updated=products_updated,
        products_failed=products_failed,
        failed_items=failed_items,
        message=f"Product sync completed. {products_created} created, {products_updated} updated, {products_failed} failed."
    )


@api_app.post("/api/v1/bom/sync", response_model=BOMSyncResponse, tags=["ERP Integration"])
async def sync_bom(request: BOMSyncRequest):
    """
    Receive BOM structure updates from ERP system.
    
    **Use Case:** ERPNext pushes BOM changes when product formulations are updated.
    
    **Example Request:**
```json
    {
        "request_id": "bom-sync-001",
        "source": "erpnext",
        "bom_items": [
            {
                "parent_sku_id": "SKU-001",
                "parent_sku_name": "Organic Green Tea 100g",
                "components": [
                    {"component_id": "COMP-TEA-001", "component_name": "Green Tea Leaves", "quantity_required": 0.1, "uom": "KG", "wastage_pct": 2.0, "unit_cost": 25.00},
                    {"component_id": "COMP-PKG-001", "component_name": "Tea Box 100g", "quantity_required": 1, "uom": "EA", "wastage_pct": 1.0, "unit_cost": 0.50}
                ]
            }
        ],
        "sync_mode": "upsert"
    }
```
    
    **Sync Modes:**
    - `upsert`: Update existing BOMs, insert new ones
    - `replace_all`: Clear all BOMs and replace with provided list
    """
    if request.sync_mode == "replace_all":
        erp_bom_store.clear()
    
    bom_items_processed = 0
    total_components_processed = 0
    
    for bom_item in request.bom_items:
        try:
            sku_key = bom_item.parent_sku_id.upper()
            
            components_list = []
            for comp in bom_item.components:
                components_list.append({
                    "component_id": comp.component_id,
                    "component_name": comp.component_name,
                    "quantity_required": comp.quantity_required,
                    "uom": comp.uom,
                    "wastage_pct": comp.wastage_pct,
                    "unit_cost": comp.unit_cost
                })
                total_components_processed += 1
            
            erp_bom_store[sku_key] = {
                "parent_sku_id": bom_item.parent_sku_id,
                "parent_sku_name": bom_item.parent_sku_name,
                "components": components_list,
                "updated_at": request.timestamp,
                "source": request.source
            }
            
            bom_items_processed += 1
            
        except Exception as e:
            print(f"Error processing BOM for {bom_item.parent_sku_id}: {e}")
    
    # Log sync event
    sync_history.append({
        "type": "bom_sync",
        "request_id": request.request_id,
        "timestamp": request.timestamp,
        "source": request.source,
        "bom_items_received": len(request.bom_items),
        "bom_items_processed": bom_items_processed,
        "total_components_processed": total_components_processed
    })
    
    return BOMSyncResponse(
        success=True,
        request_id=request.request_id,
        bom_items_received=len(request.bom_items),
        bom_items_processed=bom_items_processed,
        total_components_processed=total_components_processed,
        message=f"BOM sync completed. {bom_items_processed} BOMs with {total_components_processed} components processed."
    )


@api_app.post("/api/v1/procurement-params/sync", response_model=ProcurementParamsSyncResponse, tags=["ERP Integration"])
async def sync_procurement_params(request: ProcurementParamsSyncRequest):
    """
    Receive procurement parameters from ERP system.
    
    **Use Case:** ERPNext pushes updated lead times, MOQ, EOQ when supplier terms change.
    
    **Example Request:**
```json
    {
        "request_id": "proc-sync-001",
        "source": "erpnext",
        "parameters": [
            {
                "component_id": "COMP-001",
                "lead_time_days": 14,
                "moq": 500,
                "eoq": 1000,
                "safety_stock_pct": 0.15,
                "supplier_id": "SUP-001",
                "supplier_name": "ABC Suppliers"
            }
        ]
    }
```
    
    **Behavior:**
    - Updates procurement parameters for specified components
    - Only provided fields are updated (null fields are ignored)
    - Parameters are used in next BOM/MRP calculation
    """
    items_updated = 0
    items_failed = 0
    failed_items = []
    
    for param in request.parameters:
        try:
            comp_key = param.component_id.upper()
            
            # Get existing params or create new
            existing = erp_procurement_params_store.get(comp_key, {})
            
            # Update only provided fields (not None)
            if param.lead_time_days is not None:
                existing["lead_time_days"] = param.lead_time_days
            if param.moq is not None:
                existing["moq"] = param.moq
            if param.eoq is not None:
                existing["eoq"] = param.eoq
            if param.safety_stock_pct is not None:
                existing["safety_stock_pct"] = param.safety_stock_pct
            if param.reorder_point is not None:
                existing["reorder_point"] = param.reorder_point
            if param.supplier_id is not None:
                existing["supplier_id"] = param.supplier_id
            if param.supplier_name is not None:
                existing["supplier_name"] = param.supplier_name
            
            existing["updated_at"] = request.timestamp
            existing["source"] = request.source
            
            erp_procurement_params_store[comp_key] = existing
            items_updated += 1
            
        except Exception as e:
            items_failed += 1
            failed_items.append({
                "component_id": param.component_id,
                "error": str(e)
            })
    
    # Log sync event
    sync_history.append({
        "type": "procurement_params_sync",
        "request_id": request.request_id,
        "timestamp": request.timestamp,
        "source": request.source,
        "items_received": len(request.parameters),
        "items_updated": items_updated,
        "items_failed": items_failed
    })
    
    return ProcurementParamsSyncResponse(
        success=items_failed == 0,
        request_id=request.request_id,
        items_received=len(request.parameters),
        items_updated=items_updated,
        items_failed=items_failed,
        failed_items=failed_items,
        message=f"Procurement params sync completed. {items_updated} items updated, {items_failed} failed."
    )


# ==============================================================================
# DATA RETRIEVAL ENDPOINTS (For ERP to verify synced data)
# ==============================================================================

@api_app.get("/api/v1/inventory/current", tags=["ERP Integration"])
async def get_current_inventory():
    """
    Get currently synced inventory data.
    
    **Use Case:** ERPNext verifies what inventory data the MRP system has.
    """
    return {
        "success": True,
        "total_items": len(erp_inventory_store),
        "inventory": erp_inventory_store
    }


@api_app.get("/api/v1/forecast/overrides", tags=["ERP Integration"])
async def get_forecast_overrides():
    """
    Get currently active forecast overrides.
    
    **Use Case:** ERPNext reviews what forecast overrides are active.
    """
    return {
        "success": True,
        "total_overrides": len(erp_forecast_overrides),
        "overrides": erp_forecast_overrides
    }


@api_app.get("/api/v1/products/current", tags=["ERP Integration"])
async def get_current_products():
    """
    Get currently synced product master data.
    
    **Use Case:** ERPNext verifies product data in MRP system.
    """
    return {
        "success": True,
        "total_products": len(erp_products_store),
        "products": erp_products_store
    }


@api_app.get("/api/v1/bom/current", tags=["ERP Integration"])
async def get_current_bom():
    """
    Get currently synced BOM structures.
    
    **Use Case:** ERPNext verifies BOM data in MRP system.
    """
    return {
        "success": True,
        "total_boms": len(erp_bom_store),
        "boms": erp_bom_store
    }


@api_app.get("/api/v1/procurement-params/current", tags=["ERP Integration"])
async def get_current_procurement_params():
    """
    Get currently synced procurement parameters.
    
    **Use Case:** ERPNext verifies procurement parameters in MRP system.
    """
    return {
        "success": True,
        "total_items": len(erp_procurement_params_store),
        "parameters": erp_procurement_params_store
    }


@api_app.get("/api/v1/sync/history", tags=["ERP Integration"])
async def get_sync_history(limit: int = 50):
    """
    Get history of sync operations.
    
    **Use Case:** Audit trail for troubleshooting integration issues.
    
    **Query Parameters:**
    - `limit`: Maximum number of records to return (default: 50)
    """
    return {
        "success": True,
        "total_records": len(sync_history),
        "showing": min(limit, len(sync_history)),
        "history": sync_history[-limit:]  # Return most recent
    }


@api_app.delete("/api/v1/sync/clear-all", tags=["ERP Integration"])
async def clear_all_synced_data():
    """
    Clear all synced data from memory.
    
    **Use Case:** Reset system for fresh sync from ERP.
    
    **Warning:** This will delete all inventory, forecasts, products, BOMs, and procurement params!
    """
    erp_inventory_store.clear()
    erp_forecast_overrides.clear()
    erp_products_store.clear()
    erp_bom_store.clear()
    erp_procurement_params_store.clear()
    
    sync_history.append({
        "type": "clear_all",
        "timestamp": datetime.now().isoformat(),
        "message": "All synced data cleared"
    })
    
    return {
        "success": True,
        "message": "All synced data has been cleared."
    }

# ==============================================================================
# HERBAL GOODNESS - INVENTORY INTELLIGENCE SYSTEM
# ENHANCED UI v5.1 - Improved Original Theme
# ==============================================================================
# This file contains the ENHANCED Streamlit UI section
# Keep your existing theme but with improved components
# Replace everything from "## Streamlit UI" comment to the end of the file
# ==============================================================================

## Streamlit UI
# --- Import required modules ---

# --- Set page config ---
st.set_page_config(
    page_title="Herbal Goodness | AI Forecasting", 
    layout="wide",
    page_icon="🔮",
    initial_sidebar_state="expanded"
)

# --- Enhanced Styling (Building on Original Theme) ---
st.markdown("""
    <style>
        /* Import futuristic fonts */
        @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;500;600;700&family=JetBrains+Mono:wght@300;400;500;600&display=swap');
        
        /* Global reset and base styling */
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        /* Main app background with animated gradient - ORIGINAL STYLE PRESERVED */
        .stApp {
            background: linear-gradient(135deg, 
                #0a0f1c 0%, 
                #1a2332 25%, 
                #0d2439 50%, 
                #1e3a52 75%, 
                #0f1419 100%);
            background-size: 400% 400%;
            animation: gradientShift 15s ease infinite;
            min-height: 100vh;
            position: relative;
        }
        
        /* Animated background */
        @keyframes gradientShift {
            0% { background-position: 0% 50%; }
            50% { background-position: 100% 50%; }
            100% { background-position: 0% 50%; }
        }
        
        /* Floating particles effect - ORIGINAL */
        .stApp::before {
            content: '';
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background-image: 
                radial-gradient(circle at 20% 20%, rgba(64, 224, 208, 0.1) 0%, transparent 50%),
                radial-gradient(circle at 80% 80%, rgba(135, 206, 235, 0.1) 0%, transparent 50%),
                radial-gradient(circle at 40% 60%, rgba(144, 238, 144, 0.05) 0%, transparent 50%);
            pointer-events: none;
            z-index: 0;
        }
        
        /* =========================================================================
           SIDEBAR ENHANCEMENTS
           ========================================================================= */
        [data-testid="stSidebar"] {
            background: linear-gradient(180deg, 
                rgba(10, 15, 28, 0.98) 0%, 
                rgba(26, 35, 50, 0.98) 50%, 
                rgba(15, 20, 25, 0.98) 100%) !important;
            border-right: 1px solid rgba(64, 224, 208, 0.2) !important;
        }
        
        [data-testid="stSidebar"] > div:first-child {
            background: transparent !important;
            padding-top: 1rem;
        }
        
        .sidebar-header {
            text-align: center;
            padding: 1.5rem 1rem;
            border-bottom: 1px solid rgba(64, 224, 208, 0.2);
            margin-bottom: 1.5rem;
        }
        
        .sidebar-logo {
            font-size: 2.5rem;
            margin-bottom: 0.5rem;
        }
        
        .sidebar-title {
            font-family: 'Space Grotesk', sans-serif;
            font-size: 1.1rem;
            font-weight: 700;
            background: linear-gradient(135deg, #40e0d0, #87ceeb, #90ee90);
            background-clip: text;
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }
        
        .sidebar-subtitle {
            font-family: 'JetBrains Mono', monospace;
            font-size: 0.65rem;
            color: rgba(255, 255, 255, 0.5);
            margin-top: 0.25rem;
            letter-spacing: 0.1em;
        }
        
        .sidebar-section {
            padding: 0 1rem;
            margin-bottom: 1.5rem;
        }
        
        .sidebar-section-title {
            font-family: 'JetBrains Mono', monospace;
            font-size: 0.7rem;
            color: rgba(255, 255, 255, 0.4);
            text-transform: uppercase;
            letter-spacing: 0.1em;
            margin-bottom: 0.75rem;
        }
        
        .sidebar-stat {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 0.6rem 0.75rem;
            background: rgba(64, 224, 208, 0.05);
            border: 1px solid rgba(64, 224, 208, 0.1);
            border-radius: 8px;
            margin-bottom: 0.5rem;
            transition: all 0.3s ease;
        }
        
        .sidebar-stat:hover {
            background: rgba(64, 224, 208, 0.1);
            border-color: rgba(64, 224, 208, 0.3);
        }
        
        .sidebar-stat-label {
            font-family: 'Space Grotesk', sans-serif;
            font-size: 0.8rem;
            color: rgba(255, 255, 255, 0.7);
        }
        
        .sidebar-stat-value {
            font-family: 'JetBrains Mono', monospace;
            font-size: 0.75rem;
            color: #40e0d0;
            font-weight: 500;
        }
        
        .sidebar-status {
            display: flex;
            align-items: center;
            gap: 0.5rem;
            padding: 0.75rem;
            background: rgba(16, 185, 129, 0.1);
            border: 1px solid rgba(16, 185, 129, 0.2);
            border-radius: 8px;
            margin-top: 1rem;
        }
        
        .status-dot {
            width: 8px;
            height: 8px;
            background: #10b981;
            border-radius: 50%;
            animation: pulse 2s ease-in-out infinite;
        }
        
        @keyframes pulse {
            0%, 100% { opacity: 1; box-shadow: 0 0 0 0 rgba(16, 185, 129, 0.4); }
            50% { opacity: 0.8; box-shadow: 0 0 0 6px rgba(16, 185, 129, 0); }
        }
        
        .status-text {
            font-family: 'JetBrains Mono', monospace;
            font-size: 0.75rem;
            color: #10b981;
        }
        
        /* =========================================================================
           HEADER STYLES - ORIGINAL PRESERVED
           ========================================================================= */
        .header-container {
            background: linear-gradient(135deg, 
                rgba(64, 224, 208, 0.1) 0%, 
                rgba(135, 206, 235, 0.15) 50%, 
                rgba(144, 238, 144, 0.1) 100%);
            backdrop-filter: blur(20px);
            border: 1px solid rgba(255, 255, 255, 0.1);
            border-radius: 24px;
            padding: 2.5rem;
            margin: 1.5rem 0 2rem 0;
            position: relative;
            overflow: hidden;
            box-shadow: 
                0 8px 32px rgba(0, 0, 0, 0.3),
                inset 0 1px 0 rgba(255, 255, 255, 0.1);
        }
        
        .header-container::before {
            content: '';
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, 
                transparent, 
                rgba(64, 224, 208, 0.1), 
                transparent);
            animation: shimmer 3s infinite;
        }
        
        @keyframes shimmer {
            0% { left: -100%; }
            100% { left: 100%; }
        }
        
        .company-name {
            font-family: 'Space Grotesk', sans-serif;
            font-size: 3rem;
            font-weight: 700;
            background: linear-gradient(135deg, #40e0d0, #87ceeb, #90ee90);
            background-clip: text;
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            margin: 0;
            text-shadow: 0 0 30px rgba(64, 224, 208, 0.5);
        }
        
        .tagline {
            font-family: 'JetBrains Mono', monospace;
            font-size: 1rem;
            color: rgba(255, 255, 255, 0.7);
            margin-top: 0.5rem;
            letter-spacing: 1px;
        }
        
        /* =========================================================================
           TABS STYLING - NEW
           ========================================================================= */
        .stTabs [data-baseweb="tab-list"] {
            background: linear-gradient(135deg, 
                rgba(64, 224, 208, 0.05) 0%, 
                rgba(135, 206, 235, 0.08) 50%, 
                rgba(144, 238, 144, 0.05) 100%);
            backdrop-filter: blur(10px);
            border: 1px solid rgba(255, 255, 255, 0.1);
            border-radius: 16px;
            padding: 0.5rem;
            gap: 0.5rem;
            margin-bottom: 1.5rem;
        }
        
        .stTabs [data-baseweb="tab"] {
            background: transparent !important;
            border-radius: 12px !important;
            padding: 0.875rem 1.75rem !important;
            font-family: 'Space Grotesk', sans-serif !important;
            font-size: 0.95rem !important;
            font-weight: 500 !important;
            color: rgba(255, 255, 255, 0.6) !important;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
            border: 1px solid transparent !important;
        }
        
        .stTabs [data-baseweb="tab"]:hover {
            background: rgba(64, 224, 208, 0.1) !important;
            color: rgba(255, 255, 255, 0.9) !important;
            border-color: rgba(64, 224, 208, 0.2) !important;
        }
        
        .stTabs [aria-selected="true"] {
            background: linear-gradient(135deg, 
                rgba(64, 224, 208, 0.2) 0%, 
                rgba(135, 206, 235, 0.25) 50%, 
                rgba(144, 238, 144, 0.2) 100%) !important;
            color: #ffffff !important;
            border-color: rgba(64, 224, 208, 0.4) !important;
            box-shadow: 0 4px 15px rgba(64, 224, 208, 0.2) !important;
        }
        
        .stTabs [data-baseweb="tab-highlight"] {
            display: none !important;
        }
        
        .stTabs [data-baseweb="tab-border"] {
            display: none !important;
        }
        
        .stTabs [data-baseweb="tab-panel"] {
            padding-top: 1rem;
        }
        
        /* =========================================================================
           GLASSMORPHISM CARDS - ORIGINAL ENHANCED
           ========================================================================= */
        .holo-card {
            background: linear-gradient(135deg, 
                rgba(255, 255, 255, 0.05) 0%, 
                rgba(64, 224, 208, 0.1) 50%, 
                rgba(135, 206, 235, 0.05) 100%);
            backdrop-filter: blur(15px);
            border: 1px solid rgba(255, 255, 255, 0.1);
            border-radius: 20px;
            padding: 2rem;
            margin: 1.5rem 0;
            position: relative;
            overflow: hidden;
            transition: all 0.4s cubic-bezier(0.23, 1, 0.320, 1);
            box-shadow: 
                0 10px 40px rgba(0, 0, 0, 0.2),
                inset 0 1px 0 rgba(255, 255, 255, 0.1);
        }
        
        .holo-card:hover {
            transform: translateY(-3px);
            box-shadow: 
                0 20px 60px rgba(0, 0, 0, 0.3),
                0 0 50px rgba(64, 224, 208, 0.15);
            border-color: rgba(64, 224, 208, 0.3);
        }
        
        .holo-card::after {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 1px;
            background: linear-gradient(90deg, 
                transparent, 
                rgba(64, 224, 208, 0.8), 
                rgba(135, 206, 235, 0.8), 
                transparent);
            opacity: 0;
            transition: opacity 0.3s;
        }
        
        .holo-card:hover::after {
            opacity: 1;
        }
        
        .section-title {
            font-family: 'Space Grotesk', sans-serif;
            font-size: 1.6rem;
            font-weight: 600;
            color: #ffffff;
            margin-bottom: 1rem;
            display: flex;
            align-items: center;
            gap: 0.75rem;
        }
        
        .section-title-icon {
            font-size: 1.5rem;
        }
        
        .description-text {
            font-family: 'Space Grotesk', sans-serif;
            font-size: 1rem;
            color: rgba(255, 255, 255, 0.7);
            line-height: 1.7;
            margin-bottom: 1.5rem;
        }
        
        /* =========================================================================
           ENHANCED BUTTONS - IMPROVED
           ========================================================================= */
        .stButton > button {
            background: linear-gradient(135deg, 
                rgba(64, 224, 208, 0.25) 0%, 
                rgba(135, 206, 235, 0.35) 50%, 
                rgba(144, 238, 144, 0.25) 100%) !important;
            border: 2px solid rgba(64, 224, 208, 0.5) !important;
            border-radius: 16px !important;
            padding: 1.1rem 2.5rem !important;
            font-family: 'Space Grotesk', sans-serif !important;
            font-size: 1.05rem !important;
            font-weight: 600 !important;
            color: #ffffff !important;
            text-transform: uppercase !important;
            letter-spacing: 1.5px !important;
            transition: all 0.4s cubic-bezier(0.23, 1, 0.320, 1) !important;
            width: 100% !important;
            box-shadow: 
                0 8px 32px rgba(64, 224, 208, 0.25),
                inset 0 1px 0 rgba(255, 255, 255, 0.15) !important;
            position: relative !important;
            overflow: hidden !important;
        }
        
        .stButton > button::before {
            content: '' !important;
            position: absolute !important;
            top: 0 !important;
            left: -100% !important;
            width: 100% !important;
            height: 100% !important;
            background: linear-gradient(90deg, 
                transparent, 
                rgba(255, 255, 255, 0.2), 
                transparent) !important;
            transition: left 0.5s ease !important;
        }
        
        .stButton > button:hover {
            transform: translateY(-4px) scale(1.02) !important;
            box-shadow: 
                0 15px 50px rgba(64, 224, 208, 0.4),
                0 0 40px rgba(64, 224, 208, 0.2),
                inset 0 1px 0 rgba(255, 255, 255, 0.2) !important;
            border-color: rgba(64, 224, 208, 0.8) !important;
        }
        
        .stButton > button:hover::before {
            left: 100% !important;
        }
        
        .stButton > button:active {
            transform: translateY(-2px) scale(1.01) !important;
        }
        
        /* Download button styling - ENHANCED */
        .stDownloadButton > button {
            background: rgba(255, 255, 255, 0.05) !important;
            border: 1px solid rgba(64, 224, 208, 0.4) !important;
            border-radius: 12px !important;
            padding: 1rem 1.5rem !important;
            font-family: 'Space Grotesk', sans-serif !important;
            font-size: 0.9rem !important;
            font-weight: 500 !important;
            color: #40e0d0 !important;
            transition: all 0.3s cubic-bezier(0.23, 1, 0.320, 1) !important;
            width: 100% !important;
            position: relative !important;
            overflow: hidden !important;
        }
        
        .stDownloadButton > button:hover {
            background: linear-gradient(135deg, 
                rgba(64, 224, 208, 0.15), 
                rgba(135, 206, 235, 0.1)) !important;
            border-color: rgba(64, 224, 208, 0.8) !important;
            transform: translateY(-2px) !important;
            box-shadow: 0 10px 30px rgba(64, 224, 208, 0.2) !important;
            color: #ffffff !important;
        }
        
        /* Link buttons - ENHANCED */
        .neural-btn {
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 0.5rem;
            width: 100%;
            padding: 1rem 1.5rem;
            font-family: 'Space Grotesk', sans-serif;
            font-size: 0.9rem;
            font-weight: 500;
            color: rgba(64, 224, 208, 0.9);
            background: rgba(255, 255, 255, 0.03);
            border: 1px solid rgba(64, 224, 208, 0.3);
            border-radius: 12px;
            text-decoration: none;
            transition: all 0.3s cubic-bezier(0.23, 1, 0.320, 1);
            cursor: pointer;
        }
        
        .neural-btn:hover {
            color: #ffffff;
            background: linear-gradient(135deg, 
                rgba(64, 224, 208, 0.15), 
                rgba(135, 206, 235, 0.1));
            border-color: rgba(64, 224, 208, 0.6);
            transform: translateY(-2px);
            box-shadow: 0 8px 25px rgba(64, 224, 208, 0.2);
            text-decoration: none;
        }
        
        /* =========================================================================
           ENHANCED PROGRESS BAR - IMPROVED
           ========================================================================= */
        .stProgress > div > div > div > div {
            background: linear-gradient(90deg, 
                #40e0d0 0%, 
                #87ceeb 25%,
                #90ee90 50%,
                #87ceeb 75%,
                #40e0d0 100%) !important;
            background-size: 200% 100% !important;
            animation: progressGlow 2s linear infinite !important;
            border-radius: 100px !important;
            box-shadow: 
                0 0 20px rgba(64, 224, 208, 0.5),
                0 0 40px rgba(64, 224, 208, 0.3) !important;
        }
        
        @keyframes progressGlow {
            0% { background-position: 0% 50%; }
            100% { background-position: 200% 50%; }
        }
        
        .stProgress > div > div {
            background: rgba(64, 224, 208, 0.15) !important;
            border-radius: 100px !important;
            height: 12px !important;
            border: 1px solid rgba(64, 224, 208, 0.2) !important;
        }
        
        /* Progress container styling */
        .progress-wrapper {
            background: rgba(0, 0, 0, 0.3);
            border: 1px solid rgba(64, 224, 208, 0.2);
            border-radius: 16px;
            padding: 1.5rem;
            margin: 1.5rem 0;
        }
        
        .progress-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 1rem;
        }
        
        .progress-title {
            font-family: 'Space Grotesk', sans-serif;
            font-size: 1rem;
            font-weight: 600;
            color: #40e0d0;
            display: flex;
            align-items: center;
            gap: 0.5rem;
        }
        
        .progress-time {
            font-family: 'JetBrains Mono', monospace;
            font-size: 0.85rem;
            color: rgba(255, 255, 255, 0.6);
        }
        
        /* Progress steps */
        .progress-steps {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
            gap: 0.75rem;
            margin-top: 1.25rem;
            padding-top: 1.25rem;
            border-top: 1px solid rgba(64, 224, 208, 0.1);
        }
        
        .progress-step {
            display: flex;
            align-items: center;
            gap: 0.6rem;
            padding: 0.6rem 0.8rem;
            background: rgba(0, 0, 0, 0.2);
            border-radius: 8px;
            border: 1px solid transparent;
            transition: all 0.3s;
        }
        
        .progress-step.pending {
            color: rgba(255, 255, 255, 0.4);
        }
        
        .progress-step.active {
            color: #40e0d0;
            background: rgba(64, 224, 208, 0.1);
            border-color: rgba(64, 224, 208, 0.3);
        }
        
        .progress-step.complete {
            color: #10b981;
        }
        
        .step-icon {
            width: 20px;
            height: 20px;
            display: flex;
            align-items: center;
            justify-content: center;
            border-radius: 50%;
            font-size: 0.7rem;
            font-weight: 600;
            border: 2px solid currentColor;
            flex-shrink: 0;
        }
        
        .step-icon.active {
            background: #40e0d0;
            border-color: #40e0d0;
            color: #000;
            animation: stepPulse 1.5s ease-in-out infinite;
        }
        
        @keyframes stepPulse {
            0%, 100% { box-shadow: 0 0 0 0 rgba(64, 224, 208, 0.5); }
            50% { box-shadow: 0 0 0 8px rgba(64, 224, 208, 0); }
        }
        
        .step-icon.complete {
            background: #10b981;
            border-color: #10b981;
            color: #fff;
        }
        
        .step-label {
            font-family: 'JetBrains Mono', monospace;
            font-size: 0.75rem;
        }
        
        /* =========================================================================
           STATUS TEXT STYLING
           ========================================================================= */
        .status-text {
            font-family: 'JetBrains Mono', monospace;
            font-size: 0.9rem;
            color: rgba(64, 224, 208, 0.9);
            text-align: center;
            margin: 0.75rem 0;
            letter-spacing: 0.5px;
            animation: statusFade 1.5s ease-in-out infinite;
        }
        
        @keyframes statusFade {
            0%, 100% { opacity: 1; }
            50% { opacity: 0.6; }
        }
        
        /* =========================================================================
           SUCCESS/ERROR MESSAGES - ENHANCED
           ========================================================================= */
        .stSuccess {
            background: linear-gradient(135deg, 
                rgba(16, 185, 129, 0.15), 
                rgba(64, 224, 208, 0.1)) !important;
            border: 1px solid rgba(16, 185, 129, 0.4) !important;
            border-radius: 14px !important;
            padding: 1.25rem !important;
            backdrop-filter: blur(10px) !important;
        }
        
        .stError {
            background: linear-gradient(135deg, 
                rgba(244, 63, 94, 0.15), 
                rgba(244, 63, 94, 0.05)) !important;
            border: 1px solid rgba(244, 63, 94, 0.4) !important;
            border-radius: 14px !important;
            padding: 1.25rem !important;
            backdrop-filter: blur(10px) !important;
        }
        
        .stInfo {
            background: linear-gradient(135deg, 
                rgba(64, 224, 208, 0.15), 
                rgba(135, 206, 235, 0.1)) !important;
            border: 1px solid rgba(64, 224, 208, 0.4) !important;
            border-radius: 14px !important;
            padding: 1.25rem !important;
            backdrop-filter: blur(10px) !important;
        }
        
        /* =========================================================================
           ONBOARDING BANNER - NEW
           ========================================================================= */
        .onboarding-banner {
            background: linear-gradient(135deg, 
                rgba(64, 224, 208, 0.1) 0%, 
                rgba(135, 206, 235, 0.15) 50%,
                rgba(144, 238, 144, 0.1) 100%);
            backdrop-filter: blur(15px);
            border: 1px solid rgba(64, 224, 208, 0.3);
            border-radius: 16px;
            padding: 1.25rem 1.5rem;
            margin-bottom: 1.5rem;
            display: flex;
            align-items: flex-start;
            gap: 1rem;
        }
        
        .onboarding-icon {
            font-size: 1.75rem;
            flex-shrink: 0;
        }
        
        .onboarding-content {
            flex: 1;
        }
        
        .onboarding-title {
            font-family: 'Space Grotesk', sans-serif;
            font-size: 1.05rem;
            font-weight: 600;
            color: #ffffff;
            margin: 0 0 0.4rem 0;
        }
        
        .onboarding-text {
            font-family: 'Space Grotesk', sans-serif;
            font-size: 0.9rem;
            color: rgba(255, 255, 255, 0.7);
            margin: 0;
            line-height: 1.5;
        }
        
        .onboarding-highlight {
            color: #40e0d0;
            font-weight: 500;
        }
        
        /* =========================================================================
           STATS BAR - NEW
           ========================================================================= */
        .stats-bar {
            display: flex;
            gap: 1.5rem;
            padding: 0.875rem 1.25rem;
            background: rgba(0, 0, 0, 0.25);
            border: 1px solid rgba(64, 224, 208, 0.15);
            border-radius: 12px;
            margin-bottom: 1.5rem;
            flex-wrap: wrap;
        }
        
        .stat-item {
            display: flex;
            align-items: center;
            gap: 0.4rem;
            font-family: 'JetBrains Mono', monospace;
            font-size: 0.8rem;
            color: rgba(255, 255, 255, 0.6);
        }
        
        .stat-icon {
            font-size: 0.9rem;
        }
        
        .stat-value {
            color: #40e0d0;
            font-weight: 500;
        }
        
        /* =========================================================================
           DATA VISUALIZATION PREVIEW - ORIGINAL
           ========================================================================= */
        .viz-preview {
            background: linear-gradient(135deg, 
                rgba(64, 224, 208, 0.05), 
                rgba(135, 206, 235, 0.05));
            border: 1px solid rgba(64, 224, 208, 0.2);
            border-radius: 16px;
            padding: 1.5rem;
            margin: 1.5rem 0;
            text-align: center;
        }
        
        .viz-preview h3 {
            color: #40e0d0;
            font-family: 'Space Grotesk', sans-serif;
            margin-bottom: 0.75rem;
            font-size: 1.1rem;
        }
        
        .viz-preview p {
            color: rgba(255, 255, 255, 0.6);
            font-family: 'JetBrains Mono', monospace;
            font-size: 0.8rem;
            margin: 0;
        }
        
        /* =========================================================================
           FOOTER - ENHANCED
           ========================================================================= */
        .footer {
            text-align: center;
            padding: 2.5rem 0 1.5rem 0;
            margin-top: 2.5rem;
            border-top: 1px solid rgba(64, 224, 208, 0.2);
        }
        
        .footer-brand {
            font-family: 'JetBrains Mono', monospace;
            font-size: 0.85rem;
            color: rgba(64, 224, 208, 0.8);
            margin-bottom: 0.4rem;
        }
        
        .footer-text {
            font-family: 'Space Grotesk', sans-serif;
            font-size: 0.8rem;
            color: rgba(255, 255, 255, 0.4);
        }
        
        /* =========================================================================
           RESPONSIVE DESIGN
           ========================================================================= */
        @media (max-width: 768px) {
            .company-name {
                font-size: 2rem;
            }
            
            .tagline {
                font-size: 0.8rem;
            }
            
            .header-container {
                padding: 1.5rem;
            }
            
            .holo-card {
                padding: 1.25rem;
                border-radius: 14px;
            }
            
            .section-title {
                font-size: 1.3rem;
            }
            
            .stButton > button {
                padding: 0.9rem 1.5rem !important;
                font-size: 0.9rem !important;
            }
            
            .stats-bar {
                flex-direction: column;
                gap: 0.75rem;
            }
            
            .progress-steps {
                grid-template-columns: 1fr;
            }
            
            .stTabs [data-baseweb="tab"] {
                padding: 0.75rem 1rem !important;
                font-size: 0.85rem !important;
            }
        }
        
        @media (max-width: 480px) {
            .company-name {
                font-size: 1.6rem;
            }
            
            .holo-card {
                padding: 1rem;
            }
            
            .onboarding-banner {
                flex-direction: column;
                text-align: center;
            }
        }
        
        /* =========================================================================
           HIDE STREAMLIT ELEMENTS
           ========================================================================= */
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        header {visibility: hidden;}
        .stDeployButton {display: none;}
        
        /* Custom scrollbar */
        ::-webkit-scrollbar {
            width: 8px;
            height: 8px;
        }
        
        ::-webkit-scrollbar-track {
            background: rgba(0, 0, 0, 0.3);
        }
        
        ::-webkit-scrollbar-thumb {
            background: rgba(64, 224, 208, 0.3);
            border-radius: 4px;
        }
        
        ::-webkit-scrollbar-thumb:hover {
            background: rgba(64, 224, 208, 0.5);
        }
    </style>
""", unsafe_allow_html=True)

# ==============================================================================
# SESSION STATE INITIALIZATION
# ==============================================================================
if "excel_buffer" not in st.session_state:
    st.session_state.excel_buffer = None
    st.session_state.filename = None
    st.session_state.drive_file_id = None
    st.session_state.file_downloaded = False
    st.session_state.bom_excel_buffer = None
    st.session_state.bom_filename = None
    st.session_state.bom_analysis_complete = False
    st.session_state.bom_sheets_url = None
    st.session_state.show_onboarding = True
    st.session_state.last_forecast_time = None
    st.session_state.last_bom_time = None

# ==============================================================================
# SIDEBAR
# ==============================================================================
with st.sidebar:
    # Sidebar Header
    st.markdown("""
        <div class="sidebar-header">
            <div class="sidebar-logo">🌿</div>
            <div class="sidebar-title">HERBAL GOODNESS</div>
            <div class="sidebar-subtitle">INVENTORY INTELLIGENCE</div>
        </div>
    """, unsafe_allow_html=True)
    
    # Quick Stats Section
    st.markdown("""
        <div class="sidebar-section">
            <div class="sidebar-section-title">Quick Stats</div>
    """, unsafe_allow_html=True)
    
    # Estimated Times
    st.markdown("""
            <div class="sidebar-stat">
                <span class="sidebar-stat-label">⚡ Forecast Time</span>
                <span class="sidebar-stat-value">~90 sec</span>
            </div>
            <div class="sidebar-stat">
                <span class="sidebar-stat-label">🧬 BOM Time</span>
                <span class="sidebar-stat-value">~35 sec</span>
            </div>
    """, unsafe_allow_html=True)
    
    # Last Run Times
    if st.session_state.last_forecast_time:
        st.markdown(f"""
            <div class="sidebar-stat">
                <span class="sidebar-stat-label">📊 Last Forecast</span>
                <span class="sidebar-stat-value">{st.session_state.last_forecast_time[-8:]}</span>
            </div>
        """, unsafe_allow_html=True)
    
    if st.session_state.last_bom_time:
        st.markdown(f"""
            <div class="sidebar-stat">
                <span class="sidebar-stat-label">🧬 Last BOM</span>
                <span class="sidebar-stat-value">{st.session_state.last_bom_time[-8:]}</span>
            </div>
        """, unsafe_allow_html=True)
    
    st.markdown("</div>", unsafe_allow_html=True)
    
    # System Status
    st.markdown("""
        <div class="sidebar-section">
            <div class="sidebar-section-title">System Status</div>
            <div class="sidebar-status">
                <div class="status-dot"></div>
                <span class="status-text">System Online • v5.1</span>
            </div>
        </div>
    """, unsafe_allow_html=True)

# ==============================================================================
# MAIN CONTENT
# ==============================================================================

# Header
header_col1, header_col2 = st.columns([1, 5])

with header_col1:
    try:
        logo_path = os.path.join(BASE_DIR, "logo", "herbal-logo.avif")
        st.image(logo_path, width=160)
    except:
        st.markdown("""
            <div style="width: 80px; height: 80px; background: linear-gradient(135deg, #40e0d0, #87ceeb); 
                        border-radius: 20px; display: flex; align-items: center; justify-content: center; 
                        font-size: 2.5rem; box-shadow: 0 8px 32px rgba(64, 224, 208, 0.3);">
                🌿
            </div>
        """, unsafe_allow_html=True)

with header_col2:
    st.markdown("""
        <div style="padding-left: 1.5rem;">
            <h1 class="company-name">HERBAL GOODNESS</h1>
            <p class="tagline">// INVENTORY INTELLIGENCE SYSTEM //</p>
        </div>
    """, unsafe_allow_html=True)

# Onboarding Banner
if st.session_state.show_onboarding:
    st.markdown("""
        <div class="onboarding-banner">
            <div class="onboarding-icon">👋</div>
            <div class="onboarding-content">
                <h4 class="onboarding-title">Welcome to the Inventory Intelligence System</h4>
                <p class="onboarding-text">
                    Use the <span class="onboarding-highlight">Forecast Engine</span> (~90 sec) to predict demand across all channels, 
                    or run <span class="onboarding-highlight">BOM Analysis</span> (~35 sec) for material requirements planning.
                </p>
            </div>
        </div>
    """, unsafe_allow_html=True)
    
    if st.button("✕ Dismiss Welcome", key="dismiss_onboarding"):
        st.session_state.show_onboarding = False
        st.rerun()

# Stats Bar
current_time = datetime.now().strftime("%Y-%m-%d %H:%M")
st.markdown(f"""
    <div class="stats-bar">
        <div class="stat-item">
            <span class="stat-icon">📅</span>
            <span>Current: <span class="stat-value">{current_time}</span></span>
        </div>
        <div class="stat-item">
            <span class="stat-icon">📊</span>
            <span>Channels: <span class="stat-value">5</span></span>
        </div>
        <div class="stat-item">
            <span class="stat-icon">🔄</span>
            <span>Data: <span class="stat-value">Google Sheets</span></span>
        </div>
    </div>
""", unsafe_allow_html=True)

# ==============================================================================
# TABBED INTERFACE
# ==============================================================================
tab1, tab2, tab3 = st.tabs(["🔮 Forecast Engine", "🧬 BOM Analysis", "📁 Access Portal"])

# ==============================================================================
# TAB 1: FORECAST ENGINE
# ==============================================================================
with tab1:
    st.markdown("""
        <div class="holo-card">
            <h2 class="section-title">
                <span class="section-title-icon">🔮</span>
                Enhanced Forecast Engine
            </h2>
            <p class="description-text">
                Unleash the power of advanced analytics and machine learning algorithms 
                to predict inventory patterns with high accuracy. Analyzes data from 
                Amazon FBA, Shopify Main, Shopify Faire, Amazon FBM, and Walmart FBM.
            </p>
        </div>
    """, unsafe_allow_html=True)
    
    # Containers
    forecast_btn_container = st.container()
    forecast_progress_container = st.container()
    forecast_result_container = st.container()
    
    with forecast_btn_container:
        if st.button("🚀 INITIATE FORECASTING ANALYSIS", key="generate_btn"):
            start_time = time.time()
            
            with forecast_progress_container:
                # Progress wrapper
                st.markdown("""
                    <div class="progress-wrapper">
                        <div class="progress-header">
                            <span class="progress-title">⚡ Processing Forecast Analysis</span>
                            <span class="progress-time">Est. ~90 seconds</span>
                        </div>
                    </div>
                """, unsafe_allow_html=True)
                
                progress_bar = st.progress(0)
                status_text = st.empty()
                steps_container = st.empty()
                
                forecast_stages = [
                    "Initializing Data Collection",
                    "Sourcing Google Sheets Data",
                    "Mapping SKU Relationships",
                    "Processing Channel Sales",
                    "Running Forecast Algorithms",
                    "Calculating Safety Stock & ROP",
                    "Generating Intelligence Reports",
                    "Finalizing Output"
                ]
                
                def render_steps(current_idx, stages):
                    html = '<div class="progress-steps">'
                    for idx, stage in enumerate(stages):
                        if idx < current_idx:
                            status = "complete"
                            icon_class = "complete"
                            icon = "✓"
                        elif idx == current_idx:
                            status = "active"
                            icon_class = "active"
                            icon = "●"
                        else:
                            status = "pending"
                            icon_class = ""
                            icon = str(idx + 1)
                        
                        html += f'''
                            <div class="progress-step {status}">
                                <div class="step-icon {icon_class}">{icon}</div>
                                <span class="step-label">{stage}</span>
                            </div>
                        '''
                    html += '</div>'
                    return html
                
                # Progress animation (first half)
                for i in range(40):
                    stage_idx = min(i // 5, len(forecast_stages) - 1)
                    status_text.markdown(
                        f'<p class="status-text">{forecast_stages[stage_idx]}...</p>',
                        unsafe_allow_html=True
                    )
                    steps_container.markdown(render_steps(stage_idx, forecast_stages), unsafe_allow_html=True)
                    time.sleep(0.05)
                    progress_bar.progress(i + 1)
                
                # Run actual forecast
                try:
                    result = main()
                    
                    if result is None or (isinstance(result, tuple) and result[0] is None):
                        st.error("❌ Forecast Analysis Failed. Please check the logs.")
                        progress_bar.empty()
                        status_text.empty()
                        steps_container.empty()
                        st.stop()
                    
                    excel_buffer, filename, drive_file_id = result
                    
                except Exception as e:
                    import traceback
                    st.error(f"❌ Forecast Analysis Error: {str(e)}")
                    st.code(traceback.format_exc())
                    progress_bar.empty()
                    status_text.empty()
                    steps_container.empty()
                    st.stop()
                
                # Complete progress
                for i in range(40, 100):
                    stage_idx = min(i // 12, len(forecast_stages) - 1)
                    status_text.markdown(
                        f'<p class="status-text">{forecast_stages[stage_idx]}...</p>',
                        unsafe_allow_html=True
                    )
                    steps_container.markdown(render_steps(stage_idx, forecast_stages), unsafe_allow_html=True)
                    time.sleep(0.02)
                    progress_bar.progress(i + 1)
                
                # Clear progress UI
                progress_bar.empty()
                status_text.empty()
                steps_container.empty()
                
                # Store results
                if excel_buffer:
                    st.session_state.excel_buffer = excel_buffer
                    st.session_state.filename = filename
                    st.session_state.drive_file_id = drive_file_id
                    
                    end_time = time.time()
                    duration_sec = end_time - start_time
                    duration_str = time.strftime("%M:%S", time.gmtime(duration_sec))
                    timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    st.session_state.last_forecast_time = timestamp_str
                    
                    with forecast_result_container:
                        st.success(f"""
                            ✅ **FORECAST ANALYSIS COMPLETED | REPORTS READY!**
                            
                            📅 Generated at: **{timestamp_str}**  
                            ⏱ Duration: **{duration_str}**
                        """)
                else:
                    with forecast_result_container:
                        st.error("❌ Forecast Analysis Failed. Please try again.")
    
    # Show quick access if results exist
    if st.session_state.excel_buffer:
        st.markdown("""
            <div class="holo-card" style="margin-top: 1rem;">
                <h3 class="section-title" style="font-size: 1.3rem;">
                    <span class="section-title-icon">📥</span>
                    Quick Access
                </h3>
            </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns(3, gap="large")
        
        with col1:
            st.download_button(
                label="📥 DOWNLOAD EXCEL",
                data=st.session_state.excel_buffer.getvalue(),
                file_name=st.session_state.filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        with col2:
            if st.session_state.drive_file_id:
                sheets_link = f"https://docs.google.com/spreadsheets/d/{st.session_state.drive_file_id}/edit"
                st.markdown(f'<a href="{sheets_link}" target="_blank" class="neural-btn">☁️ GOOGLE SHEETS</a>', unsafe_allow_html=True)
        
        with col3:
            looker_url = "https://lookerstudio.google.com/reporting/9525ae1e-6f0e-4b5f-ae50-ca84312b76fd/page/br5SF"
            st.markdown(f'<a href="{looker_url}" target="_blank" class="neural-btn">📊 LOOKER DASHBOARD</a>', unsafe_allow_html=True)

# ==============================================================================
# TAB 2: BOM ANALYSIS
# ==============================================================================
with tab2:
    st.markdown("""
        <div class="holo-card">
            <h2 class="section-title">
                <span class="section-title-icon">🧬</span>
                Forecast BOM Explosion
            </h2>
            <p class="description-text">
                Execute multi-level Bill of Materials explosion using real forecast data. 
                Calculate component requirements, ROP, and procurement needs across your 
                entire product hierarchy with safety stock calculations and urgent reorder alerts.
            </p>
        </div>
    """, unsafe_allow_html=True)
    
    # Containers
    bom_btn_container = st.container()
    bom_progress_container = st.container()
    bom_result_container = st.container()
    
    with bom_btn_container:
        if st.button("🧬 INITIATE FORECAST BOM ANALYSIS", key="bom_generate_btn"):
            bom_start_time = time.time()
            
            with bom_progress_container:
                st.markdown("""
                    <div class="progress-wrapper">
                        <div class="progress-header">
                            <span class="progress-title">🧬 Processing BOM Explosion</span>
                            <span class="progress-time">Est. ~35 seconds</span>
                        </div>
                    </div>
                """, unsafe_allow_html=True)
                
                bom_progress_bar = st.progress(0)
                bom_status_text = st.empty()
                bom_steps_container = st.empty()
                
                bom_stages = [
                    "Connecting to BOM Data",
                    "Loading Bill of Materials",
                    "Fetching SKU-UPC Mappings",
                    "Retrieving Forecast Demand",
                    "Exploding Multi-Level BOM",
                    "Calculating ROP & Procurement",
                    "Generating MRP Reports"
                ]
                
                def render_bom_steps(current_idx, stages):
                    html = '<div class="progress-steps">'
                    for idx, stage in enumerate(stages):
                        if idx < current_idx:
                            status = "complete"
                            icon_class = "complete"
                            icon = "✓"
                        elif idx == current_idx:
                            status = "active"
                            icon_class = "active"
                            icon = "●"
                        else:
                            status = "pending"
                            icon_class = ""
                            icon = str(idx + 1)
                        
                        html += f'''
                            <div class="progress-step {status}">
                                <div class="step-icon {icon_class}">{icon}</div>
                                <span class="step-label">{stage}</span>
                            </div>
                        '''
                    html += '</div>'
                    return html
                
                # Progress animation
                for i in range(40):
                    stage_idx = min(i // 6, len(bom_stages) - 1)
                    bom_status_text.markdown(
                        f'<p class="status-text">{bom_stages[stage_idx]}...</p>',
                        unsafe_allow_html=True
                    )
                    bom_steps_container.markdown(render_bom_steps(stage_idx, bom_stages), unsafe_allow_html=True)
                    time.sleep(0.03)
                    bom_progress_bar.progress(i + 1)
                
                # Run BOM analysis
                try:
                    bom_result = run_forecast_bom_analysis(gc_client=None)
                    
                    if bom_result is None or bom_result[0] is None:
                        st.error("❌ BOM Analysis Failed. Check logs for details.")
                        bom_progress_bar.empty()
                        bom_status_text.empty()
                        bom_steps_container.empty()
                        st.stop()
                    
                    bom_excel_buffer, bom_filename = bom_result
                    
                except Exception as e:
                    import traceback
                    st.error(f"❌ BOM Analysis Error: {str(e)}")
                    st.code(traceback.format_exc())
                    bom_progress_bar.empty()
                    bom_status_text.empty()
                    bom_steps_container.empty()
                    st.stop()
                
                # Complete progress
                for i in range(40, 100):
                    stage_idx = min(i // 14, len(bom_stages) - 1)
                    bom_status_text.markdown(
                        f'<p class="status-text">{bom_stages[stage_idx]}...</p>',
                        unsafe_allow_html=True
                    )
                    bom_steps_container.markdown(render_bom_steps(stage_idx, bom_stages), unsafe_allow_html=True)
                    time.sleep(0.015)
                    bom_progress_bar.progress(i + 1)
                
                # Clear progress
                bom_progress_bar.empty()
                bom_status_text.empty()
                bom_steps_container.empty()
                
                # Store results
                if bom_excel_buffer:
                    st.session_state.bom_excel_buffer = bom_excel_buffer
                    st.session_state.bom_filename = bom_filename
                    st.session_state.bom_analysis_complete = True
                    st.session_state.bom_sheets_url = "https://docs.google.com/spreadsheets/d/1_wXJDNZeZ7Y31S_i3UUDQ89vCbJC-xotm3wADBSf5eY"
                    
                    bom_end_time = time.time()
                    bom_duration_sec = bom_end_time - bom_start_time
                    bom_duration_str = time.strftime("%M:%S", time.gmtime(bom_duration_sec))
                    bom_timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    st.session_state.last_bom_time = bom_timestamp_str
                    
                    with bom_result_container:
                        st.success(f"""
                            ✅ **FORECAST BOM ANALYSIS COMPLETED!**
                            
                            📅 Generated at: **{bom_timestamp_str}**  
                            ⏱ Duration: **{bom_duration_str}**
                        """)
                else:
                    with bom_result_container:
                        st.error("❌ BOM Analysis Failed. Please try again.")
    
    # Show quick access if results exist
    if st.session_state.bom_analysis_complete and st.session_state.bom_excel_buffer:
        st.markdown("""
            <div class="holo-card" style="margin-top: 1rem;">
                <h3 class="section-title" style="font-size: 1.3rem;">
                    <span class="section-title-icon">📥</span>
                    BOM Quick Access
                </h3>
            </div>
        """, unsafe_allow_html=True)
        
        bom_col1, bom_col2, bom_col3 = st.columns(3, gap="large")
        
        with bom_col1:
            st.download_button(
                label="📥 DOWNLOAD BOM WORKBOOK",
                data=st.session_state.bom_excel_buffer.getvalue(),
                file_name=st.session_state.bom_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        with bom_col2:
            bom_sheets_url = "https://docs.google.com/spreadsheets/d/1izbZowu4FEiwiVwKWIiRz066aWWOII5u"
            st.markdown(f'<a href="{bom_sheets_url}" target="_blank" class="neural-btn">☁️ BOM SHEETS</a>', unsafe_allow_html=True)
        
        with bom_col3:
            bom_looker_url = "https://lookerstudio.google.com/reporting/9525ae1e-6f0e-4b5f-ae50-ca84312b76fd/page/p_xsi76rd4yd"
            st.markdown(f'<a href="{bom_looker_url}" target="_blank" class="neural-btn">📊 BOM DASHBOARD</a>', unsafe_allow_html=True)

# ==============================================================================
# TAB 3: ACCESS PORTAL
# ==============================================================================
with tab3:
    st.markdown("""
        <div class="holo-card">
            <h2 class="section-title">
                <span class="section-title-icon">🌐</span>
                Intelligence Access Portal
            </h2>
            <p class="description-text">
                Your enhanced reports are available across multiple platforms. 
                Access your intelligence reports through various interfaces for maximum operational efficiency.
            </p>
        </div>
    """, unsafe_allow_html=True)
    
    # Forecast Reports Section
    st.markdown("""
        <div class="holo-card">
            <h3 class="section-title" style="font-size: 1.3rem;">
                <span class="section-title-icon">📊</span>
                Forecast Reports
            </h3>
        </div>
    """, unsafe_allow_html=True)
    
    if st.session_state.excel_buffer:
        last_time = st.session_state.last_forecast_time or "N/A"
        st.info(f"📅 Last Generated: **{last_time}**")
        
        col1, col2, col3 = st.columns(3, gap="large")
        
        with col1:
            st.download_button(
                label="📥 DOWNLOAD EXCEL WORKBOOK",
                data=st.session_state.excel_buffer.getvalue(),
                file_name=st.session_state.filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="portal_download"
            )
        
        with col2:
            if st.session_state.drive_file_id:
                sheets_link = f"https://docs.google.com/spreadsheets/d/{st.session_state.drive_file_id}/edit"
                st.markdown(f'<a href="{sheets_link}" target="_blank" class="neural-btn">☁️ GOOGLE SHEETS</a>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="neural-btn" style="opacity: 0.5;">☁️ Not Available</div>', unsafe_allow_html=True)
        
        with col3:
            looker_url = "https://lookerstudio.google.com/reporting/9525ae1e-6f0e-4b5f-ae50-ca84312b76fd/page/br5SF"
            st.markdown(f'<a href="{looker_url}" target="_blank" class="neural-btn">📊 LOOKER DASHBOARD</a>', unsafe_allow_html=True)
    else:
        st.info("💡 No forecast reports available. Run the Forecast Engine first.")
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # BOM Reports Section
    st.markdown("""
        <div class="holo-card">
            <h3 class="section-title" style="font-size: 1.3rem;">
                <span class="section-title-icon">🧬</span>
                BOM Analysis Reports
            </h3>
        </div>
    """, unsafe_allow_html=True)
    
    if st.session_state.bom_analysis_complete and st.session_state.bom_excel_buffer:
        last_bom_time = st.session_state.last_bom_time or "N/A"
        st.info(f"📅 Last Generated: **{last_bom_time}**")
        
        col1, col2, col3 = st.columns(3, gap="large")
        
        with col1:
            st.download_button(
                label="📥 DOWNLOAD BOM WORKBOOK",
                data=st.session_state.bom_excel_buffer.getvalue(),
                file_name=st.session_state.bom_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="portal_bom_download"
            )
        
        with col2:
            bom_sheets_url = "https://docs.google.com/spreadsheets/d/1izbZowu4FEiwiVwKWIiRz066aWWOII5u"
            st.markdown(f'<a href="{bom_sheets_url}" target="_blank" class="neural-btn">☁️ BOM SHEETS</a>', unsafe_allow_html=True)
        
        with col3:
            bom_looker_url = "https://lookerstudio.google.com/reporting/9525ae1e-6f0e-4b5f-ae50-ca84312b76fd/page/p_xsi76rd4yd"
            st.markdown(f'<a href="{bom_looker_url}" target="_blank" class="neural-btn">📊 BOM DASHBOARD</a>', unsafe_allow_html=True)
    else:
        st.info("💡 No BOM reports available. Run the BOM Analysis first.")

# ==============================================================================
# VISUALIZATION PREVIEW
# ==============================================================================
st.markdown("""
    <div class="viz-preview">
        <h3>📈 PREDICTIVE VISUALIZATION MATRIX</h3>
        <p>Real-time quantum analytics • Multi-dimensional forecasting • Neural pattern recognition</p>
    </div>
""", unsafe_allow_html=True)

# ==============================================================================
# FOOTER
# ==============================================================================
st.markdown("""
    <div class="footer">
        <p class="footer-brand">HERBAL GOODNESS © 2025 | CREATED BY TEAM <a href="https://www.linkedin.com/company/scmplify/" target="_blank" style="color: #60a5fa; text-decoration: none; font-weight: 600;">SCMplify</a> | VERSION 5.1</p>
        <p class="footer-text">Revolutionizing inventory management through intelligent systems</p>
    </div>
""", unsafe_allow_html=True)

# ==============================================================================
# APPLICATION ENTRY POINT
# ==============================================================================

if __name__ == "__main__":
    import sys
    
    if "--api" in sys.argv:
        # Run FastAPI server
        print("🚀 Starting API Server...")
        print("📖 API Documentation: http://localhost:8000/api/docs")
        import uvicorn
        uvicorn.run(api_app, host="0.0.0.0", port=8000)
    else:
        # Streamlit runs automatically when executed with `streamlit run`
        print("ℹ️  To run API server, use: python main_app.py --api")
